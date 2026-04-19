import json
import random
import base64
import io
import os
import secrets
import urllib.request
import urllib.parse
import urllib.error as _urllib_err
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo

# Fuseau horaire local (Corse / métropole France)
_LOCAL_TZ = ZoneInfo("Europe/Paris")

def to_local(dt: datetime) -> datetime:
    """Convertit un datetime UTC naïf en heure locale Europe/Paris."""
    if dt is None:
        return dt
    return dt.replace(tzinfo=timezone.utc).astimezone(_LOCAL_TZ)
from dotenv import load_dotenv
load_dotenv(override=True)
from typing import Optional, List
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, Response
from sqlalchemy.orm import Session
from sqlalchemy import func
from pydantic import BaseModel

# ── Sessions en mémoire ────────────────────────────────────────────────────
_sessions: dict = {}       # token → {role, expires}
_pin_failures: dict = {}   # ip → {count, locked_until}

from database import get_db, engine, SessionLocal
from models import (
    Base, Supplier, Product, Cocktail, CocktailIngredient,
    CashpadMapping, ImportLog, StockHistory, InventorySession, ProductSupplier,
    SupplierOrder, SupplierOrderItem, Event, EventRequirement, ManualLoss, AppSetting, ServiceAlert,
    Staff, StockSnapshot, DeliveryCheck, DeliveryCheckItem, PriceHistory,
    OrderTemplate, OrderTemplateItem
)

Base.metadata.create_all(bind=engine)

# Migrations colonnes manquantes
from sqlalchemy import text as _text
with engine.connect() as _conn:
    # details_json sur imports_log
    try:
        _conn.execute(_text("ALTER TABLE imports_log ADD COLUMN details_json TEXT DEFAULT '[]'"))
        _conn.commit()
    except Exception:
        pass
    # email sur suppliers
    try:
        _conn.execute(_text("ALTER TABLE suppliers ADD COLUMN email TEXT DEFAULT ''"))
        _conn.commit()
    except Exception:
        pass
    # barcode sur products
    try:
        _conn.execute(_text("ALTER TABLE products ADD COLUMN barcode TEXT DEFAULT ''"))
        _conn.commit()
    except Exception:
        pass
    # archived sur products
    try:
        _conn.execute(_text("ALTER TABLE products ADD COLUMN archived INTEGER DEFAULT 0"))
        _conn.commit()
    except Exception:
        pass
    # unit_price_ht sur delivery_check_items (capture prix depuis BL)
    try:
        _conn.execute(_text("ALTER TABLE delivery_check_items ADD COLUMN unit_price_ht REAL"))
        _conn.commit()
    except Exception:
        pass
    # stock_applied sur delivery_check_items (anti-doublon partiel)
    try:
        _conn.execute(_text("ALTER TABLE delivery_check_items ADD COLUMN stock_applied INTEGER DEFAULT 0"))
        _conn.commit()
    except Exception:
        pass
    # bl_photo + bl_photo_type sur delivery_checks
    try:
        _conn.execute(_text("ALTER TABLE delivery_checks ADD COLUMN bl_photo TEXT DEFAULT ''"))
        _conn.commit()
    except Exception:
        pass
    try:
        _conn.execute(_text("ALTER TABLE delivery_checks ADD COLUMN bl_photo_type TEXT DEFAULT ''"))
        _conn.commit()
    except Exception:
        pass
    # skipped_count sur delivery_checks (validation BL sans comptage serveur)
    try:
        _conn.execute(_text("ALTER TABLE delivery_checks ADD COLUMN skipped_count INTEGER DEFAULT 0"))
        _conn.commit()
    except Exception:
        pass
    # vat_rate sur products (0.20 alcool / 0.10 softs)
    try:
        _conn.execute(_text("ALTER TABLE products ADD COLUMN vat_rate REAL DEFAULT 0.20"))
        _conn.commit()
        # Initialiser les taux selon la catégorie pour les produits existants
        _CAT_VAT_10 = ["Eaux", "Sodas", "Cocktails SA"]
        _conn.execute(_text(
            "UPDATE products SET vat_rate = 0.10 WHERE category IN ('Eaux','Sodas','Cocktails SA')"
        ))
        _conn.execute(_text(
            "UPDATE products SET vat_rate = 0.20 WHERE category NOT IN ('Eaux','Sodas','Cocktails SA')"
        ))
        _conn.commit()
    except Exception:
        pass
    # end_date / start_time / end_time sur events
    for _col, _ddl in [
        ("end_date",   "ALTER TABLE events ADD COLUMN end_date DATETIME"),
        ("start_time", "ALTER TABLE events ADD COLUMN start_time TEXT DEFAULT ''"),
        ("end_time",   "ALTER TABLE events ADD COLUMN end_time TEXT DEFAULT ''"),
    ]:
        try:
            _conn.execute(_text(_ddl))
            _conn.commit()
        except Exception:
            pass

    # app_settings table
    try:
        _conn.execute(_text("SELECT 1 FROM app_settings LIMIT 1"))
    except Exception:
        _conn.execute(_text(
            "CREATE TABLE IF NOT EXISTS app_settings ("
            "key TEXT PRIMARY KEY, "
            "value TEXT DEFAULT '', "
            "updated_at DATETIME DEFAULT CURRENT_TIMESTAMP)"
        ))
        _conn.commit()

    # events table
    try:
        _conn.execute(_text("SELECT 1 FROM events LIMIT 1"))
    except Exception:
        _conn.execute(_text(
            "CREATE TABLE IF NOT EXISTS events ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT, "
            "name TEXT NOT NULL, "
            "event_type TEXT DEFAULT 'Autre', "
            "date DATETIME NOT NULL, "
            "notes TEXT DEFAULT '', "
            "created_at DATETIME DEFAULT CURRENT_TIMESTAMP)"
        ))
        _conn.commit()

    # manual_losses table
    try:
        _conn.execute(_text("SELECT 1 FROM manual_losses LIMIT 1"))
    except Exception:
        _conn.execute(_text(
            "CREATE TABLE IF NOT EXISTS manual_losses ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT, "
            "product_id INTEGER NOT NULL, "
            "quantity REAL NOT NULL, "
            "reason TEXT DEFAULT 'Autre', "
            "notes TEXT DEFAULT '', "
            "date DATETIME DEFAULT CURRENT_TIMESTAMP, "
            "staff_name TEXT DEFAULT '', "
            "stock_updated INTEGER DEFAULT 1)"
        ))
        _conn.commit()

    # Migration données : product_suppliers depuis supplier_id + purchase_price existants
    try:
        rows = _conn.execute(_text(
            "SELECT id, supplier_id, purchase_price FROM products WHERE supplier_id IS NOT NULL"
        )).mappings().all()
        for row in rows:
            exists = _conn.execute(_text(
                "SELECT 1 FROM product_suppliers WHERE product_id=:pid AND supplier_id=:sid"
            ), {"pid": row["id"], "sid": row["supplier_id"]}).first()
            if not exists:
                _conn.execute(_text(
                    "INSERT INTO product_suppliers "
                    "(product_id, supplier_id, purchase_price, is_primary, updated_at) "
                    "VALUES (:pid, :sid, :price, 1, CURRENT_TIMESTAMP)"
                ), {"pid": row["id"], "sid": row["supplier_id"], "price": row["purchase_price"]})
        _conn.commit()
    except Exception:
        pass  # table pas encore créée au 1er démarrage → create_all s'en charge

app = FastAPI(title="Marina di Lava — Gestion Stock")
app.mount("/static", StaticFiles(directory="static"), name="static")


# ── Middleware auth ────────────────────────────────────────────────────────
OPEN_PATHS = {"/api/auth", "/api/auth/service", "/api/staff/public"}

@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    path = request.url.path
    # Laisser passer : pages HTML, assets statiques, endpoints d'auth
    if not path.startswith("/api/") or path in OPEN_PATHS or path.startswith("/api/auth/avatar/"):
        return await call_next(request)
    # Vérifier le token
    auth = request.headers.get("Authorization", "")
    token = auth[7:] if auth.startswith("Bearer ") else ""
    session = _sessions.get(token)
    if not session:
        return JSONResponse({"detail": "Non authentifié"}, status_code=401)
    if datetime.utcnow() > session["expires"]:
        _sessions.pop(token, None)
        return JSONResponse({"detail": "Session expirée"}, status_code=401)
    # Renouvelle la session à chaque appel (inactivité 30 min)
    session["expires"] = datetime.utcnow() + timedelta(minutes=30)
    return await call_next(request)


# ── helpers ────────────────────────────────────────────────────────────────

def calc_product(p: Product) -> dict:
    cout_unitaire = None
    marge = None
    marge_color = "gray"
    valeur_stock = None
    vat_rate = p.vat_rate if p.vat_rate is not None else 0.20

    if p.purchase_price is not None:
        cout_unitaire = p.purchase_price  # déjà en prix/unité individuelle
        valeur_stock = p.stock * cout_unitaire

    if cout_unitaire is not None and p.sale_price_ttc:
        prix_ht = p.sale_price_ttc / (1 + vat_rate)
        if prix_ht > 0:
            marge = (prix_ht - cout_unitaire) / prix_ht * 100
            if marge >= 70:
                marge_color = "green"
            elif marge >= 50:
                marge_color = "orange"
            else:
                marge_color = "red"

    if p.is_estimated:
        marge_color = "gray"

    return {
        "id": p.id,
        "nom": p.name,
        "name": p.name,
        "categorie": p.category,
        "category": p.category,
        "fournisseur_id": p.supplier_id,
        "supplier_id": p.supplier_id,
        "fournisseur_nom": p.supplier_rel.name if p.supplier_rel else "",
        "supplier_name": p.supplier_rel.name if p.supplier_rel else "",
        "stock": p.stock,
        "unite": p.unit,
        "unit": p.unit,
        "qte_conditionnement": p.qty_per_pack,
        "qty_per_pack": p.qty_per_pack,
        "volume_cl": p.volume_cl,
        "seuil_alerte": p.alert_threshold,
        "alert_threshold": p.alert_threshold,
        "prix_achat_ht": p.purchase_price,
        "purchase_price": p.purchase_price,
        "prix_vente_ttc": p.sale_price_ttc,
        "sale_price_ttc": p.sale_price_ttc,
        "prix_estime": p.is_estimated,
        "is_estimated": p.is_estimated,
        "barcode": p.barcode or "",
        "archived": bool(p.archived),
        "vat_rate": vat_rate,
        "cout_unitaire": round(cout_unitaire, 4) if cout_unitaire is not None else None,
        "marge": round(marge, 1) if marge is not None else None,
        "marge_color": marge_color,
        "valeur_stock": round(valeur_stock, 2) if valeur_stock is not None else None,
        "suppliers": [
            {
                "supplier_id": ps.supplier_id,
                "supplier_name": ps.supplier.name if ps.supplier else "",
                "purchase_price": ps.purchase_price,
                "is_primary": ps.is_primary,
            }
            for ps in sorted(p.product_suppliers or [], key=lambda x: (not x.is_primary, x.id))
        ],
    }


def calc_cocktail(c: Cocktail) -> dict:
    cout_matiere = 0.0
    ingredients_detail = []
    for ing in c.ingredients:
        p = ing.product
        if p and p.purchase_price is not None and p.volume_cl:
            cout_cl = p.purchase_price / p.volume_cl  # purchase_price déjà en prix/unité
            cost = cout_cl * ing.dose_cl
            cout_matiere += cost
            ingredients_detail.append({
                "product_id": p.id,
                "produit_id": p.id,
                "product_name": p.name,
                "produit_nom": p.name,
                "dose_cl": ing.dose_cl,
                "cost": round(cost, 4),
            })
        else:
            ingredients_detail.append({
                "product_id": ing.product_id,
                "produit_id": ing.product_id,
                "product_name": p.name if p else "?",
                "produit_nom": p.name if p else "?",
                "dose_cl": ing.dose_cl,
                "cost": None,
            })

    # TVA du cocktail : 20 % si au moins un ingrédient est à 20 % (alcool), sinon 10 %
    cocktail_vat = 0.10
    for ing in c.ingredients:
        p = ing.product
        if p and (p.vat_rate is not None) and p.vat_rate >= 0.20:
            cocktail_vat = 0.20
            break

    marge = None
    marge_color = "gray"
    prix_vente_ht = None
    if c.sale_price_ttc:
        prix_vente_ht = c.sale_price_ttc / (1 + cocktail_vat)
        if prix_vente_ht > 0 and cout_matiere is not None:
            marge = (prix_vente_ht - cout_matiere) / prix_vente_ht * 100
            if marge >= 70:
                marge_color = "green"
            elif marge >= 50:
                marge_color = "orange"
            else:
                marge_color = "red"

    return {
        "id": c.id,
        "nom": c.name,
        "name": c.name,
        "prix_vente_ttc": c.sale_price_ttc,
        "sale_price_ttc": c.sale_price_ttc,
        "prix_vente_ht": round(prix_vente_ht, 4) if prix_vente_ht else None,
        "cout_matiere": round(cout_matiere, 4),
        "marge": round(marge, 1) if marge is not None else None,
        "marge_color": marge_color,
        "vat_rate": cocktail_vat,
        "ingredients": ingredients_detail,
    }


def log_event(db: Session, event_type: str, description: str, data: dict = None):
    h = StockHistory(
        event_type=event_type,
        description=description,
        data_json=json.dumps(data or {}, ensure_ascii=False),
    )
    db.add(h)
    db.flush()  # pour obtenir h.id immédiatement
    return h


def _build_backup_zip() -> tuple:
    """
    Construit un ZIP des tables importantes en CSV.
    Retourne (bytes, filename).
    """
    import csv
    import zipfile
    from io import StringIO, BytesIO

    db = SessionLocal()
    try:
        def _csv(header, rows):
            buf = StringIO()
            w = csv.writer(buf, delimiter=";")
            w.writerow(header)
            w.writerows(rows)
            return buf.getvalue().encode("utf-8-sig")

        zip_buf = BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            # Produits
            zf.writestr("products.csv", _csv(
                ["id","name","category","stock","unit","alert_threshold","purchase_price",
                 "sale_price_ttc","vat_rate","supplier_id","archived","is_estimated"],
                [[p.id, p.name, p.category, p.stock, p.unit, p.alert_threshold,
                  p.purchase_price, p.sale_price_ttc, p.vat_rate, p.supplier_id,
                  bool(p.archived), bool(p.is_estimated)]
                 for p in db.query(Product).all()]
            ))
            # Fournisseurs
            zf.writestr("suppliers.csv", _csv(
                ["id","name","contact","phone","email","categories"],
                [[s.id, s.name, s.contact, s.phone, s.email, s.categories]
                 for s in db.query(Supplier).all()]
            ))
            # Cocktails
            zf.writestr("cocktails.csv", _csv(
                ["id","name","sale_price_ttc"],
                [[c.id, c.name, c.sale_price_ttc] for c in db.query(Cocktail).all()]
            ))
            zf.writestr("cocktail_ingredients.csv", _csv(
                ["cocktail_id","product_id","dose_cl"],
                [[i.cocktail_id, i.product_id, i.dose_cl] for i in db.query(CocktailIngredient).all()]
            ))
            # Historique stock
            zf.writestr("stock_history.csv", _csv(
                ["id","event_type","description","data_json","created_at"],
                [[h.id, h.event_type, h.description, h.data_json,
                  h.created_at.isoformat() if h.created_at else ""]
                 for h in db.query(StockHistory).all()]
            ))
            # Commandes fournisseur
            zf.writestr("supplier_orders.csv", _csv(
                ["id","reference","supplier_id","status","notes","created_at","sent_at","received_at"],
                [[o.id, o.reference, o.supplier_id, o.status, o.notes,
                  o.created_at.isoformat() if o.created_at else "",
                  o.sent_at.isoformat() if o.sent_at else "",
                  o.received_at.isoformat() if o.received_at else ""]
                 for o in db.query(SupplierOrder).all()]
            ))
            zf.writestr("supplier_order_items.csv", _csv(
                ["order_id","product_id","product_name","qty_ordered","unit_price_ht"],
                [[i.order_id, i.product_id, i.product_name, i.qty_ordered, i.unit_price_ht]
                 for i in db.query(SupplierOrderItem).all()]
            ))
            # Événements
            zf.writestr("events.csv", _csv(
                ["id","name","event_type","date","end_date","start_time","end_time","notes"],
                [[e.id, e.name, e.event_type,
                  e.date.isoformat() if e.date else "",
                  e.end_date.isoformat() if e.end_date else "",
                  e.start_time, e.end_time, e.notes]
                 for e in db.query(Event).all()]
            ))
            # Contrôles livraison
            zf.writestr("delivery_checks.csv", _csv(
                ["id","supplier_id","order_id","status","checked_by","validated_by",
                 "bl_reference","created_at","validated_at"],
                [[c.id, c.supplier_id, c.order_id, c.status, c.checked_by,
                  c.validated_by, c.bl_reference,
                  c.created_at.isoformat() if c.created_at else "",
                  c.validated_at.isoformat() if c.validated_at else ""]
                 for c in db.query(DeliveryCheck).all()]
            ))
            # Historique prix
            zf.writestr("price_history.csv", _csv(
                ["product_id","old_price","new_price","source","supplier_id","reference","changed_at"],
                [[h.product_id, h.old_price, h.new_price, h.source, h.supplier_id, h.reference,
                  h.changed_at.isoformat() if h.changed_at else ""]
                 for h in db.query(PriceHistory).all()]
            ))
            # Pertes manuelles
            zf.writestr("manual_losses.csv", _csv(
                ["product_id","quantity","reason","notes","date","staff_name"],
                [[l.product_id, l.quantity, l.reason, l.notes,
                  l.date.isoformat() if l.date else "",
                  l.staff_name]
                 for l in db.query(ManualLoss).all()]
            ))

        data = zip_buf.getvalue()
        ts = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
        return data, f"marina-backup-{ts}.zip"
    finally:
        db.close()


def _send_mailjet_with_attachment(subject: str, html: str, to_emails: list,
                                   attachment_bytes: bytes, attachment_name: str,
                                   attachment_mime: str = "application/zip") -> bool:
    """Envoie un email via Mailjet avec un fichier joint."""
    import urllib.request as _urllib
    mailjet_key    = os.getenv("MAILJET_API_KEY", "").strip()
    mailjet_secret = os.getenv("MAILJET_SECRET_KEY", "").strip()
    if not mailjet_key or not mailjet_secret or not to_emails:
        return False
    from_addr = os.getenv("FROM_EMAIL", "marinadilava.commandes@gmail.com").strip()
    creds = base64.b64encode(f"{mailjet_key}:{mailjet_secret}".encode()).decode()
    b64 = base64.standard_b64encode(attachment_bytes).decode()
    payload = json.dumps({
        "Messages": [{
            "From":     {"Email": from_addr, "Name": "Marina di Lava Backup"},
            "To":       [{"Email": e.strip()} for e in to_emails if e.strip()],
            "Subject":  subject,
            "HTMLPart": html,
            "Attachments": [{
                "ContentType":   attachment_mime,
                "Filename":      attachment_name,
                "Base64Content": b64,
            }],
        }]
    }).encode("utf-8")
    try:
        req = _urllib.Request(
            "https://api.mailjet.com/v3.1/send",
            data=payload,
            headers={"Authorization": f"Basic {creds}", "Content-Type": "application/json"},
            method="POST",
        )
        with _urllib.urlopen(req, timeout=30) as resp:
            return 200 <= resp.status < 300
    except Exception as e:
        print(f"[Backup email] ❌ {e}")
        return False


def _run_weekly_backup():
    """Job scheduler : backup + envoi email aux destinataires configurés."""
    db = SessionLocal()
    try:
        recipients_raw = _get_setting(db, "backup_emails", "").strip()
        recipients = [e.strip() for e in recipients_raw.split(",") if e.strip()]
        if not recipients:
            print("[Backup] skippé — aucun destinataire configuré")
            return
    finally:
        db.close()

    try:
        data, filename = _build_backup_zip()
    except Exception as e:
        print(f"[Backup] échec construction ZIP : {e}")
        return

    subject = f"📦 Backup Marina di Lava — {datetime.utcnow().strftime('%d/%m/%Y')}"
    html = f"""<div style="font-family:Arial,sans-serif;max-width:520px;margin:0 auto;padding:20px">
      <h2 style="color:#14532D">📦 Backup hebdomadaire</h2>
      <p>Archive complète des données Marina di Lava ci-jointe.</p>
      <p><strong>Fichier</strong> : {filename}<br>
         <strong>Taille</strong> : {len(data)/1024:.1f} KB</p>
      <p style="color:#666;font-size:12px;margin-top:20px">Conservez ce mail — c'est votre parachute en cas de crash serveur.</p>
    </div>"""
    ok = _send_mailjet_with_attachment(subject, html, recipients, data, filename)
    print(f"[Backup] {'✓ envoyé' if ok else '❌ échec envoi'} à {len(recipients)} destinataire(s)")


def _send_mailjet_email(subject: str, html_body: str, to_emails: list, from_name: str = "Marina di Lava Alertes") -> bool:
    """
    Envoie un email via Mailjet. Retourne True si ok, False sinon.
    Ne lève pas d'exception — les alertes ne doivent jamais bloquer le flux principal.
    """
    import urllib.request as _urllib
    mailjet_key    = os.getenv("MAILJET_API_KEY", "").strip()
    mailjet_secret = os.getenv("MAILJET_SECRET_KEY", "").strip()
    if not mailjet_key or not mailjet_secret or not to_emails:
        return False
    from_addr = os.getenv("FROM_EMAIL", "marinadilava.commandes@gmail.com").strip()
    creds = base64.b64encode(f"{mailjet_key}:{mailjet_secret}".encode()).decode()
    recipients = [{"Email": e.strip()} for e in to_emails if e.strip()]
    if not recipients:
        return False
    payload = json.dumps({
        "Messages": [{
            "From":     {"Email": from_addr, "Name": from_name},
            "To":       recipients,
            "Subject":  subject,
            "HTMLPart": html_body,
        }]
    }).encode("utf-8")
    try:
        req = _urllib.Request(
            "https://api.mailjet.com/v3.1/send",
            data=payload,
            headers={"Authorization": f"Basic {creds}", "Content-Type": "application/json"},
            method="POST",
        )
        with _urllib.urlopen(req, timeout=10) as resp:
            return 200 <= resp.status < 300
    except Exception as e:
        print(f"[Alert email] ❌ {e}")
        return False


def _trigger_rupture_alerts(db: Session, context: str = ""):
    """
    Scanne les produits actifs. Pour ceux qui viennent de passer sous le seuil,
    envoie un email (cool-down 6h pour éviter le spam).
    Appelé après les opérations qui baissent le stock (Cashpad, sorties).
    """
    recipients_raw = _get_setting(db, "alert_emails", "").strip()
    if not recipients_raw:
        return   # pas configuré → skip
    recipients = [e.strip() for e in recipients_raw.split(",") if e.strip()]
    if not recipients:
        return

    threshold_mode = _get_setting(db, "alert_mode", "rupture")   # "rupture" | "low_stock"
    cooldown_hours = 6
    now = datetime.utcnow()

    products = db.query(Product).filter(
        (Product.archived == False) | (Product.archived == None)
    ).all()

    to_alert = []
    for p in products:
        is_rupture = (p.stock or 0) <= 0
        is_low     = not is_rupture and (p.stock or 0) <= (p.alert_threshold or 0)
        triggered  = (is_rupture or (threshold_mode == "low_stock" and is_low))
        if not triggered:
            continue

        key = f"last_alert_{p.id}"
        last_raw = _get_setting(db, key, "")
        if last_raw:
            try:
                last_dt = datetime.fromisoformat(last_raw)
                if (now - last_dt).total_seconds() < cooldown_hours * 3600:
                    continue   # cooldown pas encore écoulé
            except Exception:
                pass

        to_alert.append({
            "id": p.id, "name": p.name, "stock": p.stock or 0,
            "threshold": p.alert_threshold or 0, "unit": p.unit or "u",
            "is_rupture": is_rupture,
        })
        _set_setting(db, key, now.isoformat())

    if not to_alert:
        return

    # Construire l'email
    rupture_rows = "".join(
        f'<tr><td style="padding:8px;border-bottom:1px solid #eee"><strong>{a["name"]}</strong></td>'
        f'<td style="padding:8px;border-bottom:1px solid #eee;color:#DC2626;font-weight:700">RUPTURE</td>'
        f'<td style="padding:8px;border-bottom:1px solid #eee">{a["unit"]}</td></tr>'
        for a in to_alert if a["is_rupture"]
    )
    low_rows = "".join(
        f'<tr><td style="padding:8px;border-bottom:1px solid #eee"><strong>{a["name"]}</strong></td>'
        f'<td style="padding:8px;border-bottom:1px solid #eee;color:#F59E0B;font-weight:700">{a["stock"]:.1f} {a["unit"]} (seuil {a["threshold"]})</td>'
        f'<td style="padding:8px;border-bottom:1px solid #eee">Stock bas</td></tr>'
        for a in to_alert if not a["is_rupture"]
    )
    n_rup = sum(1 for a in to_alert if a["is_rupture"])
    n_low = sum(1 for a in to_alert if not a["is_rupture"])
    subject = f"🔴 Marina di Lava — {n_rup} rupture(s)" + (f" + {n_low} stock bas" if n_low else "")
    html = f"""<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:20px;background:#f7f7f7">
        <div style="background:linear-gradient(135deg,#0D2818,#14532D);color:#fff;padding:18px 22px;border-radius:10px 10px 0 0">
          <h2 style="margin:0;font-size:20px;font-weight:700">🔴 Alerte stock — Marina di Lava</h2>
          <p style="margin:4px 0 0;opacity:0.85;font-size:13px">{context or 'Suite à un mouvement de stock'}</p>
        </div>
        <div style="background:#fff;padding:20px;border-radius:0 0 10px 10px">
          <p style="margin:0 0 12px">Les produits suivants nécessitent une action :</p>
          <table style="width:100%;border-collapse:collapse;font-size:14px">
            <thead><tr style="background:#f3f4f6"><th style="padding:8px;text-align:left">Produit</th><th style="padding:8px;text-align:left">Niveau</th><th style="padding:8px;text-align:left">Action</th></tr></thead>
            <tbody>{rupture_rows}{low_rows}</tbody>
          </table>
          <p style="margin:16px 0 0;font-size:12px;color:#666">Cet email est envoyé automatiquement. Pour désactiver, videz la liste des destinataires dans les paramètres de l'application.</p>
        </div></div>"""
    _send_mailjet_email(subject, html, recipients, from_name="Marina di Lava Alertes")
    db.commit()


class BackupSettingsIn(BaseModel):
    emails: str


@app.get("/api/backup-settings")
def get_backup_settings(db: Session = Depends(get_db)):
    return {"emails": _get_setting(db, "backup_emails", "")}


@app.post("/api/backup-settings")
def set_backup_settings(body: BackupSettingsIn, db: Session = Depends(get_db)):
    _set_setting(db, "backup_emails", (body.emails or "").strip())
    db.commit()
    return {"ok": True}


@app.post("/api/backup/run")
def run_backup_now(db: Session = Depends(get_db)):
    """Lance un backup + envoi email (manuel)."""
    recipients_raw = _get_setting(db, "backup_emails", "").strip()
    recipients = [e.strip() for e in recipients_raw.split(",") if e.strip()]
    if not recipients:
        raise HTTPException(400, "Aucun destinataire configuré dans les paramètres.")
    try:
        data, filename = _build_backup_zip()
    except Exception as e:
        raise HTTPException(500, f"Erreur construction backup : {e}")
    html = f"""<div style="font-family:Arial,sans-serif;max-width:520px;margin:0 auto;padding:20px">
      <h2 style="color:#14532D">📦 Backup manuel Marina di Lava</h2>
      <p>Archive ci-jointe.</p>
      <p><strong>Fichier</strong> : {filename}<br><strong>Taille</strong> : {len(data)/1024:.1f} KB</p>
    </div>"""
    subject = f"📦 Backup Marina di Lava — {datetime.utcnow().strftime('%d/%m/%Y %H:%M')}"
    ok = _send_mailjet_with_attachment(subject, html, recipients, data, filename)
    if not ok:
        raise HTTPException(500, "Échec envoi — vérifiez MAILJET_API_KEY/SECRET.")
    return {"ok": True, "size_kb": round(len(data)/1024, 1), "sent_to": recipients}


@app.get("/api/backup/download")
def download_backup():
    """Télécharge directement le ZIP de backup (sans email)."""
    from fastapi.responses import Response
    try:
        data, filename = _build_backup_zip()
    except Exception as e:
        raise HTTPException(500, f"Erreur : {e}")
    return Response(
        content=data, media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


class AlertSettingsIn(BaseModel):
    emails: str                         # liste CSV
    mode: str = "rupture"               # rupture | low_stock


@app.get("/api/alert-settings")
def get_alert_settings(db: Session = Depends(get_db)):
    return {
        "emails": _get_setting(db, "alert_emails", ""),
        "mode":   _get_setting(db, "alert_mode", "rupture"),
    }


@app.post("/api/alert-settings")
def set_alert_settings(body: AlertSettingsIn, db: Session = Depends(get_db)):
    _set_setting(db, "alert_emails", (body.emails or "").strip())
    _set_setting(db, "alert_mode", body.mode if body.mode in ("rupture", "low_stock") else "rupture")
    db.commit()
    return {"ok": True}


@app.post("/api/alert-settings/test")
def test_alert_email(db: Session = Depends(get_db)):
    """Envoie un email de test aux destinataires configurés."""
    recipients_raw = _get_setting(db, "alert_emails", "").strip()
    recipients = [e.strip() for e in recipients_raw.split(",") if e.strip()]
    if not recipients:
        raise HTTPException(400, "Aucun destinataire configuré")
    html = """<div style="font-family:Arial,sans-serif;max-width:520px;margin:0 auto;padding:20px">
      <h2 style="color:#14532D">✓ Email de test — Marina di Lava</h2>
      <p>Cet email confirme que les alertes de rupture sont bien configurées.</p>
      <p style="color:#666;font-size:12px">Tu recevras automatiquement un email à chaque rupture ou stock bas détecté.</p>
    </div>"""
    ok = _send_mailjet_email("Test alertes Marina di Lava", html, recipients)
    if not ok:
        raise HTTPException(500, "Échec de l'envoi — vérifiez MAILJET_API_KEY et MAILJET_SECRET_KEY sur Railway.")
    return {"ok": True, "sent_to": recipients}


def record_price_change(db: Session, product, new_price: float, source: str = "",
                        supplier_id: Optional[int] = None, reference: str = ""):
    """Enregistre une variation de prix d'achat si différente de l'ancienne."""
    if new_price is None:
        return
    old = product.purchase_price
    if old is not None and abs(float(old) - float(new_price)) < 0.001:
        return   # pas de changement significatif
    ph = PriceHistory(
        product_id=product.id,
        old_price=float(old) if old is not None else None,
        new_price=float(new_price),
        source=source or "",
        supplier_id=supplier_id,
        reference=reference or "",
    )
    db.add(ph)


def extract_sales(h) -> list:
    """
    Retourne la liste des lignes de vente d'une entrée StockHistory (import_cashpad).
    Gère l'ancien et le nouveau format. Chaque ligne contient au moins :
    qty_sold, sale_price_ttc, purchase_price, vat_rate, name/product_name.
    """
    try:
        data = json.loads(h.data_json or "{}")
    except Exception:
        return []
    if isinstance(data, dict):
        return data.get("sales") or []
    if isinstance(data, list):
        return data
    return []


# ── root ───────────────────────────────────────────────────────────────────

@app.get("/")
def root():
    # index.html = jamais caché → on veut que Safari iOS voie immédiatement
    # les nouvelles versions de app.js et style.css référencées par ?v=...
    return FileResponse(
        "static/index.html",
        headers={
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


class PinIn(BaseModel):
    pin: str

class ServiceLoginIn(BaseModel):
    pin: Optional[str] = None
    staff_id: Optional[int] = None


@app.get("/api/staff/public")
def list_staff_public(db: Session = Depends(get_db)):
    """Liste publique (sans PIN) pour l'écran de login service."""
    rows = db.query(Staff).filter(Staff.is_active == True).order_by(Staff.name).all()
    return [{"id": s.id, "name": s.name, "slug": s.slug} for s in rows]


@app.post("/api/auth/service")
def auth_service(body: ServiceLoginIn = None, db: Session = Depends(get_db)):
    """
    Connexion service.
    - Si aucun staff n'est créé : connexion directe sans PIN (mode simple).
    - Sinon : staff_id + PIN requis → retour user_name dans la session.
    """
    staff_exists = db.query(Staff).filter(Staff.is_active == True).count() > 0

    if not staff_exists:
        token = secrets.token_urlsafe(32)
        _sessions[token] = {
            "role": "service",
            "user_name": "",
            "expires": datetime.utcnow() + timedelta(hours=12),
        }
        return {"ok": True, "role": "service", "token": token, "user_name": ""}

    if not body or not body.staff_id or not body.pin:
        raise HTTPException(400, detail="Choisissez votre nom et entrez votre PIN.")

    staff = db.query(Staff).get(body.staff_id)
    if not staff or not staff.is_active or staff.pin != body.pin:
        raise HTTPException(401, detail="PIN incorrect.")

    token = secrets.token_urlsafe(32)
    _sessions[token] = {
        "role": "service",
        "user_name": staff.name,
        "staff_id": staff.id,
        "expires": datetime.utcnow() + timedelta(hours=12),
    }
    return {
        "ok": True, "role": "service", "token": token,
        "user_name": staff.name, "staff_id": staff.id,
    }


# ── Gestion Staff (direction uniquement) ─────────────────────────────────
class StaffIn(BaseModel):
    name: str
    pin: str
    is_active: Optional[bool] = True


def _slugify_staff(name: str) -> str:
    import re, unicodedata
    s = unicodedata.normalize("NFKD", name or "").encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-zA-Z0-9]+", "", s).lower()
    return s or "staff"


def _require_manager_session(request: Request):
    auth = request.headers.get("Authorization", "") if request else ""
    token = auth[7:] if auth.startswith("Bearer ") else ""
    session = _sessions.get(token)
    if not session or session.get("role") != "manager":
        raise HTTPException(401, detail="Accès réservé à la direction.")


@app.get("/api/staff")
def list_staff(request: Request, db: Session = Depends(get_db)):
    _require_manager_session(request)
    rows = db.query(Staff).order_by(Staff.is_active.desc(), Staff.name).all()
    return [
        {
            "id": s.id, "name": s.name, "slug": s.slug, "pin": s.pin,
            "is_active": bool(s.is_active),
            "created_at": s.created_at.strftime("%d/%m/%Y %H:%M") if s.created_at else "",
        }
        for s in rows
    ]


@app.post("/api/staff")
def create_staff(body: StaffIn, request: Request, db: Session = Depends(get_db)):
    _require_manager_session(request)
    name = (body.name or "").strip()
    pin = (body.pin or "").strip()
    if not name or not pin:
        raise HTTPException(400, detail="Nom et PIN requis.")
    if not pin.isdigit() or not (3 <= len(pin) <= 8):
        raise HTTPException(400, detail="PIN : 3 à 8 chiffres.")
    # Générer un slug unique
    base_slug = _slugify_staff(name)
    slug = base_slug
    i = 2
    while db.query(Staff).filter(Staff.slug == slug).first():
        slug = f"{base_slug}{i}"
        i += 1
    s = Staff(name=name, slug=slug, pin=pin, is_active=bool(body.is_active))
    db.add(s)
    db.commit()
    db.refresh(s)
    return {"id": s.id, "name": s.name, "slug": s.slug, "pin": s.pin, "is_active": s.is_active}


@app.put("/api/staff/{sid}")
def update_staff(sid: int, body: StaffIn, request: Request, db: Session = Depends(get_db)):
    _require_manager_session(request)
    s = db.query(Staff).get(sid)
    if not s:
        raise HTTPException(404)
    if body.name is not None and body.name.strip():
        s.name = body.name.strip()
    if body.pin is not None and body.pin.strip():
        pin = body.pin.strip()
        if not pin.isdigit() or not (3 <= len(pin) <= 8):
            raise HTTPException(400, detail="PIN : 3 à 8 chiffres.")
        s.pin = pin
    if body.is_active is not None:
        s.is_active = bool(body.is_active)
    db.commit()
    return {"ok": True}


@app.delete("/api/staff/{sid}")
def delete_staff(sid: int, request: Request, db: Session = Depends(get_db)):
    """Supprime un staff (les logs conservent le nom copié dans staff_name)."""
    _require_manager_session(request)
    s = db.query(Staff).get(sid)
    if not s:
        raise HTTPException(404)
    db.delete(s)
    db.commit()
    return {"ok": True}

@app.post("/api/auth")
def auth_pin(body: PinIn, request: Request, db: Session = Depends(get_db)):
    """Connexion direction — vérifie PIN avec anti-brute-force. Supporte plusieurs utilisateurs."""
    ip = request.client.host if request.client else "unknown"
    failure = _pin_failures.get(ip, {})
    locked_until = failure.get("locked_until")
    if locked_until and datetime.utcnow() < locked_until:
        secs = int((locked_until - datetime.utcnow()).total_seconds())
        raise HTTPException(429, f"Trop de tentatives. Réessayez dans {secs}s")

    # Utilisateurs direction : PIN → profil (configurable via env)
    direction_users = {
        os.environ.get("PIN_JMARC", "0034"): {"name": "J-Marc", "slug": "jmarc"},
        os.environ.get("PIN_LISANDRU", "1143"): {"name": "Lisandru", "slug": "lisandru"},
    }
    # Fallback : ancien MANAGER_PIN (rétrocompatible)
    legacy_pin = os.environ.get("MANAGER_PIN", "")
    if legacy_pin and legacy_pin not in direction_users:
        direction_users[legacy_pin] = {"name": "Direction", "slug": "direction"}

    user = direction_users.get(body.pin)
    if user:
        _pin_failures.pop(ip, None)
        slug = user["slug"]
        # Chercher un avatar en base (persistant), sinon SVG par défaut
        if _get_setting(db, f"avatar_{slug}_data", ""):
            photo = f"/api/auth/avatar/{slug}"
        else:
            photo = f"/static/avatars/{slug}.svg"
        token = secrets.token_urlsafe(32)
        _sessions[token] = {
            "role": "manager",
            "user_name": user["name"],
            "user_photo": photo,
            "expires": datetime.utcnow() + timedelta(minutes=30),
        }
        return {"ok": True, "role": "manager", "token": token,
                "user_name": user["name"], "user_photo": photo}

    count = failure.get("count", 0) + 1
    if count >= 3:
        _pin_failures[ip] = {"count": count, "locked_until": datetime.utcnow() + timedelta(minutes=15)}
        raise HTTPException(429, "3 erreurs — accès bloqué 15 minutes")
    else:
        _pin_failures[ip] = {"count": count}
        remaining = 3 - count
        raise HTTPException(401, f"Code PIN incorrect — {remaining} tentative(s) restante(s)")


def _apply_exif_orientation(content: bytes, content_type: str) -> tuple[bytes, str]:
    """
    Réoriente une image selon ses données EXIF (photos iPhone portrait/paysage).
    Retourne (bytes, mime). Pour PDF ou si Pillow échoue, retourne tel quel.
    """
    if content_type and "pdf" in content_type.lower():
        return content, content_type
    try:
        import io as _io
        from PIL import Image as _Image, ImageOps as _ImageOps
        img = _Image.open(_io.BytesIO(content))
        img = _ImageOps.exif_transpose(img)
        if img.mode in ("RGBA", "P", "LA"):
            img = img.convert("RGB")
        buf = _io.BytesIO()
        img.save(buf, format="JPEG", quality=90, optimize=True)
        return buf.getvalue(), "image/jpeg"
    except Exception:
        return content, content_type or "image/jpeg"


@app.post("/api/auth/avatar")
async def upload_avatar(file: UploadFile = File(...), request: Request = None, db: Session = Depends(get_db)):
    """Upload une photo de profil — stockée en base (persistant sur Railway)."""
    auth = request.headers.get("Authorization", "")
    token = auth[7:] if auth.startswith("Bearer ") else ""
    session = _sessions.get(token)
    if not session or session.get("role") != "manager":
        raise HTTPException(401, detail="Non authentifié")
    user_name = session.get("user_name", "")
    if not user_name:
        raise HTTPException(400, detail="Utilisateur inconnu")

    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(400, detail="Image trop grande (max 5 Mo)")

    slug = user_name.lower().replace("-", "").replace(" ", "")
    ext = file.filename.rsplit(".", 1)[-1].lower() if file.filename and "." in file.filename else "jpg"
    media_map = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png", "webp": "image/webp"}
    media_type = media_map.get(ext, "image/jpeg")

    # Réoriente l'avatar selon EXIF (iPhone) — pas pour les PDF (impossible ici)
    content, media_type = _apply_exif_orientation(content, media_type)

    b64 = base64.standard_b64encode(content).decode("utf-8")
    _set_setting(db, f"avatar_{slug}_data", b64)
    _set_setting(db, f"avatar_{slug}_type", media_type)
    db.commit()

    photo_url = f"/api/auth/avatar/{slug}?t={int(datetime.utcnow().timestamp())}"
    session["user_photo"] = photo_url
    return {"ok": True, "photo": photo_url}


@app.get("/api/auth/avatar/{slug}")
def serve_avatar(slug: str, db: Session = Depends(get_db)):
    """Sert un avatar stocké en base de données."""
    b64 = _get_setting(db, f"avatar_{slug}_data", "")
    if not b64:
        raise HTTPException(404, detail="Avatar non trouvé")
    media_type = _get_setting(db, f"avatar_{slug}_type", "image/jpeg")
    data = base64.standard_b64decode(b64)
    return Response(content=data, media_type=media_type,
                    headers={"Cache-Control": "public, max-age=86400"})


@app.get("/api/auth/me")
def auth_me(request: Request):
    """Retourne le profil de l'utilisateur connecté."""
    auth = request.headers.get("Authorization", "")
    token = auth[7:] if auth.startswith("Bearer ") else ""
    session = _sessions.get(token)
    if not session:
        raise HTTPException(401, detail="Non authentifié")
    return {
        "role": session.get("role", "service"),
        "user_name": session.get("user_name", ""),
        "user_photo": session.get("user_photo", ""),
    }


# ══════════════════════════════════════════════════════════════════════════
# FOURNISSEURS / SUPPLIERS
# ══════════════════════════════════════════════════════════════════════════

class SupplierIn(BaseModel):
    name: str
    contact: str = ""
    phone: str = ""
    email: str = ""
    categories: str = ""


def _supplier_dict(s: Supplier) -> dict:
    return {
        "id": s.id,
        "nom": s.name,
        "name": s.name,
        "contact": s.contact,
        "telephone": s.phone,
        "phone": s.phone,
        "email": s.email or "",
        "categories": s.categories,
    }


@app.get("/api/fournisseurs")
@app.get("/api/suppliers")
def get_suppliers(db: Session = Depends(get_db)):
    return [_supplier_dict(s) for s in db.query(Supplier).all()]


@app.post("/api/fournisseurs")
@app.post("/api/suppliers")
def create_supplier(body: SupplierIn, db: Session = Depends(get_db)):
    s = Supplier(**body.model_dump())
    db.add(s)
    db.commit()
    db.refresh(s)
    return _supplier_dict(s)


@app.put("/api/fournisseurs/{sid}")
@app.put("/api/suppliers/{sid}")
def update_supplier(sid: int, body: SupplierIn, db: Session = Depends(get_db)):
    s = db.query(Supplier).get(sid)
    if not s:
        raise HTTPException(404)
    for k, v in body.model_dump().items():
        setattr(s, k, v)
    db.commit()
    return {"ok": True}


@app.delete("/api/fournisseurs/{sid}")
@app.delete("/api/suppliers/{sid}")
def delete_supplier(sid: int, db: Session = Depends(get_db)):
    s = db.query(Supplier).get(sid)
    if not s:
        raise HTTPException(404)
    db.delete(s)
    db.commit()
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════
# COMMANDES FOURNISSEURS / SUPPLIER ORDERS
# ══════════════════════════════════════════════════════════════════════════

def _order_dict(o: SupplierOrder) -> dict:
    total_ht = sum(
        (it.qty_ordered or 0) * (it.unit_price_ht or 0)
        for it in o.items
    )
    return {
        "id": o.id,
        "reference": o.reference,
        "supplier_id": o.supplier_id,
        "supplier_name": o.supplier.name if o.supplier else "",
        "supplier_email": (o.supplier.email or "") if o.supplier else "",
        "status": o.status,
        "notes": o.notes or "",
        "total_ht": round(total_ht, 2),
        "items_count": len(o.items),
        "created_at": to_local(o.created_at).strftime("%d/%m/%Y %H:%M") if o.created_at else "",
        "sent_at": to_local(o.sent_at).strftime("%d/%m/%Y %H:%M") if o.sent_at else None,
        "received_at": to_local(o.received_at).strftime("%d/%m/%Y %H:%M") if o.received_at else None,
        "items": [
            {
                "id": it.id,
                "product_id": it.product_id,
                "product_name": it.product_name or (it.product.name if it.product else ""),
                "qty_ordered": it.qty_ordered,
                "unit_price_ht": it.unit_price_ht,
                "line_total": round((it.qty_ordered or 0) * (it.unit_price_ht or 0), 2) if it.unit_price_ht else None,
                "current_stock": it.product.stock if it.product else None,
                "alert_threshold": it.product.alert_threshold if it.product else None,
            }
            for it in o.items
        ],
    }


def _gen_order_ref(db: Session) -> str:
    """Génère une référence unique CMD-YYYYMMDD-NNN."""
    today = datetime.utcnow().strftime("%Y%m%d")
    prefix = f"CMD-{today}-"
    existing = db.query(SupplierOrder).filter(
        SupplierOrder.reference.like(f"{prefix}%")
    ).count()
    return f"{prefix}{existing + 1:03d}"


class OrderItemIn(BaseModel):
    product_id: Optional[int] = None
    product_name: str = ""
    qty_ordered: float = 0
    unit_price_ht: Optional[float] = None


class OrderIn(BaseModel):
    supplier_id: int
    notes: str = ""
    items: List[OrderItemIn] = []


@app.get("/api/orders")
def get_orders(db: Session = Depends(get_db)):
    orders = db.query(SupplierOrder).order_by(SupplierOrder.created_at.desc()).all()
    return [_order_dict(o) for o in orders]


@app.get("/api/orders/suggestions/{supplier_id}")
def get_order_suggestions(supplier_id: int, db: Session = Depends(get_db)):
    """Retourne les produits du fournisseur avec suggestions de quantité et
    comparaison multi-fournisseurs (alternative moins chère si elle existe)."""
    # Produits rattachés à ce fournisseur : via supplier_id principal OU via ProductSupplier
    primary_products = db.query(Product).filter(Product.supplier_id == supplier_id).all()
    via_link = (
        db.query(Product)
        .join(ProductSupplier, ProductSupplier.product_id == Product.id)
        .filter(ProductSupplier.supplier_id == supplier_id)
        .all()
    )
    # Uniq sur id
    seen = set()
    products = []
    for p in list(primary_products) + list(via_link):
        if p.id not in seen:
            seen.add(p.id)
            products.append(p)

    result = []
    for p in products:
        stock = p.stock or 0
        threshold = p.alert_threshold or 0
        target = threshold * 3
        suggested = max(0, target - stock)

        # Prix chez CE fournisseur (via ProductSupplier si dispo, sinon purchase_price global)
        link_here = next((ps for ps in (p.product_suppliers or []) if ps.supplier_id == supplier_id), None)
        price_here = link_here.purchase_price if link_here and link_here.purchase_price is not None else p.purchase_price

        # Alternatives : tous les autres fournisseurs avec un prix connu
        alternatives = []
        for ps in (p.product_suppliers or []):
            if ps.supplier_id == supplier_id:
                continue
            if ps.purchase_price is None:
                continue
            alternatives.append({
                "supplier_id":    ps.supplier_id,
                "supplier_name":  ps.supplier.name if ps.supplier else "",
                "price":          round(ps.purchase_price, 4),
            })
        alternatives.sort(key=lambda a: a["price"])

        # Moins cher ailleurs ?
        cheapest_elsewhere = alternatives[0] if alternatives else None
        is_cheapest = True
        savings_per_unit = 0.0
        if cheapest_elsewhere and price_here is not None:
            if cheapest_elsewhere["price"] < price_here:
                is_cheapest = False
                savings_per_unit = round(price_here - cheapest_elsewhere["price"], 4)

        result.append({
            "product_id":      p.id,
            "product_name":    p.name,
            "category":        p.category,
            "stock":           stock,
            "alert_threshold": threshold,
            "unit_price_ht":   price_here,
            "suggested_qty":   round(suggested, 0) if suggested > 0 else 0,
            "stock_status":    "rupture" if stock == 0 else ("low" if stock <= threshold else "ok"),
            "alternatives":    alternatives,
            "is_cheapest":     is_cheapest,
            "cheapest_elsewhere": cheapest_elsewhere,
            "savings_per_unit": savings_per_unit,
        })
    order_map = {"rupture": 0, "low": 1, "ok": 2}
    result.sort(key=lambda x: (order_map[x["stock_status"]], x["category"], x["product_name"]))
    return result


@app.get("/api/orders/best-supplier-map")
def get_best_supplier_map(db: Session = Depends(get_db)):
    """
    Pour chaque produit actif, retourne le fournisseur le moins cher (parmi ses
    ProductSuppliers). Permet de générer un plan de commandes optimisé.
    """
    products = db.query(Product).filter(
        (Product.archived == False) | (Product.archived == None)
    ).all()
    rows = []
    for p in products:
        links = [ps for ps in (p.product_suppliers or []) if ps.purchase_price is not None]
        if not links:
            # Fallback : supplier_id principal + purchase_price global
            if p.supplier_id and p.purchase_price is not None:
                rows.append({
                    "product_id":    p.id,
                    "product_name":  p.name,
                    "category":      p.category,
                    "stock":         p.stock or 0,
                    "alert_threshold": p.alert_threshold or 0,
                    "best_supplier_id":   p.supplier_id,
                    "best_supplier_name": p.supplier_rel.name if p.supplier_rel else "",
                    "best_price":    round(p.purchase_price, 4),
                    "alternatives":  [],
                })
            continue
        links_sorted = sorted(links, key=lambda x: x.purchase_price)
        best = links_sorted[0]
        rows.append({
            "product_id":    p.id,
            "product_name":  p.name,
            "category":      p.category,
            "stock":         p.stock or 0,
            "alert_threshold": p.alert_threshold or 0,
            "best_supplier_id":   best.supplier_id,
            "best_supplier_name": best.supplier.name if best.supplier else "",
            "best_price":    round(best.purchase_price, 4),
            "alternatives":  [
                {"supplier_id": l.supplier_id,
                 "supplier_name": l.supplier.name if l.supplier else "",
                 "price": round(l.purchase_price, 4)}
                for l in links_sorted[1:]
            ],
        })
    rows.sort(key=lambda x: (x["best_supplier_name"], x["category"], x["product_name"]))
    return rows


@app.post("/api/orders")
def create_order(body: OrderIn, db: Session = Depends(get_db)):
    ref = _gen_order_ref(db)
    order = SupplierOrder(
        reference=ref,
        supplier_id=body.supplier_id,
        notes=body.notes,
        status="draft",
    )
    db.add(order)
    db.flush()
    for it in body.items:
        if it.qty_ordered <= 0:
            continue
        product_name = it.product_name
        if it.product_id and not product_name:
            p = db.query(Product).get(it.product_id)
            product_name = p.name if p else ""
        db.add(SupplierOrderItem(
            order_id=order.id,
            product_id=it.product_id,
            product_name=product_name,
            qty_ordered=it.qty_ordered,
            unit_price_ht=it.unit_price_ht,
        ))
    db.commit()
    db.refresh(order)
    return _order_dict(order)


@app.get("/api/orders/{order_id}")
def get_order(order_id: int, db: Session = Depends(get_db)):
    o = db.query(SupplierOrder).get(order_id)
    if not o:
        raise HTTPException(404)
    return _order_dict(o)


@app.put("/api/orders/{order_id}")
def update_order(order_id: int, body: OrderIn, db: Session = Depends(get_db)):
    o = db.query(SupplierOrder).get(order_id)
    if not o:
        raise HTTPException(404)
    if o.status != "draft":
        raise HTTPException(400, detail="Seuls les brouillons sont modifiables")
    o.supplier_id = body.supplier_id
    o.notes = body.notes
    # Replace items
    for it in o.items:
        db.delete(it)
    db.flush()
    for it in body.items:
        if it.qty_ordered <= 0:
            continue
        product_name = it.product_name
        if it.product_id and not product_name:
            p = db.query(Product).get(it.product_id)
            product_name = p.name if p else ""
        db.add(SupplierOrderItem(
            order_id=o.id,
            product_id=it.product_id,
            product_name=product_name,
            qty_ordered=it.qty_ordered,
            unit_price_ht=it.unit_price_ht,
        ))
    db.commit()
    db.refresh(o)
    return _order_dict(o)


@app.delete("/api/orders/{order_id}")
def delete_order(order_id: int, db: Session = Depends(get_db)):
    o = db.query(SupplierOrder).get(order_id)
    if not o:
        raise HTTPException(404)
    db.delete(o)
    db.commit()
    return {"ok": True}


class OrderStatusIn(BaseModel):
    status: str  # sent / partial / received


@app.patch("/api/orders/{order_id}/status")
def update_order_status(order_id: int, body: OrderStatusIn, db: Session = Depends(get_db)):
    o = db.query(SupplierOrder).get(order_id)
    if not o:
        raise HTTPException(404)
    o.status = body.status
    if body.status == "sent" and not o.sent_at:
        o.sent_at = datetime.utcnow()
    if body.status == "received" and not o.received_at:
        o.received_at = datetime.utcnow()
    db.commit()
    return _order_dict(o)


@app.post("/api/orders/{order_id}/send-email")
def send_order_email(order_id: int, db: Session = Depends(get_db)):
    """Envoie la commande par email.
    Priorité : 1) Resend API (HTTP, fonctionne sur Railway)
               2) SMTP classique (fallback hors Railway)
               3) Retourne no_smtp=True pour fallback mailto côté client.
    """
    import urllib.request as _urllib
    import urllib.error as _urllib_err

    o = db.query(SupplierOrder).get(order_id)
    if not o:
        raise HTTPException(404)

    supplier = o.supplier
    to_email = supplier.email if supplier else ""
    reply_to = os.getenv("FROM_EMAIL", os.getenv("SMTP_USER", ""))
    subject  = f"Bon de commande {o.reference} — Marina di Lava"

    # ── Construction HTML (sans prix — le fournisseur n'en a pas besoin) ──────
    items_rows = ""
    for it in o.items:
        items_rows += f"""
        <tr>
          <td style="padding:10px 12px;border-bottom:1px solid #eee;font-size:14px">{it.product_name}</td>
          <td style="padding:10px 12px;border-bottom:1px solid #eee;text-align:center;font-size:16px;font-weight:700;color:#1a1a2e">{int(it.qty_ordered)}</td>
        </tr>"""

    notes_block = f'<p style="margin-top:16px;padding:12px 14px;background:#fffbeb;border-left:3px solid #C9A84C;font-style:italic;color:#666">{o.notes}</p>' if o.notes else ""

    html_body = f"""
    <div style="font-family:Arial,sans-serif;max-width:560px;margin:0 auto;color:#333">
      <div style="background:linear-gradient(135deg,#1a1a2e,#2d1f0e);padding:24px 28px;border-radius:10px 10px 0 0">
        <h1 style="color:#C9A84C;margin:0;font-size:22px;letter-spacing:.05em">Marina di Lava</h1>
        <p style="color:rgba(201,168,76,.7);margin:4px 0 0;font-size:13px">Bon de Commande</p>
      </div>
      <div style="background:#fff;padding:24px 28px;border:1px solid #e5e7eb;border-top:none;border-radius:0 0 10px 10px">
        <table style="width:100%;margin-bottom:20px">
          <tr><td style="color:#6b7280;font-size:12px;padding-bottom:2px">Référence</td><td style="font-weight:700;text-align:right">{o.reference}</td></tr>
          <tr><td style="color:#6b7280;font-size:12px;padding-bottom:2px">Fournisseur</td><td style="font-weight:700;text-align:right">{supplier.name if supplier else ''}</td></tr>
          <tr><td style="color:#6b7280;font-size:12px">Date</td><td style="font-weight:700;text-align:right">{to_local(datetime.utcnow()).strftime('%d/%m/%Y')}</td></tr>
        </table>
        {notes_block}
        <table style="width:100%;border-collapse:collapse;margin-top:16px">
          <thead><tr style="background:#f9fafb">
            <th style="padding:10px 12px;text-align:left;font-size:11px;text-transform:uppercase;letter-spacing:.06em;color:#6b7280;border-bottom:2px solid #e5e7eb">Produit</th>
            <th style="padding:10px 12px;text-align:center;font-size:11px;text-transform:uppercase;letter-spacing:.06em;color:#6b7280;border-bottom:2px solid #e5e7eb">Quantité</th>
          </tr></thead>
          <tbody>{items_rows}</tbody>
        </table>
        <p style="margin-top:24px;font-size:12px;color:#9ca3af;border-top:1px solid #f3f4f6;padding-top:16px">
          Bon de commande généré par le système de gestion Marina di Lava.
          {f'<br>Répondre à : <a href="mailto:{reply_to}" style="color:#C9A84C">{reply_to}</a>' if reply_to else ''}
        </p>
      </div>
    </div>"""

    # ── 1. Mailjet API (HTTP — gratuit 200/jour, pas de domaine requis) ──────
    mailjet_key    = os.getenv("MAILJET_API_KEY", "")
    mailjet_secret = os.getenv("MAILJET_SECRET_KEY", "")
    if mailjet_key and mailjet_secret:
        if not to_email:
            raise HTTPException(400, detail="Adresse email du fournisseur non renseignée")
        from_addr = reply_to or "marinadilava.commandes@gmail.com"
        import base64 as _b64
        creds   = _b64.b64encode(f"{mailjet_key}:{mailjet_secret}".encode()).decode()
        payload = json.dumps({
            "Messages": [{
                "From":     {"Email": from_addr, "Name": "Marina di Lava Commandes"},
                "To":       [{"Email": to_email}],
                "ReplyTo":  {"Email": from_addr},
                "Subject":  subject,
                "HTMLPart": html_body,
            }]
        }).encode("utf-8")
        req = _urllib.Request(
            "https://api.mailjet.com/v3.1/send",
            data=payload,
            headers={"Authorization": f"Basic {creds}", "Content-Type": "application/json"},
            method="POST",
        )
        try:
            with _urllib.urlopen(req, timeout=15) as resp:
                json.loads(resp.read())
        except _urllib_err.HTTPError as e:
            body = e.read().decode("utf-8", errors="ignore")
            raise HTTPException(500, detail=f"Erreur Mailjet {e.code}: {body[:300]}")
        except Exception as e:
            raise HTTPException(500, detail=f"Erreur Mailjet : {str(e)}")
        o.status = "sent"
        if not o.sent_at:
            o.sent_at = datetime.utcnow()
        db.commit()
        return {"ok": True, "to": to_email, "provider": "mailjet"}

    # ── 2. Resend API (nécessite domaine vérifié) ──────────────
    resend_key = os.getenv("RESEND_API_KEY", "")
    if resend_key:
        if not to_email:
            raise HTTPException(400, detail="Adresse email du fournisseur non renseignée")
        payload = json.dumps({
            "from":     "Marina di Lava Commandes <onboarding@resend.dev>",
            "reply_to": reply_to or None,
            "to":       [to_email],
            "subject":  subject,
            "html":     html_body,
        }).encode("utf-8")
        req = _urllib.Request(
            "https://api.resend.com/emails",
            data=payload,
            headers={"Authorization": f"Bearer {resend_key}", "Content-Type": "application/json"},
            method="POST",
        )
        try:
            with _urllib.urlopen(req, timeout=15) as resp:
                json.loads(resp.read())
        except Exception as e:
            raise HTTPException(500, detail=f"Erreur Resend : {str(e)}")
        o.status = "sent"
        if not o.sent_at:
            o.sent_at = datetime.utcnow()
        db.commit()
        return {"ok": True, "to": to_email, "provider": "resend"}

    # ── 2. SMTP classique (fallback) ───────────────────────────
    import smtplib
    from email.mime.multipart import MIMEMultipart as _MIME
    from email.mime.text import MIMEText as _MIMEText

    smtp_host = os.getenv("SMTP_HOST", "")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USER", "")
    smtp_pass = os.getenv("SMTP_PASS", "")
    from_email_smtp = os.getenv("FROM_EMAIL", smtp_user)

    if smtp_host and smtp_user:
        if not to_email:
            raise HTTPException(400, detail="Adresse email du fournisseur non renseignée")
        try:
            msg = _MIME("alternative")
            msg["Subject"] = subject
            msg["From"]    = from_email_smtp
            msg["To"]      = to_email
            msg.attach(_MIMEText(html_body, "html", "utf-8"))
            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.ehlo(); server.starttls()
                server.login(smtp_user, smtp_pass)
                server.sendmail(from_email_smtp, to_email, msg.as_string())
            o.status = "sent"
            if not o.sent_at:
                o.sent_at = datetime.utcnow()
            db.commit()
            return {"ok": True, "to": to_email, "provider": "smtp"}
        except Exception as e:
            raise HTTPException(500, detail=f"Erreur envoi email : {str(e)}")

    # ── 3. Pas de config → fallback mailto côté client ─────────
    return {
        "ok": False, "no_smtp": True,
        "to": to_email,
        "subject": subject,
    }


# ══════════════════════════════════════════════════════════════════════════
# CONTRÔLE DE LIVRAISON — flux 2-étapes anti-triche
# ══════════════════════════════════════════════════════════════════════════
# 1) serveur compte en aveugle      (POST counts)
# 2) direction saisit le BL + valide (PUT bl, POST validate)
# Le stock n'est mis à jour QU'à la validation (un seul point d'entrée).

class DeliveryCheckCreateIn(BaseModel):
    supplier_id: int
    order_id: Optional[int] = None
    bl_reference: Optional[str] = ""
    items: Optional[List[dict]] = None  # [{product_id, product_name, qty_expected}]


class DeliveryCheckCountsIn(BaseModel):
    checked_by: str
    counts: List[dict]   # [{item_id, qty_physical, notes}]


class DeliveryCheckBlIn(BaseModel):
    bl_reference: Optional[str] = ""
    bls: List[dict]      # [{item_id, qty_bl, unit_price_ht}]


class DeliveryCheckValidateIn(BaseModel):
    validated_by: str
    overrides: Optional[List[dict]] = None  # [{item_id, qty_validated, notes}]
    apply_prices: bool = True   # mettre à jour Product.purchase_price depuis unit_price_ht
    partial: bool = False        # validation partielle : le reste arrivera plus tard


class DeliveryCheckValidateBlOnlyIn(BaseModel):
    """
    Validation depuis BL seul (sans comptage serveur).
    Le BL reçu par email fait foi → qty_validated = qty_bl.
    Trace : skipped_count = True.
    """
    validated_by: str
    bl_reference: Optional[str] = ""
    bls: List[dict]      # [{item_id, qty_bl, unit_price_ht}]
    apply_prices: bool = True


def _delivery_check_dict(c: DeliveryCheck, for_role: str = "manager") -> dict:
    """
    Sérialise un DeliveryCheck. En mode 'service' (comptage aveugle),
    on MASQUE qty_expected, qty_bl, qty_validated pour éviter la triche.
    """
    hide_details = (for_role == "service" and c.status == "pending_count")
    return {
        "id":            c.id,
        "supplier_id":   c.supplier_id,
        "supplier_name": c.supplier.name if c.supplier else "",
        "order_id":      c.order_id,
        "order_reference": c.order.reference if c.order else "",
        "status":        c.status,
        "checked_by":    c.checked_by or "",
        "validated_by":  c.validated_by or "",
        "bl_reference":  c.bl_reference or "",
        "notes":         c.notes or "",
        "skipped_count": bool(c.skipped_count),
        "has_photo":     bool(c.bl_photo),
        "photo_type":    c.bl_photo_type or "",
        "created_at":    c.created_at.isoformat() + "Z" if c.created_at else None,
        "counted_at":    c.counted_at.isoformat() + "Z" if c.counted_at else None,
        "validated_at":  c.validated_at.isoformat() + "Z" if c.validated_at else None,
        "items": [
            {
                "id":           it.id,
                "product_id":   it.product_id,
                "product_name": it.product_name,
                "unit":         it.product.unit if it.product else "u",
                "current_purchase_price": it.product.purchase_price if it.product else None,
                "qty_expected": None if hide_details else it.qty_expected,
                "qty_bl":       None if hide_details else it.qty_bl,
                "qty_physical": it.qty_physical,
                "qty_validated": None if hide_details else it.qty_validated,
                "unit_price_ht": None if hide_details else it.unit_price_ht,
                "stock_applied": bool(it.stock_applied),
                "notes":        it.notes or "",
            }
            for it in (c.items or [])
        ],
    }


@app.get("/api/delivery-checks")
def list_delivery_checks(
    role: str = "manager",
    status: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """
    Liste les contrôles. role=service → vue aveugle (pour compter).
    status filtre optionnel.
    """
    q = db.query(DeliveryCheck).order_by(DeliveryCheck.created_at.desc())
    if status:
        q = q.filter(DeliveryCheck.status == status)
    return [_delivery_check_dict(c, for_role=role) for c in q.all()]


@app.get("/api/delivery-checks/{cid}")
def get_delivery_check(cid: int, role: str = "manager", db: Session = Depends(get_db)):
    c = db.query(DeliveryCheck).get(cid)
    if not c:
        raise HTTPException(404)
    return _delivery_check_dict(c, for_role=role)


@app.post("/api/delivery-checks")
def create_delivery_check(body: DeliveryCheckCreateIn, db: Session = Depends(get_db)):
    """
    Crée un contrôle en attente de comptage.
    Si order_id fourni : les items sont copiés depuis la commande (qty_expected = qty_ordered).
    Sinon : items libres (saisie spontanée direction).
    """
    supplier = db.query(Supplier).get(body.supplier_id)
    if not supplier:
        raise HTTPException(404, "Fournisseur introuvable")

    c = DeliveryCheck(
        supplier_id=body.supplier_id,
        order_id=body.order_id,
        bl_reference=body.bl_reference or "",
        status="pending_count",
    )
    db.add(c)
    db.flush()

    if body.order_id:
        order = db.query(SupplierOrder).get(body.order_id)
        if order:
            for it in order.items:
                db.add(DeliveryCheckItem(
                    check_id=c.id,
                    product_id=it.product_id,
                    product_name=it.product_name,
                    qty_expected=it.qty_ordered,
                ))
    elif body.items:
        for it in body.items:
            p = db.query(Product).get(it.get("product_id")) if it.get("product_id") else None
            db.add(DeliveryCheckItem(
                check_id=c.id,
                product_id=it.get("product_id"),
                product_name=it.get("product_name") or (p.name if p else ""),
                qty_expected=float(it.get("qty_expected") or 0),
            ))
    db.commit()
    db.refresh(c)
    return _delivery_check_dict(c, for_role="manager")


@app.put("/api/delivery-checks/{cid}/counts")
def submit_physical_counts(cid: int, body: DeliveryCheckCountsIn, db: Session = Depends(get_db)):
    """Serveur → saisit les quantités physiques (comptage aveugle)."""
    c = db.query(DeliveryCheck).get(cid)
    if not c:
        raise HTTPException(404)
    if c.status not in ("pending_count",):
        raise HTTPException(400, f"Contrôle déjà en statut {c.status}, comptage impossible.")

    item_map = {it.id: it for it in c.items}
    for row in body.counts:
        iid = row.get("item_id")
        it = item_map.get(iid)
        if not it:
            continue
        try:
            it.qty_physical = float(row.get("qty_physical") or 0)
        except Exception:
            it.qty_physical = 0
        it.notes = (row.get("notes") or "")[:500]

    c.checked_by = body.checked_by or c.checked_by or ""
    c.counted_at = datetime.utcnow()
    c.status = "counted"
    db.commit()
    return _delivery_check_dict(c, for_role="manager")


@app.put("/api/delivery-checks/{cid}/bl")
def submit_bl_quantities(cid: int, body: DeliveryCheckBlIn, db: Session = Depends(get_db)):
    """Direction → saisit les quantités du BL papier."""
    c = db.query(DeliveryCheck).get(cid)
    if not c:
        raise HTTPException(404)
    if c.status == "validated":
        raise HTTPException(400, "Contrôle déjà validé, BL figé.")

    item_map = {it.id: it for it in c.items}
    for row in body.bls:
        iid = row.get("item_id")
        it = item_map.get(iid)
        if not it:
            continue
        try:
            it.qty_bl = float(row.get("qty_bl")) if row.get("qty_bl") is not None else None
        except Exception:
            pass
        if row.get("unit_price_ht") is not None and row.get("unit_price_ht") != "":
            try:
                it.unit_price_ht = float(row["unit_price_ht"])
            except Exception:
                pass

    if body.bl_reference is not None:
        c.bl_reference = body.bl_reference or ""
    db.commit()
    return _delivery_check_dict(c, for_role="manager")


@app.post("/api/delivery-checks/{cid}/validate")
def validate_delivery_check(cid: int, body: DeliveryCheckValidateIn, db: Session = Depends(get_db)):
    """
    Direction → valide : applique le stock = qty_validated (ou qty_physical par défaut).
    C'est le SEUL endroit où le stock est augmenté.
    """
    c = db.query(DeliveryCheck).get(cid)
    if not c:
        raise HTTPException(404)
    if c.status == "validated":
        raise HTTPException(400, "Déjà validé.")

    # Appliquer éventuels overrides (qty finale décidée par la direction)
    override_map = {o.get("item_id"): o for o in (body.overrides or [])}
    applied = []
    for it in c.items:
        # Skip items déjà appliqués lors d'une validation partielle précédente
        if it.stock_applied:
            continue

        qty_final = None
        if it.id in override_map and override_map[it.id].get("qty_validated") is not None:
            try:
                qty_final = float(override_map[it.id]["qty_validated"])
            except Exception:
                qty_final = None
            if override_map[it.id].get("notes"):
                it.notes = override_map[it.id]["notes"][:500]
        if qty_final is None:
            qty_final = it.qty_physical if it.qty_physical is not None else (it.qty_bl if it.qty_bl is not None else 0)

        # En validation partielle : on ne traite que les items avec une qté > 0
        if body.partial and (not qty_final or qty_final <= 0):
            continue

        it.qty_validated = qty_final

        if it.product_id and qty_final and qty_final > 0:
            p = db.query(Product).get(it.product_id)
            if p:
                p.stock = (p.stock or 0) + qty_final
                it.stock_applied = True
                applied.append({"product_id": p.id, "product_name": p.name, "added": qty_final})
                if body.apply_prices and it.unit_price_ht is not None and it.unit_price_ht > 0:
                    old_price = p.purchase_price
                    new_price = float(it.unit_price_ht)
                    if old_price is None or abs(old_price - new_price) > 0.001:
                        record_price_change(
                            db, p, new_price,
                            source="bl_validation",
                            supplier_id=c.supplier_id,
                            reference=c.bl_reference or (c.order.reference if c.order else ""),
                        )
                        p.purchase_price = new_price
                        p.is_estimated = False

    c.validated_by = body.validated_by or c.validated_by or ""
    # Statut final
    remaining = [it for it in c.items if not it.stock_applied]
    if body.partial and remaining:
        c.status = "partial"
        # ne pas figer validated_at en cas de partiel (attendu la suite)
    else:
        c.status = "validated"
        c.validated_at = datetime.utcnow()

    # Si lié à une commande → maj statut (partial ou received)
    if c.order_id:
        order = db.query(SupplierOrder).get(c.order_id)
        if order and order.status != "received":
            if c.status == "validated":
                order.status = "received"
                order.received_at = datetime.utcnow()
            elif c.status == "partial" and order.status != "partial":
                order.status = "partial"

    log_event(
        db, "livraison",
        f"Contrôle livraison validé — {c.supplier.name if c.supplier else ''} ({len(applied)} produits)",
        {"delivery_check_id": c.id, "bl_reference": c.bl_reference, "items": applied},
    )
    db.commit()
    return _delivery_check_dict(c, for_role="manager")


@app.post("/api/delivery-checks/{cid}/validate-bl-only")
def validate_delivery_check_bl_only(
    cid: int,
    body: DeliveryCheckValidateBlOnlyIn,
    db: Session = Depends(get_db),
):
    """
    Direction → valide directement depuis le BL sans comptage serveur.
    qty_validated = qty_bl. Marque skipped_count = True pour traçabilité.
    Seule la direction utilise ce chemin (cas où personne n'a pu compter).
    """
    c = db.query(DeliveryCheck).get(cid)
    if not c:
        raise HTTPException(404)
    if c.status != "pending_count":
        raise HTTPException(400, f"Validation BL seul autorisée uniquement depuis 'pending_count' (actuel : {c.status}).")

    # Applique le BL : qty_bl + unit_price_ht par item
    item_map = {it.id: it for it in c.items}
    for row in body.bls:
        iid = row.get("item_id")
        it = item_map.get(iid)
        if not it:
            continue
        try:
            it.qty_bl = float(row.get("qty_bl")) if row.get("qty_bl") is not None else None
        except Exception:
            pass
        if row.get("unit_price_ht") not in (None, ""):
            try:
                it.unit_price_ht = float(row["unit_price_ht"])
            except Exception:
                pass

    if body.bl_reference is not None:
        c.bl_reference = body.bl_reference or c.bl_reference or ""

    # Applique le stock : qty_validated = qty_bl
    applied = []
    for it in c.items:
        if it.stock_applied:
            continue
        qty_final = it.qty_bl if it.qty_bl is not None else 0
        it.qty_validated = qty_final
        if it.product_id and qty_final and qty_final > 0:
            p = db.query(Product).get(it.product_id)
            if p:
                p.stock = (p.stock or 0) + qty_final
                it.stock_applied = True
                applied.append({"product_id": p.id, "product_name": p.name, "added": qty_final})
                if body.apply_prices and it.unit_price_ht is not None and it.unit_price_ht > 0:
                    old_price = p.purchase_price
                    new_price = float(it.unit_price_ht)
                    if old_price is None or abs(old_price - new_price) > 0.001:
                        record_price_change(
                            db, p, new_price,
                            source="bl_only_validation",
                            supplier_id=c.supplier_id,
                            reference=c.bl_reference or (c.order.reference if c.order else ""),
                        )
                        p.purchase_price = new_price
                        p.is_estimated = False

    c.validated_by = body.validated_by or c.validated_by or ""
    c.skipped_count = True
    c.status = "validated"
    c.validated_at = datetime.utcnow()

    if c.order_id:
        order = db.query(SupplierOrder).get(c.order_id)
        if order and order.status != "received":
            order.status = "received"
            order.received_at = datetime.utcnow()

    log_event(
        db, "livraison",
        f"Contrôle livraison validé depuis BL seul — {c.supplier.name if c.supplier else ''} ({len(applied)} produits, sans comptage)",
        {"delivery_check_id": c.id, "bl_reference": c.bl_reference, "items": applied, "skipped_count": True},
    )
    db.commit()
    return _delivery_check_dict(c, for_role="manager")


class AddDcItemIn(BaseModel):
    product_id: Optional[int] = None
    product_name: Optional[str] = ""
    qty_expected: float = 0


@app.post("/api/delivery-checks/{cid}/items")
def add_delivery_check_item(cid: int, body: AddDcItemIn, db: Session = Depends(get_db)):
    """Ajoute un produit à un contrôle existant (livraison spontanée, produit oublié...)."""
    c = db.query(DeliveryCheck).get(cid)
    if not c:
        raise HTTPException(404)
    if c.status == "validated":
        raise HTTPException(400, "Contrôle déjà validé, impossible de modifier.")
    p = db.query(Product).get(body.product_id) if body.product_id else None
    it = DeliveryCheckItem(
        check_id=c.id,
        product_id=body.product_id,
        product_name=body.product_name or (p.name if p else ""),
        qty_expected=float(body.qty_expected or 0),
    )
    db.add(it)
    db.commit()
    db.refresh(it)
    return _delivery_check_dict(c, for_role="manager")


@app.delete("/api/delivery-checks/{cid}/items/{iid}")
def delete_delivery_check_item(cid: int, iid: int, db: Session = Depends(get_db)):
    c = db.query(DeliveryCheck).get(cid)
    if not c or c.status == "validated":
        raise HTTPException(400, "Impossible de modifier.")
    it = db.query(DeliveryCheckItem).filter(DeliveryCheckItem.id == iid, DeliveryCheckItem.check_id == cid).first()
    if not it:
        raise HTTPException(404)
    db.delete(it)
    db.commit()
    return {"ok": True}


@app.post("/api/delivery-checks/{cid}/reject")
def reject_delivery_check(cid: int, db: Session = Depends(get_db)):
    """Direction refuse (erreur manifeste, à recompter) → reset pour nouveau comptage."""
    c = db.query(DeliveryCheck).get(cid)
    if not c:
        raise HTTPException(404)
    if c.status == "validated":
        raise HTTPException(400, "Impossible de rejeter un contrôle déjà validé.")
    c.status = "pending_count"
    for it in c.items:
        it.qty_physical = None
    c.counted_at = None
    c.checked_by = ""
    db.commit()
    return _delivery_check_dict(c, for_role="manager")


@app.post("/api/delivery-checks/{cid}/photo")
async def upload_bl_photo(cid: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Attache une photo / PDF du BL au contrôle."""
    c = db.query(DeliveryCheck).get(cid)
    if not c:
        raise HTTPException(404)
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:   # 5 MB max
        raise HTTPException(400, "Fichier trop gros (max 5 MB).")
    ext_map = {
        "jpg": "image/jpeg", "jpeg": "image/jpeg",
        "png": "image/png", "webp": "image/webp",
        "heic": "image/jpeg", "gif": "image/gif",
        "pdf": "application/pdf",
    }
    fname = (file.filename or "").lower()
    ext = fname.rsplit(".", 1)[-1] if "." in fname else "jpg"
    media_type = file.content_type or ext_map.get(ext, "image/jpeg")
    # Réoriente la photo selon EXIF (iPhone) — bypass pour PDF
    content, media_type = _apply_exif_orientation(content, media_type)
    c.bl_photo = base64.standard_b64encode(content).decode("utf-8")
    c.bl_photo_type = media_type
    db.commit()
    return {"ok": True, "has_photo": True, "photo_type": media_type}


@app.get("/api/delivery-checks/{cid}/photo")
def get_bl_photo(cid: int, db: Session = Depends(get_db)):
    """Retourne la photo/PDF du BL (décodé)."""
    from fastapi.responses import Response
    c = db.query(DeliveryCheck).get(cid)
    if not c or not c.bl_photo:
        raise HTTPException(404)
    try:
        data = base64.standard_b64decode(c.bl_photo)
    except Exception:
        raise HTTPException(500, "Photo corrompue")
    return Response(content=data, media_type=c.bl_photo_type or "image/jpeg")


@app.delete("/api/delivery-checks/{cid}/photo")
def delete_bl_photo(cid: int, db: Session = Depends(get_db)):
    c = db.query(DeliveryCheck).get(cid)
    if not c:
        raise HTTPException(404)
    c.bl_photo = ""
    c.bl_photo_type = ""
    db.commit()
    return {"ok": True}


@app.delete("/api/delivery-checks/{cid}")
def delete_delivery_check(cid: int, rollback_stock: bool = False, db: Session = Depends(get_db)):
    """
    Supprime un contrôle de livraison.
    - pending/counted/partial : suppression directe.
    - validated : nécessite rollback_stock=true (décrémente le stock appliqué).
    """
    c = db.query(DeliveryCheck).get(cid)
    if not c:
        raise HTTPException(404)

    if c.status == "validated":
        if not rollback_stock:
            raise HTTPException(
                400,
                "Contrôle validé — ajouter ?rollback_stock=true pour le supprimer et décrémenter le stock.",
            )
        # Rollback : décrémente le stock pour chaque item appliqué
        rolled = []
        for it in c.items:
            if it.stock_applied and it.qty_validated and it.product_id:
                p = db.query(Product).get(it.product_id)
                if p:
                    p.stock = max(0, (p.stock or 0) - it.qty_validated)
                    rolled.append({"product_id": p.id, "product_name": p.name, "removed": it.qty_validated})
        log_event(
            db, "livraison",
            f"Suppression contrôle validé #{c.id} — rollback stock sur {len(rolled)} items",
            {"delivery_check_id": c.id, "rollback": rolled},
        )

    db.delete(c)
    db.commit()
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════
# PRODUITS / PRODUCTS
# ══════════════════════════════════════════════════════════════════════════

class ProductIn(BaseModel):
    name: str
    category: str
    supplier_id: Optional[int] = None
    stock: float = 0
    unit: str = "Bouteille"
    qty_per_pack: float = 1
    volume_cl: float = 70
    alert_threshold: float = 2
    purchase_price: Optional[float] = None
    sale_price_ttc: Optional[float] = None
    is_estimated: bool = False
    barcode: str = ""
    vat_rate: float = 0.20


@app.get("/api/produits")
@app.get("/api/products")
def get_products(db: Session = Depends(get_db)):
    products = db.query(Product).all()
    return [calc_product(p) for p in products]


@app.post("/api/produits")
@app.post("/api/products")
def create_product(body: ProductIn, db: Session = Depends(get_db)):
    p = Product(**body.model_dump())
    db.add(p)
    db.commit()
    db.refresh(p)
    return calc_product(p)


@app.get("/api/products/by-barcode/{code}")
def get_product_by_barcode(code: str, db: Session = Depends(get_db)):
    p = db.query(Product).filter(Product.barcode == code).first()
    if not p:
        raise HTTPException(404, detail="Produit non trouvé")
    return calc_product(p)


@app.put("/api/produits/{pid}")
@app.put("/api/products/{pid}")
def update_product(pid: int, body: ProductIn, db: Session = Depends(get_db)):
    p = db.query(Product).get(pid)
    if not p:
        raise HTTPException(404)
    # Traquer les changements de prix avant override
    payload = body.model_dump()
    new_price = payload.get("purchase_price")
    if new_price is not None:
        record_price_change(db, p, new_price, source="manual_edit", supplier_id=p.supplier_id)
    for k, v in payload.items():
        setattr(p, k, v)
    p.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(p)
    return calc_product(p)


@app.delete("/api/produits/{pid}")
@app.delete("/api/products/{pid}")
def delete_product(pid: int, db: Session = Depends(get_db)):
    p = db.query(Product).get(pid)
    if not p:
        raise HTTPException(404)
    db.delete(p)
    db.commit()
    return {"ok": True}


class ArchiveIn(BaseModel):
    archived: bool = True
    product_ids: Optional[List[int]] = None


@app.post("/api/products/{pid}/archive")
def archive_product(pid: int, body: ArchiveIn, db: Session = Depends(get_db)):
    p = db.query(Product).get(pid)
    if not p:
        raise HTTPException(404)
    p.archived = bool(body.archived)
    db.commit()
    return {"ok": True, "archived": p.archived}


@app.post("/api/products/archive-bulk")
def archive_products_bulk(body: ArchiveIn, db: Session = Depends(get_db)):
    if not body.product_ids:
        return {"ok": True, "updated": 0}
    count = (
        db.query(Product)
        .filter(Product.id.in_(body.product_ids))
        .update({"archived": bool(body.archived)}, synchronize_session=False)
    )
    db.commit()
    return {"ok": True, "updated": count, "archived": bool(body.archived)}


# ── Multi-fournisseurs par produit ─────────────────────────────────────────

class ProductSupplierIn(BaseModel):
    supplier_id: int
    purchase_price: Optional[float] = None
    is_primary: bool = False


@app.get("/api/products/{pid}/suppliers")
def get_product_suppliers(pid: int, db: Session = Depends(get_db)):
    rows = db.query(ProductSupplier).filter(ProductSupplier.product_id == pid).all()
    return [
        {
            "id": r.id,
            "supplier_id": r.supplier_id,
            "supplier_name": r.supplier.name if r.supplier else "",
            "purchase_price": r.purchase_price,
            "is_primary": r.is_primary,
        }
        for r in sorted(rows, key=lambda x: (not x.is_primary, x.id))
    ]


@app.put("/api/products/{pid}/suppliers")
def update_product_suppliers(pid: int, body: List[ProductSupplierIn], db: Session = Depends(get_db)):
    p = db.query(Product).get(pid)
    if not p:
        raise HTTPException(404)

    # S'assurer qu'il y a exactement un is_primary
    has_primary = any(s.is_primary for s in body)
    if body and not has_primary:
        body[0].is_primary = True

    # Remplacer tous les liens fournisseurs
    db.query(ProductSupplier).filter(ProductSupplier.product_id == pid).delete()
    for s in body:
        db.add(ProductSupplier(
            product_id=pid,
            supplier_id=s.supplier_id,
            purchase_price=s.purchase_price,
            is_primary=s.is_primary,
        ))

    # Synchroniser les champs legacy depuis le fournisseur principal
    primary = next((s for s in body if s.is_primary), body[0] if body else None)
    if primary:
        p.supplier_id = primary.supplier_id
        if primary.purchase_price is not None:
            p.purchase_price = primary.purchase_price
            p.is_estimated = False

    db.commit()
    db.refresh(p)
    return calc_product(p)


# ══════════════════════════════════════════════════════════════════════════
# RECETTES / COCKTAILS
# ══════════════════════════════════════════════════════════════════════════

class IngredientIn(BaseModel):
    product_id: int
    dose_cl: float


class CocktailIn(BaseModel):
    name: str
    sale_price_ttc: Optional[float] = None
    ingredients: List[IngredientIn] = []


@app.get("/api/recettes")
@app.get("/api/cocktails")
def get_cocktails(db: Session = Depends(get_db)):
    return [calc_cocktail(c) for c in db.query(Cocktail).all()]


@app.post("/api/recettes")
@app.post("/api/cocktails")
def create_cocktail(body: CocktailIn, db: Session = Depends(get_db)):
    c = Cocktail(name=body.name, sale_price_ttc=body.sale_price_ttc)
    db.add(c)
    db.flush()
    for ing in body.ingredients:
        db.add(CocktailIngredient(cocktail_id=c.id, product_id=ing.product_id, dose_cl=ing.dose_cl))
    db.commit()
    db.refresh(c)
    return calc_cocktail(c)


@app.put("/api/recettes/{cid}")
@app.put("/api/cocktails/{cid}")
def update_cocktail(cid: int, body: CocktailIn, db: Session = Depends(get_db)):
    c = db.query(Cocktail).get(cid)
    if not c:
        raise HTTPException(404)
    c.name = body.name
    c.sale_price_ttc = body.sale_price_ttc
    db.query(CocktailIngredient).filter(CocktailIngredient.cocktail_id == cid).delete()
    for ing in body.ingredients:
        db.add(CocktailIngredient(cocktail_id=c.id, product_id=ing.product_id, dose_cl=ing.dose_cl))
    db.commit()
    db.refresh(c)
    return calc_cocktail(c)


@app.delete("/api/recettes/{cid}")
@app.delete("/api/cocktails/{cid}")
def delete_cocktail(cid: int, db: Session = Depends(get_db)):
    c = db.query(Cocktail).get(cid)
    if not c:
        raise HTTPException(404)
    db.delete(c)
    db.commit()
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════
# ALERTES
# ══════════════════════════════════════════════════════════════════════════

@app.get("/api/predictions")
def get_predictions(db: Session = Depends(get_db)):
    """Prédit la date de rupture de chaque produit basé sur la consommation cashpad des 14 derniers jours."""
    from collections import defaultdict

    PERIODE = 14  # jours d'analyse
    cutoff = datetime.utcnow() - timedelta(days=PERIODE)

    # Récupère tous les imports cashpad des 14 derniers jours
    rows = db.query(StockHistory).filter(
        StockHistory.created_at >= cutoff,
        StockHistory.event_type == "import_cashpad"
    ).order_by(StockHistory.created_at).all()

    # Calcule la consommation totale par product_id
    consumption = defaultdict(float)   # product_id → total unités consommées
    days_seen   = defaultdict(set)     # product_id → set de jours avec vente

    for row in rows:
        try:
            d = json.loads(row.data_json)
        except Exception:
            continue
        day = row.created_at.strftime("%Y-%m-%d")
        for ded in d.get("deductions", []):
            pid = ded.get("product_id")
            qty = abs(float(ded.get("quantity", ded.get("qty", 0)) or 0))
            if pid and qty > 0:
                consumption[pid] += qty
                days_seen[pid].add(day)

    # Besoins événementiels à venir (jusqu'à 14j) — par produit
    today_dt = datetime.combine(datetime.utcnow().date(), datetime.min.time())
    horizon_dt = today_dt + timedelta(days=PERIODE)
    upcoming = (
        db.query(Event)
        .filter(Event.date <= horizon_dt)
        .filter(
            (Event.date >= today_dt) |
            ((Event.end_date != None) & (Event.end_date >= today_dt))
        )
        .all()
    )
    event_needs = defaultdict(list)   # product_id → [{qty, event_name, days_until}]
    for ev in upcoming:
        ev_start = ev.date.date() if hasattr(ev.date, "date") else ev.date
        days_until = max(0, (ev_start - datetime.utcnow().date()).days)
        for req in (ev.requirements or []):
            if req.product_id and req.quantity:
                event_needs[req.product_id].append({
                    "qty":         float(req.quantity),
                    "event_name":  ev.name,
                    "days_until":  days_until,
                    "event_date":  ev.date.strftime("%d/%m/%Y"),
                })

    # Boost saisonnier lié à la météo : on regarde la température prévue demain
    weather_boost = 1.0
    weather_label = None
    try:
        cached = _get_setting(db, "weather_cache", "")
        if cached:
            w = json.loads(cached)
            level = w.get("alert_level")
            if level == "canicule":
                weather_boost = 1.30   # +30 % par canicule
                weather_label = "canicule prévue (+30%)"
            elif level == "chaud":
                weather_boost = 1.15   # +15 % par forte chaleur
                weather_label = "forte chaleur prévue (+15%)"
    except Exception:
        pass

    # Catégories plus sensibles à la chaleur (softs, eaux, bières, rosé)
    _HEAT_SENSITIVE = {"Eaux", "Sodas", "Bières", "Vins Rosés", "Cocktails SA"}

    products = db.query(Product).filter((Product.archived == False) | (Product.archived == None)).all()
    predictions = []

    for p in products:
        if p.id not in consumption:
            continue
        total_consumed = consumption[p.id]
        n_days = max(len(days_seen[p.id]), 1)
        avg_per_active_day = total_consumed / n_days
        base_avg_daily = total_consumed / PERIODE

        # Appliquer le boost météo aux catégories sensibles
        avg_daily = base_avg_daily
        applied_weather = None
        if weather_boost > 1.0 and p.category in _HEAT_SENSITIVE:
            avg_daily = base_avg_daily * weather_boost
            applied_weather = weather_label

        if avg_daily <= 0 or p.stock <= 0:
            continue

        # Stock effectif : on soustrait les besoins événements à venir dans l'horizon
        event_reservations = event_needs.get(p.id, [])
        total_event_need = sum(r["qty"] for r in event_reservations)
        effective_stock = p.stock - total_event_need

        # Si les événements à eux seuls vident le stock → rupture programmée à la date du 1er event
        if effective_stock <= 0 and event_reservations:
            first_ev = min(event_reservations, key=lambda r: r["days_until"])
            days_left = float(first_ev["days_until"])
        else:
            days_left = max(effective_stock, 0) / avg_daily

        predicted_date = to_local(datetime.utcnow()) + timedelta(days=days_left)

        urgency = "critique" if days_left < 3 else "warning" if days_left < 7 else "info" if days_left < 14 else None
        if not urgency:
            continue

        predictions.append({
            "product_id":    p.id,
            "product_name":  p.name,
            "category":      p.category,
            "stock":         round(p.stock, 1),
            "effective_stock": round(effective_stock, 1),
            "avg_daily":     round(avg_daily, 2),
            "base_avg_daily":round(base_avg_daily, 2),
            "days_left":     round(days_left, 1),
            "predicted_date": predicted_date.strftime("%d/%m/%Y"),
            "urgency":       urgency,
            "total_consumed_14j": round(total_consumed, 1),
            "event_reservations": event_reservations,
            "event_need_total":   round(total_event_need, 1),
            "weather_boost":      applied_weather,
        })

    predictions.sort(key=lambda x: x["days_left"])
    return predictions


@app.get("/api/alertes")
@app.get("/api/alerts")
def get_alerts(db: Session = Depends(get_db)):
    alerts = []
    for p in db.query(Product).filter((Product.archived == False) | (Product.archived == None)).all():
        if p.stock == 0:
            alerts.append({
                "type": "rupture",
                "product_id": p.id,
                "product": p.name,
                "message": f"Rupture de stock : {p.name}",
                "severity": "high"
            })
        elif p.stock <= p.alert_threshold:
            alerts.append({
                "type": "stock_bas",
                "product_id": p.id,
                "product": p.name,
                "message": f"Stock bas : {p.name} ({p.stock} {p.unit})",
                "severity": "medium"
            })
        if p.purchase_price and p.sale_price_ttc:
            cout = p.purchase_price  # déjà en prix/unité individuelle
            vat_r = p.vat_rate if p.vat_rate is not None else 0.20
            ht = p.sale_price_ttc / (1 + vat_r)
            if ht > 0:
                marge = (ht - cout) / ht * 100
                if marge < 50 and not p.is_estimated:
                    alerts.append({
                        "type": "marge",
                        "product_id": p.id,
                        "product": p.name,
                        "message": f"Marge insuffisante : {p.name} ({marge:.1f}%)",
                        "severity": "medium"
                    })
    # Also fetch inventory gap alerts from recent history
    recent_inv = db.query(StockHistory).filter(
        StockHistory.event_type == "alerte_inventaire"
    ).order_by(StockHistory.created_at.desc()).limit(10).all()
    for h in recent_inv:
        try:
            data = json.loads(h.data_json)
            alerts.append({
                "type": "ecart_inventaire",
                "product": data.get("product", ""),
                "message": f"Écart inventaire : {data.get('product', '')} (écart {data.get('diff', 0):+.3f})",
                "severity": "high",
                "date": h.created_at.isoformat() + "Z",
            })
        except Exception:
            pass

    # ── Besoins événements : stock insuffisant pour un événement à venir ──
    today_dt = datetime.combine(datetime.utcnow().date(), datetime.min.time())
    upcoming = (
        db.query(Event)
        .filter(
            (Event.date >= today_dt) |
            ((Event.end_date != None) & (Event.end_date >= today_dt))
        )
        .order_by(Event.date.asc())
        .all()
    )
    # Cumul des besoins par produit pour tous les événements à venir
    cumul = {}   # product_id → [{"event": ev, "qty": q}]
    for ev in upcoming:
        for r in (ev.requirements or []):
            if not r.product or not r.quantity:
                continue
            if r.product.archived:
                continue
            cumul.setdefault(r.product_id, []).append({"event": ev, "qty": r.quantity, "product": r.product})
    for pid, entries in cumul.items():
        p = entries[0]["product"]
        total_need = sum(e["qty"] for e in entries)
        if total_need > p.stock:
            missing = total_need - p.stock
            if len(entries) == 1:
                ev = entries[0]["event"]
                date_str = ev.date.strftime("%d/%m")
                msg = (f"Besoin événement : {p.name} — {int(entries[0]['qty']) if entries[0]['qty'].is_integer() else entries[0]['qty']} "
                       f"demandé pour « {ev.name} » ({date_str}), stock actuel {int(p.stock) if float(p.stock).is_integer() else p.stock} "
                       f"→ manque {int(missing) if float(missing).is_integer() else round(missing, 2)}")
            else:
                noms = ", ".join(f"« {e['event'].name} »" for e in entries[:3])
                if len(entries) > 3:
                    noms += f" +{len(entries)-3}"
                msg = (f"Besoin événements : {p.name} — {int(total_need) if total_need.is_integer() else total_need} demandés "
                       f"({noms}), stock {int(p.stock) if float(p.stock).is_integer() else p.stock} "
                       f"→ manque {int(missing) if float(missing).is_integer() else round(missing, 2)}")
            alerts.append({
                "type": "besoin_evenement",
                "product_id": pid,
                "product": p.name,
                "message": msg,
                "severity": "high",
            })
    return alerts


# ══════════════════════════════════════════════════════════════════════════
# ALERTES SERVEUR (signalements terrain)
# ══════════════════════════════════════════════════════════════════════════

class ServiceAlertIn(BaseModel):
    product_id: int
    reported_stock: float
    is_rupture: bool = False
    staff_name: str = ""
    notes: str = ""


@app.post("/api/service-alerts")
def create_service_alert(body: ServiceAlertIn, db: Session = Depends(get_db)):
    """Un serveur signale un stock bas ou une rupture."""
    p = db.query(Product).get(body.product_id)
    if not p:
        raise HTTPException(404, detail="Produit introuvable")

    alert = ServiceAlert(
        product_id=body.product_id,
        reported_stock=body.reported_stock,
        is_rupture=body.is_rupture,
        staff_name=body.staff_name,
        notes=body.notes,
    )
    db.add(alert)

    event_desc = f"{'Rupture' if body.is_rupture else 'Stock bas'} signalé par {body.staff_name or 'service'} : {p.name} ({body.reported_stock} restant)"
    log_event(db, "alerte_service", event_desc, {
        "product_id": p.id, "product_name": p.name,
        "reported_stock": body.reported_stock, "is_rupture": body.is_rupture,
        "staff_name": body.staff_name,
    })
    db.commit()
    db.refresh(alert)
    return {"ok": True, "id": alert.id, "product_name": p.name}


@app.get("/api/service-alerts")
def list_service_alerts(status: str = "open", db: Session = Depends(get_db)):
    """Liste les alertes serveur. status=open|all"""
    q = db.query(ServiceAlert).order_by(ServiceAlert.created_at.desc())
    if status != "all":
        q = q.filter(ServiceAlert.status == status)
    alerts = q.limit(100).all()
    result = []
    for a in alerts:
        p = a.product
        supplier_name = ""
        if p and p.product_suppliers:
            primary = next((ps for ps in p.product_suppliers if ps.is_primary), None)
            if primary and primary.supplier:
                supplier_name = primary.supplier.name
        result.append({
            "id": a.id,
            "product_id": a.product_id,
            "product_name": p.name if p else "?",
            "category": p.category if p else "",
            "reported_stock": a.reported_stock,
            "is_rupture": a.is_rupture,
            "staff_name": a.staff_name,
            "notes": a.notes,
            "status": a.status,
            "supplier_name": supplier_name,
            "created_at": to_local(a.created_at).strftime("%d/%m %H:%M") if a.created_at else "",
        })
    return result


@app.patch("/api/service-alerts/{alert_id}")
def update_service_alert(alert_id: int, db: Session = Depends(get_db)):
    """Direction acknowledge/résout une alerte."""
    a = db.query(ServiceAlert).get(alert_id)
    if not a:
        raise HTTPException(404, detail="Alerte introuvable")
    if a.status == "open":
        a.status = "acknowledged"
    elif a.status == "acknowledged":
        a.status = "ordered"
    elif a.status == "ordered":
        a.status = "resolved"
        a.resolved_at = datetime.utcnow()
    db.commit()
    return {"ok": True, "status": a.status}


# ══════════════════════════════════════════════════════════════════════════
# HISTORIQUE
# ══════════════════════════════════════════════════════════════════════════

@app.get("/api/historique")
@app.get("/api/history")
def get_history(db: Session = Depends(get_db)):
    events = db.query(StockHistory).order_by(StockHistory.created_at.desc()).limit(200).all()
    return [
        {
            "id": e.id,
            "event_type": e.event_type,
            "description": e.description,
            "data": json.loads(e.data_json),
            "created_at": e.created_at.isoformat() + "Z",
        }
        for e in events
    ]


class ManualMovementIn(BaseModel):
    product_id: int
    quantity: float
    note: str = ""


@app.post("/api/history/manual")
def manual_movement(body: ManualMovementIn, db: Session = Depends(get_db)):
    p = db.query(Product).get(body.product_id)
    if not p:
        raise HTTPException(404)
    old = p.stock
    p.stock += body.quantity
    db.commit()
    data = {"product_id": p.id, "product": p.name, "old_stock": old, "new_stock": p.stock,
            "quantity": body.quantity, "note": body.note or ""}
    h = log_event(db, "mouvement_manuel", f"Mouvement manuel : {p.name} ({'+' if body.quantity >= 0 else ''}{body.quantity})", data)
    db.commit()
    # Alerte email si le stock a baissé (sortie réserve, ajustement négatif)
    if body.quantity < 0:
        try: _trigger_rupture_alerts(db, context=f"Sortie réserve : {p.name}")
        except Exception as _e: print(f"[Alert email] skipped : {_e}")
    return {"ok": True, "new_stock": p.stock, "history_id": h.id}


# ══════════════════════════════════════════════════════════════════════════
# TEMPLATES DE COMMANDES (récurrentes)
# ══════════════════════════════════════════════════════════════════════════

class OrderTemplateItemIn(BaseModel):
    product_id: int
    default_qty: float = 0


class OrderTemplateIn(BaseModel):
    name: str
    supplier_id: int
    items: List[OrderTemplateItemIn] = []


def _template_dict(t: OrderTemplate) -> dict:
    return {
        "id":           t.id,
        "name":         t.name,
        "supplier_id":  t.supplier_id,
        "supplier_name": t.supplier.name if t.supplier else "",
        "last_used_at": t.last_used_at.isoformat() + "Z" if t.last_used_at else None,
        "created_at":   t.created_at.isoformat() + "Z" if t.created_at else None,
        "items": [
            {
                "product_id":   it.product_id,
                "product_name": it.product.name if it.product else "",
                "unit":         it.product.unit if it.product else "",
                "default_qty":  it.default_qty,
            }
            for it in (t.items or [])
        ],
    }


@app.get("/api/order-templates")
def list_order_templates(supplier_id: Optional[int] = None, db: Session = Depends(get_db)):
    q = db.query(OrderTemplate).order_by(OrderTemplate.name.asc())
    if supplier_id:
        q = q.filter(OrderTemplate.supplier_id == supplier_id)
    return [_template_dict(t) for t in q.all()]


@app.post("/api/order-templates")
def create_order_template(body: OrderTemplateIn, db: Session = Depends(get_db)):
    if not body.name.strip():
        raise HTTPException(400, "Nom requis")
    t = OrderTemplate(name=body.name.strip(), supplier_id=body.supplier_id)
    db.add(t)
    db.flush()
    for it in body.items:
        if it.default_qty and it.default_qty > 0:
            db.add(OrderTemplateItem(template_id=t.id, product_id=it.product_id, default_qty=float(it.default_qty)))
    db.commit()
    db.refresh(t)
    return _template_dict(t)


@app.put("/api/order-templates/{tid}")
def update_order_template(tid: int, body: OrderTemplateIn, db: Session = Depends(get_db)):
    t = db.query(OrderTemplate).get(tid)
    if not t:
        raise HTTPException(404)
    t.name = body.name.strip() or t.name
    t.supplier_id = body.supplier_id
    for old in list(t.items):
        db.delete(old)
    for it in body.items:
        if it.default_qty and it.default_qty > 0:
            db.add(OrderTemplateItem(template_id=t.id, product_id=it.product_id, default_qty=float(it.default_qty)))
    db.commit()
    db.refresh(t)
    return _template_dict(t)


@app.delete("/api/order-templates/{tid}")
def delete_order_template(tid: int, db: Session = Depends(get_db)):
    t = db.query(OrderTemplate).get(tid)
    if not t:
        raise HTTPException(404)
    db.delete(t)
    db.commit()
    return {"ok": True}


@app.post("/api/order-templates/{tid}/mark-used")
def mark_template_used(tid: int, db: Session = Depends(get_db)):
    t = db.query(OrderTemplate).get(tid)
    if not t:
        raise HTTPException(404)
    t.last_used_at = datetime.utcnow()
    db.commit()
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════
# CHAT IA — assistant conversationnel sur les données métier
# ══════════════════════════════════════════════════════════════════════════

class ChatMessageIn(BaseModel):
    role: str       # "user" | "assistant"
    content: str


class ChatIn(BaseModel):
    messages: List[ChatMessageIn]


def _build_chat_context(db: Session) -> str:
    """Construit un contexte textuel des données actuelles pour Claude."""
    lines = ["=== CONTEXTE MARINA DI LAVA (actualisé à la requête) ===\n"]

    # Produits actifs avec stock
    lines.append("PRODUITS (id | nom | catégorie | stock | unité | seuil | prix achat | prix vente TTC | TVA) :")
    for p in db.query(Product).filter((Product.archived == False) | (Product.archived == None)).order_by(Product.category, Product.name).all():
        pa = f"€{p.purchase_price:.3f}" if p.purchase_price is not None else "?"
        pv = f"€{p.sale_price_ttc:.2f}" if p.sale_price_ttc is not None else "?"
        vat = int((p.vat_rate or 0.20) * 100)
        lines.append(f"  {p.id} | {p.name} | {p.category} | {p.stock:.1f} | {p.unit} | seuil {p.alert_threshold} | {pa} | {pv} | {vat}%")

    # Fournisseurs
    lines.append("\nFOURNISSEURS :")
    for s in db.query(Supplier).all():
        email_str = s.email or "(pas d'email)"
        lines.append(f"  {s.id} | {s.name} | {email_str} | catégories: {s.categories}")

    # Événements à venir (30 prochains jours)
    today_dt = datetime.utcnow()
    horizon = today_dt + timedelta(days=30)
    lines.append("\nÉVÉNEMENTS À VENIR (30 jours) :")
    for e in db.query(Event).filter(Event.date <= horizon).filter(
        (Event.date >= today_dt.replace(hour=0, minute=0, second=0, microsecond=0)) |
        ((Event.end_date != None) & (Event.end_date >= today_dt.replace(hour=0, minute=0, second=0, microsecond=0)))
    ).all():
        dates = e.date.strftime("%d/%m")
        if e.end_date and e.end_date != e.date:
            dates += f"→{e.end_date.strftime('%d/%m')}"
        reqs = [f"{r.product.name if r.product else '?'} x{r.quantity}" for r in (e.requirements or [])]
        lines.append(f"  {dates} | {e.event_type} | {e.name}" + (f" | besoins : {', '.join(reqs)}" if reqs else ""))

    # Commandes fournisseurs en cours
    lines.append("\nCOMMANDES EN COURS :")
    for o in db.query(SupplierOrder).filter(SupplierOrder.status.in_(["draft", "sent", "partial"])).all():
        lines.append(f"  {o.reference} | {o.supplier.name if o.supplier else '?'} | {o.status} | {len(o.items or [])} produit(s)")

    # Contrôles livraison en attente
    lines.append("\nCONTRÔLES LIVRAISON EN ATTENTE :")
    for c in db.query(DeliveryCheck).filter(DeliveryCheck.status.in_(["pending_count", "counted", "partial"])).all():
        lines.append(f"  {c.id} | {c.supplier.name if c.supplier else '?'} | statut {c.status} | BL {c.bl_reference or '—'}")

    # Ventes récentes (7 derniers jours)
    week_start = datetime.utcnow() - timedelta(days=7)
    sales_hist = db.query(StockHistory).filter(
        StockHistory.event_type.in_(["import_cashpad", "cashpad_sync"]),
        StockHistory.created_at >= week_start,
    ).all()
    totals = {}
    ca = 0.0
    for h in sales_hist:
        for item in extract_sales(h):
            qty = float(item.get("qty_sold", 0) or 0)
            price = float(item.get("sale_price_ttc", 0) or 0)
            name = item.get("product_name") or item.get("name", "")
            ca += qty * price
            if name:
                totals[name] = totals.get(name, 0) + qty
    top = sorted(totals.items(), key=lambda x: -x[1])[:10]
    lines.append(f"\nVENTES 7 DERNIERS JOURS (CA={ca:.0f}€, top 10) :")
    for name, qty in top:
        lines.append(f"  {name} : {qty:.0f}")

    # Historique prix récent
    recent_prices = db.query(PriceHistory).order_by(PriceHistory.changed_at.desc()).limit(10).all()
    if recent_prices:
        lines.append("\nVARIATIONS PRIX D'ACHAT RÉCENTES (10 dernières) :")
        for h in recent_prices:
            pct = (h.new_price - h.old_price) / h.old_price * 100 if h.old_price else 0
            lines.append(f"  {h.product.name if h.product else '?'} : {h.old_price}→{h.new_price} ({pct:+.1f}%) via {h.source}")

    # Pertes déclarées du mois
    month_start = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    losses = db.query(ManualLoss).filter(ManualLoss.date >= month_start).all()
    if losses:
        lines.append(f"\nPERTES DÉCLARÉES CE MOIS ({len(losses)}) :")
        for l in losses[:20]:
            lines.append(f"  {l.product.name if l.product else '?'} × {l.quantity} | {l.reason}")

    # Météo cache
    try:
        w = json.loads(_get_setting(db, "weather_cache", "") or "{}")
        if w:
            lines.append(f"\nMÉTÉO : {w.get('current_temp')}°C · {w.get('alert_label', '')}")
    except Exception:
        pass

    return "\n".join(lines)


_CHAT_SYSTEM_INSTRUCTIONS = """Tu es **l'assistant IA de Marina di Lava**, un hôtel-restaurant-bar de plage à Ajaccio (Corse).
Ton rôle : aider Jean-Marc (le gérant) à gérer son stock, ses ventes et son activité.

Règles :
- Toujours répondre en **français**, ton direct et pragmatique (J-Marc n'aime pas les réponses alambiquées).
- Baser toutes tes réponses sur le **CONTEXTE** fourni ci-dessous. Ne pas inventer de chiffres.
- Si tu ne peux pas répondre avec les données disponibles, dis-le honnêtement.
- Format : réponses courtes et exploitables. Utilise des listes à puces si pertinent.
- Pour les montants : utilise le format 123,45 € (virgule + espace insécable avant €).
- Si la question implique une action concrète (passer une commande, archiver un produit, créer un événement…), suggère clairement l'endroit où la faire dans l'app :
  * Stock & Marges : ajustements produits, seuils d'alerte
  * Cocktails & Marges : recettes
  * Alertes : tous les signaux (ruptures, événements sans stock…)
  * Démarque : pertes déclarées, démarque automatique, rapprochement sortie/Cashpad
  * Contrôle livraison : réceptionner un BL + admin + backups
  * Commandes : passer une commande fournisseur
  * Événements : agenda + besoins boissons clients
  * Statistiques : CA, consommation, historique prix
- Ne pas donner de conseils médicaux, juridiques ou comportementaux hors du champ de la gestion du bar."""


@app.post("/api/chat")
def chat_endpoint(body: ChatIn, db: Session = Depends(get_db)):
    """Chat conversationnel avec Claude, contextualisé sur les données métier."""
    api_key = os.getenv("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        raise HTTPException(400, "Clé API Anthropic non configurée (ANTHROPIC_API_KEY)")

    if not body.messages:
        raise HTTPException(400, "Message vide")

    # Normaliser les messages (rôles autorisés : user / assistant)
    msgs = []
    for m in body.messages[-20:]:   # on limite à 20 derniers tours pour maîtriser le coût
        if m.role not in ("user", "assistant"):
            continue
        if not m.content or not m.content.strip():
            continue
        msgs.append({"role": m.role, "content": m.content.strip()})
    if not msgs or msgs[-1]["role"] != "user":
        raise HTTPException(400, "Dernier message doit venir de l'utilisateur")

    # Construire le contexte actuel des données
    context = _build_chat_context(db)

    import anthropic as _anthropic
    client = _anthropic.Anthropic(api_key=api_key)

    try:
        response = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1024,
            system=[
                {"type": "text", "text": _CHAT_SYSTEM_INSTRUCTIONS},
                # Le contexte change à chaque requête → pas de cache ici, mais
                # les instructions système, elles, sont cachées par Anthropic
                # automatiquement (préfixe identique).
                {"type": "text", "text": context},
            ],
            messages=msgs,
        )
    except _anthropic.AuthenticationError:
        raise HTTPException(401, "Clé API Anthropic invalide")
    except _anthropic.BadRequestError as e:
        raise HTTPException(400, f"Requête invalide : {str(e)[:200]}")
    except Exception as e:
        raise HTTPException(500, f"Erreur IA : {str(e)[:200]}")

    reply = ""
    try:
        for block in response.content:
            if getattr(block, "type", "") == "text":
                reply += block.text
        reply = reply.strip()
    except Exception:
        reply = "(Réponse vide)"

    return {
        "reply": reply,
        "usage": {
            "input_tokens":  response.usage.input_tokens,
            "output_tokens": response.usage.output_tokens,
        },
    }


@app.get("/api/price-history")
def price_history_summary(limit: int = 30, db: Session = Depends(get_db)):
    """Liste les dernières variations de prix d'achat, tous produits confondus."""
    rows = (
        db.query(PriceHistory)
        .order_by(PriceHistory.changed_at.desc())
        .limit(limit)
        .all()
    )
    return [
        {
            "id":            h.id,
            "product_id":    h.product_id,
            "product_name":  h.product.name if h.product else "",
            "category":      h.product.category if h.product else "",
            "old_price":     round(h.old_price, 4) if h.old_price is not None else None,
            "new_price":     round(h.new_price, 4),
            "delta_pct":     round(((h.new_price - h.old_price) / h.old_price) * 100, 1) if h.old_price and h.old_price > 0 else None,
            "changed_at":    to_local(h.changed_at).strftime("%d/%m/%Y %H:%M") if h.changed_at else "",
            "source":        h.source or "",
            "supplier_name": h.supplier.name if h.supplier else "",
            "reference":     h.reference or "",
        }
        for h in rows
    ]


@app.get("/api/products/{pid}/price-history")
def price_history_product(pid: int, db: Session = Depends(get_db)):
    """Historique détaillé d'un produit (pour graphique)."""
    rows = (
        db.query(PriceHistory)
        .filter(PriceHistory.product_id == pid)
        .order_by(PriceHistory.changed_at.asc())
        .all()
    )
    return [
        {
            "id":        h.id,
            "old_price": round(h.old_price, 4) if h.old_price is not None else None,
            "new_price": round(h.new_price, 4),
            "changed_at": h.changed_at.isoformat() + "Z" if h.changed_at else "",
            "source":    h.source or "",
            "supplier_name": h.supplier.name if h.supplier else "",
            "reference": h.reference or "",
        }
        for h in rows
    ]


@app.get("/api/reconciliation")
def reconciliation(period: str = "today", db: Session = Depends(get_db)):
    """
    Rapproche les sorties réserve (mouvement_manuel négatif) avec les ventes
    Cashpad sur la même période. Un écart = produit parti sans être vendu
    (offert, cassé, volé, erreur de saisie).

    period: today | yesterday | week (7 jours)
    """
    today = datetime.utcnow().date()
    if period == "yesterday":
        start_dt = datetime.combine(today - timedelta(days=1), datetime.min.time())
        end_dt   = datetime.combine(today, datetime.min.time())
        label    = "hier"
    elif period == "week":
        start_dt = datetime.combine(today - timedelta(days=7), datetime.min.time())
        end_dt   = datetime.utcnow()
        label    = "7 derniers jours"
    else:
        start_dt = datetime.combine(today, datetime.min.time())
        end_dt   = datetime.utcnow()
        label    = "aujourd'hui"

    # Sorties réserve (mouvement_manuel quantity < 0) groupées par produit + staff
    sorties_hist = (
        db.query(StockHistory)
        .filter(
            StockHistory.event_type == "mouvement_manuel",
            StockHistory.created_at >= start_dt,
            StockHistory.created_at <  end_dt,
        )
        .all()
    )
    sorties = {}   # product_id → {qty: float, staff: [names]}
    for h in sorties_hist:
        try:
            d = json.loads(h.data_json or "{}")
        except Exception:
            continue
        qty = float(d.get("quantity") or 0)
        if qty >= 0:
            continue  # on ne garde que les sorties (négatifs)
        pid = d.get("product_id")
        if not pid:
            continue
        note = (d.get("note") or "").lower()
        if "sortie réserve" not in note and "sortie reserve" not in note:
            continue   # ignorer les autres mouvements manuels (ajustements)
        rec = sorties.setdefault(pid, {"qty": 0.0, "staff": set()})
        rec["qty"] += abs(qty)
        # Extraire le nom après "par"
        if " par " in note:
            try:
                name = note.split(" par ")[1].split(":")[0].strip().title()
                if name:
                    rec["staff"].add(name)
            except Exception:
                pass

    # Ventes Cashpad par produit sur la période
    ventes_hist = (
        db.query(StockHistory)
        .filter(
            StockHistory.event_type.in_(["import_cashpad", "cashpad_sync"]),
            StockHistory.created_at >= start_dt,
            StockHistory.created_at <  end_dt,
        )
        .all()
    )
    ventes = {}   # product_id → qty
    for h in ventes_hist:
        # On utilise deductions (quantités effectivement retirées du stock)
        try:
            d = json.loads(h.data_json or "{}")
        except Exception:
            continue
        deds = d.get("deductions") if isinstance(d, dict) else []
        for ded in (deds or []):
            pid = ded.get("product_id")
            qty = abs(float(ded.get("quantity", ded.get("deducted", 0)) or 0))
            if pid and qty > 0:
                ventes[pid] = ventes.get(pid, 0) + qty

    # Rapprochement
    all_pids = set(sorties.keys()) | set(ventes.keys())
    products = {p.id: p for p in db.query(Product).filter(Product.id.in_(all_pids)).all()}
    rows = []
    total_gap_eur = 0.0
    for pid in all_pids:
        p = products.get(pid)
        if not p:
            continue
        sortie_qty = sorties.get(pid, {}).get("qty", 0)
        sortie_staff = sorted(sorties.get(pid, {}).get("staff", set()))
        vente_qty  = ventes.get(pid, 0)
        gap = sortie_qty - vente_qty

        # Seuil d'ignorance : écart < 5% ou < 1 unité → "normal"
        is_significant = abs(gap) >= 1 or (sortie_qty > 0 and abs(gap / sortie_qty) >= 0.05)

        severity = "ok"
        if gap > 0 and is_significant:
            severity = "warn" if gap < 3 else "danger"
        elif gap < 0 and abs(gap) >= 2:
            severity = "info"   # plus vendu que sorti → stock non tracé, anomalie inverse

        loss_eur = 0.0
        if gap > 0 and p.purchase_price:
            loss_eur = round(gap * p.purchase_price, 2)
            if severity != "ok":
                total_gap_eur += loss_eur

        rows.append({
            "product_id":    p.id,
            "product_name":  p.name,
            "category":      p.category,
            "unit":          p.unit or "u",
            "sortie_qty":    round(sortie_qty, 2),
            "ventes_qty":    round(vente_qty, 2),
            "gap":           round(gap, 2),
            "loss_eur":      loss_eur,
            "severity":      severity,
            "sortie_staff":  sortie_staff,
        })

    rows.sort(key=lambda x: -x["loss_eur"] if x["severity"] != "ok" else 1)
    return {
        "period":        period,
        "period_label":  label,
        "start":         start_dt.isoformat() + "Z",
        "end":           end_dt.isoformat() + "Z",
        "rows":          rows,
        "total_gap_eur": round(total_gap_eur, 2),
        "n_anomalies":   sum(1 for r in rows if r["severity"] in ("warn", "danger")),
    }


@app.get("/api/sorties/today")
def sorties_today(staff: str = "", db: Session = Depends(get_db)):
    """Retourne les sorties réserve du jour, filtrées par prénom si fourni."""
    from datetime import date
    today_start = datetime.combine(date.today(), datetime.min.time())
    rows = db.query(StockHistory).filter(
        StockHistory.event_type == "mouvement_manuel",
        StockHistory.created_at >= today_start
    ).order_by(StockHistory.created_at.desc()).all()

    result = []
    for r in rows:
        try:
            d = json.loads(r.data_json)
        except Exception:
            continue
        if d.get("quantity", 0) >= 0:
            continue  # garder seulement les sorties (négatif)
        note = d.get("note", "")
        if staff and staff.lower() not in note.lower():
            continue
        result.append({
            "history_id": r.id,
            "product_id": d.get("product_id"),
            "product": d.get("product", "?"),
            "quantity": d.get("quantity", 0),
            "note": note,
            "created_at": to_local(r.created_at).strftime("%H:%M"),
        })
    return result


@app.post("/api/sorties/annuler/{history_id}")
def annuler_sortie(history_id: int, db: Session = Depends(get_db)):
    """Annule une sortie en créant un mouvement inverse."""
    h = db.query(StockHistory).get(history_id)
    if not h or h.event_type != "mouvement_manuel":
        raise HTTPException(404, "Sortie introuvable")
    try:
        d = json.loads(h.data_json)
    except Exception:
        raise HTTPException(400, "Données invalides")
    qty = d.get("quantity", 0)
    if qty >= 0:
        raise HTTPException(400, "Ce n'est pas une sortie")
    product_id = d.get("product_id")
    p = db.query(Product).get(product_id)
    if not p:
        raise HTTPException(404, "Produit introuvable")
    old = p.stock
    p.stock -= qty  # inverse : réintègre la quantité
    db.commit()
    log_event(db, "mouvement_manuel",
              f"Annulation sortie : {p.name} (+{abs(qty)})",
              {"product_id": p.id, "product": p.name, "old_stock": old,
               "new_stock": p.stock, "quantity": abs(qty), "note": f"Annulation de la sortie #{history_id}"})
    db.commit()
    return {"ok": True, "new_stock": p.stock}


# ══════════════════════════════════════════════════════════════════════════
# CASHPAD MAPPING
# ══════════════════════════════════════════════════════════════════════════

class MappingIn(BaseModel):
    nom_cashpad: str
    product_id: Optional[int] = None
    cocktail_id: Optional[int] = None
    dose_cl: float = 0
    mapping_type: str = "direct"
    ignored: bool = False


def _mapping_dict(r: CashpadMapping) -> dict:
    return {
        "id": r.id,
        "nom_cashpad": r.nom_cashpad,
        "product_id": r.product_id,
        "cocktail_id": r.cocktail_id,
        "dose_cl": r.dose_cl,
        "type": r.mapping_type,
        "mapping_type": r.mapping_type,
        "ignore": r.ignored,
        "ignored": r.ignored,
        "product_name": r.product.name if r.product else None,
        "cocktail_name": db_cocktail_name(r),
    }


def db_cocktail_name(r: CashpadMapping) -> Optional[str]:
    return None  # resolved in route if needed


@app.get("/api/cashpad_mapping")
@app.get("/api/cashpad-mapping")
def get_mapping(db: Session = Depends(get_db)):
    rows = db.query(CashpadMapping).all()
    result = []
    cocktail_names = {}
    for c in db.query(Cocktail).all():
        cocktail_names[c.id] = c.name
    for r in rows:
        result.append({
            "id": r.id,
            "nom_cashpad": r.nom_cashpad,
            "product_id": r.product_id,
            "cocktail_id": r.cocktail_id,
            "dose_cl": r.dose_cl,
            "type": r.mapping_type,
            "mapping_type": r.mapping_type,
            "ignore": r.ignored,
            "ignored": r.ignored,
            "product_name": r.product.name if r.product else None,
            "cocktail_name": cocktail_names.get(r.cocktail_id) if r.cocktail_id else None,
        })
    return result


@app.post("/api/cashpad_mapping")
@app.post("/api/cashpad-mapping")
def create_mapping(body: MappingIn, db: Session = Depends(get_db)):
    m = CashpadMapping(
        nom_cashpad=body.nom_cashpad,
        product_id=body.product_id,
        cocktail_id=body.cocktail_id,
        dose_cl=body.dose_cl,
        mapping_type=body.mapping_type,
        ignored=body.ignored,
    )
    db.add(m)
    db.commit()
    db.refresh(m)
    return {"id": m.id}


@app.put("/api/cashpad_mapping/{mid}")
@app.put("/api/cashpad-mapping/{mid}")
def update_mapping(mid: int, body: MappingIn, db: Session = Depends(get_db)):
    m = db.query(CashpadMapping).get(mid)
    if not m:
        raise HTTPException(404)
    m.nom_cashpad = body.nom_cashpad
    m.product_id = body.product_id
    m.cocktail_id = body.cocktail_id
    m.dose_cl = body.dose_cl
    m.mapping_type = body.mapping_type
    m.ignored = body.ignored
    db.commit()
    return {"ok": True}


@app.delete("/api/cashpad_mapping/{mid}")
@app.delete("/api/cashpad-mapping/{mid}")
def delete_mapping(mid: int, db: Session = Depends(get_db)):
    m = db.query(CashpadMapping).get(mid)
    if not m:
        raise HTTPException(404)
    db.delete(m)
    db.commit()
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════
# IMPORT CASHPAD
# ══════════════════════════════════════════════════════════════════════════

@app.post("/api/import/cashpad")
async def import_cashpad(
    file: UploadFile = File(...),
    numero_cloture: str = Form(""),
    db: Session = Depends(get_db),
):
    if numero_cloture:
        existing = db.query(ImportLog).filter(
            ImportLog.import_type == "cashpad",
            ImportLog.reference == numero_cloture,
        ).first()
        if existing:
            raise HTTPException(
                400,
                detail=f"Clôture n°{numero_cloture} déjà importée le {to_local(existing.created_at).strftime('%d/%m/%Y %H:%M')}."
            )

    import openpyxl
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    # find sheet — prefer "Ventes-par-produit-2"
    sheet = None
    for sname in wb.sheetnames:
        if "Ventes-par-produit-2" in sname:
            sheet = wb[sname]
            break
    if sheet is None:
        for sname in wb.sheetnames:
            if "ventes" in sname.lower() or "produit" in sname.lower():
                sheet = wb[sname]
                break
    if sheet is None:
        sheet = wb.active

    # build mapping index
    mappings = {m.nom_cashpad.strip().lower(): m for m in db.query(CashpadMapping).all()}

    deductions = []      # impact stock (audit)
    sales = []           # lignes de vente (CA, marge) — prix figés au moment de la vente
    alerts_triggered = []
    unknown_products = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        product_name = str(row[1]).strip()
        # quantity is at column index 6
        try:
            qty_sold = float(row[6]) if len(row) > 6 and row[6] is not None else 0
        except (ValueError, TypeError):
            qty_sold = 0
        if qty_sold <= 0:
            continue

        key = product_name.lower()
        mapping = mappings.get(key)
        if not mapping:
            unknown_products.append(product_name)
            continue
        if mapping.ignored:
            continue

        if mapping.mapping_type == "cocktail" and mapping.cocktail_id:
            cocktail = db.query(Cocktail).get(mapping.cocktail_id)
            if cocktail:
                cost_matiere = 0.0
                has_alcohol = False
                for ing in cocktail.ingredients:
                    p = ing.product
                    if p and p.volume_cl and ing.dose_cl:
                        total_vol = p.volume_cl
                        deduct = (ing.dose_cl / total_vol) * qty_sold
                        p.stock = p.stock - deduct
                        deductions.append({
                            "product": p.name,
                            "deducted": round(deduct, 4),
                            "new_stock": round(p.stock, 4),
                        })
                        if p.purchase_price is not None:
                            cost_matiere += (p.purchase_price / total_vol) * ing.dose_cl
                        if (p.vat_rate or 0) >= 0.20:
                            has_alcohol = True
                        if p.stock <= p.alert_threshold:
                            alerts_triggered.append(p.name)
                sales.append({
                    "ref_type":       "cocktail",
                    "ref_id":         cocktail.id,
                    "product_name":   cocktail.name,
                    "name":           cocktail.name,
                    "qty_sold":       qty_sold,
                    "sale_price_ttc": cocktail.sale_price_ttc or 0,
                    "purchase_price": round(cost_matiere, 4),
                    "vat_rate":       0.20 if has_alcohol else 0.10,
                    "unit":           "u",
                })
        elif mapping.mapping_type == "direct" and mapping.product_id:
            p = db.query(Product).get(mapping.product_id)
            if p and p.volume_cl and mapping.dose_cl:
                total_vol = p.volume_cl
                deduct = (mapping.dose_cl / total_vol) * qty_sold
                p.stock = p.stock - deduct
                deductions.append({
                    "product": p.name,
                    "deducted": round(deduct, 4),
                    "new_stock": round(p.stock, 4),
                })
                unit_cost = (p.purchase_price / total_vol) * mapping.dose_cl if p.purchase_price is not None else None
                sales.append({
                    "ref_type":       "direct",
                    "ref_id":         p.id,
                    "product_id":     p.id,
                    "product_name":   mapping.nom_cashpad,
                    "name":           mapping.nom_cashpad,
                    "qty_sold":       qty_sold,
                    "sale_price_ttc": 0,  # inconnu côté Cashpad Excel, complété plus tard si possible
                    "purchase_price": round(unit_cost, 4) if unit_cost is not None else None,
                    "vat_rate":       p.vat_rate if p.vat_rate is not None else 0.20,
                    "unit":           p.unit or "u",
                })
                if p.stock <= p.alert_threshold:
                    alerts_triggered.append(p.name)

    ref = numero_cloture or f"auto-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"
    db.add(ImportLog(import_type="cashpad", reference=ref))
    log_event(
        db,
        "import_cashpad",
        f"Import Cashpad clôture n°{ref} — {len(deductions)} déductions",
        {
            "numero_cloture": ref,
            "deductions":     deductions,
            "sales":          sales,
            "alerts":         list(set(alerts_triggered)),
            "unknown":        unknown_products,
        }
    )
    db.commit()
    # Alerte email (non bloquante)
    try: _trigger_rupture_alerts(db, context=f"Import Cashpad {ref}")
    except Exception as _e: print(f"[Alert email] skipped : {_e}")

    return {
        "ok": True,
        "deductions": deductions,
        "alerts": list(set(alerts_triggered)),
        "unknown": unknown_products,
    }


# ══════════════════════════════════════════════════════════════════════════
# IMPORT BON DE LIVRAISON
# ══════════════════════════════════════════════════════════════════════════

def _dedup_char(s):
    """Supprime les caractères doublés du codage PDF SOCOBO (ex: 'AABB' → 'AB')."""
    if not s:
        return s
    result = []
    i = 0
    chars = list(s)
    while i < len(chars):
        c = chars[i]
        if i + 1 < len(chars) and chars[i + 1] == c:
            result.append(c)
            i += 2
        else:
            result.append(c)
            i += 1
    return ''.join(result)


def _dedup_line(line):
    return ' '.join(_dedup_char(w) for w in line.split())


def _parse_socobo_pdf(content_bytes):
    """
    Parse une facture SOCOBO DSAC directement via pdfplumber.
    Retourne (products_list, invoice_num).
    products_list = [{"nom":..., "quantite":..., "prix_unitaire_ht":..., "numero_facture":...}]
    """
    import pdfplumber as _pdfplumber
    import io as _io
    import re as _re

    invoice_num = None
    raw_lines = []  # (code, nom, quantite, prix, is_promo)

    with _pdfplumber.open(_io.BytesIO(content_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text(x_tolerance=3, y_tolerance=3)
            if not text:
                continue
            for line in text.split('\n'):
                dl = _dedup_line(line)

                # Numéro de facture (6 chiffres commençant par 1)
                if not invoice_num:
                    m = _re.search(r'\b(1\d{5})\b', dl)
                    if m:
                        invoice_num = m.group(1)

                # Ligne produit : commence par code numérique suivi (immédiatement) d'une lettre
                if not _re.match(r'^\d{2,5}\*?\s*[A-ZÉÀÙÊÎF]', dl):
                    continue

                upper = dl.upper()
                SKIP = ['FRAIS DE REGIE', 'FUTÀ30', 'FUTà30', 'FUT À 30',
                        'DECONSIGN', 'BON DE RETOUR', 'CO25KG', 'TUBE CO2',
                        'SANS EM ', 'CARTEBANCAIRE', 'PAR NOUS', 'SOCOBO',
                        'CLIENT DÉCLAR', 'AUCUNE ESCOMPTE', 'LE TAUX',
                        'CASESCERTIF', 'RENVOI', 'SIGNATURECLIENT']
                if any(s.upper() in upper for s in SKIP):
                    continue
                if _re.match(r'^\d{3,},\d{2}', dl):
                    continue

                is_promo = 'PROMO FOURN' in upper

                m_code = _re.match(r'^(\d+)\*?\s*', dl)
                if not m_code:
                    continue
                code = m_code.group(1)
                rest = dl[m_code.end():]
                rest_up = rest.upper()

                if not rest or not rest[0].isalpha():
                    continue

                # ── TYPE 1 : CARTON N X ou FUT N X → COLIS COLS suivent ──
                m_cx = _re.search(
                    r'(CARTON|FUT)\s+(\d+)\s+X\s+(\d+)\s+(\d+)\s+', rest_up)
                # ── TYPE 2 : LOT DE N → COLIS COLS suivent ──
                m_lot = _re.search(
                    r'LOT\s+DE\s+(\d+)\s+(\d+)\s+(\d+)\s+', rest_up)
                # ── TYPE 3 : PACK DE N → COLIS seul (prix par pack) ──
                m_pack = _re.search(
                    r'PACK\s+DE\s+(\d+)\s+(\d+)\s+', rest_up)

                if m_cx:
                    kw = m_cx.group(1)
                    n = int(m_cx.group(2))
                    colis = int(m_cx.group(3))
                    cols = int(m_cx.group(4))
                    is_fut_line = (kw == 'FUT')
                    quantite = colis if is_fut_line else cols
                    nom = (rest[:m_cx.start()].strip() + ' ' + kw + ' ' + str(n) + ' X').strip()
                    rest_after = rest_up[m_cx.end():]
                    pr = _re.findall(r'\d+[,\.]\d+', rest_after)
                    # SOCOBO affiche le prix par unité (€/L pour fûts, €/bouteille pour cartons)
                    # → multiplier par n pour obtenir le prix par fût ou par carton
                    prix_unit = float(pr[0].replace(',', '.')) if pr else None
                    prix = round(prix_unit * n, 4) if prix_unit is not None else None

                elif m_lot:
                    n = int(m_lot.group(1))
                    colis = int(m_lot.group(2))
                    cols = int(m_lot.group(3))
                    quantite = cols
                    nom = (rest[:m_lot.start()].strip() + ' LOT DE ' + str(n)).strip()
                    rest_after = rest_up[m_lot.end():]
                    pr = _re.findall(r'\d+[,\.]\d+', rest_after)
                    prix = float(pr[0].replace(',', '.')) if pr else None

                elif m_pack:
                    n = int(m_pack.group(1))
                    colis = int(m_pack.group(2))
                    quantite = colis * n
                    nom = rest[:m_pack.end(1)].strip()  # inclut "PACK DE N"
                    rest_after = rest_up[m_pack.end():]
                    pr = _re.findall(r'\d+[,\.]\d+', rest_after)
                    prix = float(pr[0].replace(',', '.')) if pr else None

                else:
                    # Produit simple : chercher COLIS suivi de PRIX PRIX
                    m_c = _re.search(r'(\d+)\s+(\d+[,\.]\d+)\s+(\d+[,\.]\d+)', rest)
                    if not m_c:
                        continue
                    colis = int(m_c.group(1))
                    prix = float(m_c.group(2).replace(',', '.'))
                    nom = rest[:m_c.start()].strip()
                    # xN ou /N dans le nom → multiplier
                    m_xn = _re.search(r'[xX×](\d+)|/(\d+)', nom)
                    if m_xn:
                        n = int(m_xn.group(1) or m_xn.group(2))
                        quantite = colis * n
                    else:
                        quantite = colis

                if not nom or quantite <= 0:
                    continue
                raw_lines.append((code, nom.strip(), quantite, prix, is_promo))

    # Agréger : PROMO FOURN ajoute au produit principal (même code de base)
    products = {}
    for code, nom, quantite, prix, is_promo in raw_lines:
        if code not in products:
            products[code] = {"nom": nom, "quantite": 0, "prix": None}
        products[code]["quantite"] += quantite
        if not is_promo:
            if prix:
                products[code]["prix"] = prix
            products[code]["nom"] = nom  # nom issu de la ligne principale

    result = []
    for v in products.values():
        if v["quantite"] > 0:
            result.append({
                "nom": v["nom"],
                "quantite": v["quantite"],
                "prix_unitaire_ht": v["prix"],
                "numero_facture": invoice_num,
            })
    return result, invoice_num


def _parse_esprit_du_vin_pdf(content_bytes):
    """
    Parse une facture SARL Esprit du Vin.
    Format : REF [GENCOD] DESIGNATION QTY,00 CONDI,00 P.U.HT TOTAL TVA%
    Retourne (products_list, invoice_num).
    """
    import pdfplumber as _pdfplumber
    import io as _io
    import re as _re

    invoice_num = None
    products_out = []

    with _pdfplumber.open(_io.BytesIO(content_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text(x_tolerance=3, y_tolerance=3)
            if not text:
                continue
            for line in text.split('\n'):
                line = line.strip()
                if not invoice_num:
                    m = _re.search(r'FACTURE\s+n°(\d+)', line)
                    if m:
                        invoice_num = m.group(1)

                # Ligne produit : REF [GENCOD?] DESIGNATION QTY,00 CONDI,00 PU TOTAL TVA%
                m_prod = _re.match(
                    r'^([A-Z]{2,6}\d*\w*)\s+'
                    r'(?:\d{13}\s+)?'
                    r'(.+?)\s+'
                    r'(\d+)[,\.]00\s+'
                    r'(\d+)[,\.]00\s+'
                    r'(\d+[,\.]\d+)\s+'
                    r'(\d[\d\s]*[,\.]\d+)\s+'
                    r'(\d+)\s*%',
                    line
                )
                if not m_prod:
                    continue

                ref = m_prod.group(1)
                nom = m_prod.group(2).strip()
                quantite = int(m_prod.group(3))
                prix_ht = float(m_prod.group(5).replace(',', '.'))

                SKIP_REFS = {'SARL', 'IBAN', 'RCS', 'TVA', 'NET', 'REF', 'TOTAL', 'MODE', 'NET'}
                if ref.upper() in SKIP_REFS:
                    continue
                if quantite <= 0 or not nom:
                    continue

                products_out.append({
                    'nom': nom,
                    'quantite': quantite,
                    'prix_unitaire_ht': prix_ht,
                    'numero_facture': invoice_num,
                })

    # Vérifier que c'est bien une facture Esprit du Vin
    if not invoice_num or not products_out:
        return [], None
    return products_out, invoice_num


def _parse_pgv_vv_pdf(content_bytes):
    """
    Parse les factures PGV Distribution (PGFA...) et Empreinte du Vin (VVFA...).
    Format : REFCODE DESIGNATION QTY,000 [C|U] COLS PRIX_BRUT PRIX_NET MONTANT TVA
    Quantité = COLS (bouteilles individuelles). Les avoirs génèrent des quantités négatives.
    """
    import pdfplumber as _pdfplumber
    import io as _io
    import re as _re

    invoice_num = None
    is_avoir = False
    products_out = []

    with _pdfplumber.open(_io.BytesIO(content_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text(x_tolerance=3, y_tolerance=3)
            if not text:
                continue
            for line in text.split('\n'):
                line = line.strip()

                if not invoice_num:
                    m = _re.search(r'(PGFA|VVFA)\s+(\d+)\s*/\s*\d+', line)
                    if m:
                        invoice_num = m.group(1) + m.group(2)

                if _re.search(r'\bAvoir\b', line, _re.I):
                    is_avoir = True

                m_prod = _re.match(
                    r'^(\d{6})\s+'
                    r'(.+?)\s+'
                    r'(\d+)[,\.]\d+\s+'
                    r'[CU]\s+'
                    r'(\d+)\s+'
                    r'(\d+[,\.]\d+)\s+'
                    r'(\d+[,\.]\d+)\s+'
                    r'(\d[\d\s]*[,\.]\d+)\s+'
                    r'(\d+)\s*',
                    line
                )
                if not m_prod:
                    continue

                nom = m_prod.group(2).strip()
                cols = int(m_prod.group(4))
                prix_ht = float(m_prod.group(6).replace(',', '.'))
                quantite = -cols if is_avoir else cols

                if any(x in nom.upper() for x in ['CONSIGN', 'DECONSIGN', 'FRAIS']):
                    continue
                if cols == 0:
                    continue

                products_out.append({
                    'nom': nom,
                    'quantite': quantite,
                    'prix_unitaire_ht': prix_ht,
                    'numero_facture': invoice_num,
                })

    if not invoice_num or not products_out:
        return [], None
    return products_out, invoice_num


def _parse_auchan_pdf(content_bytes):
    """
    Parse une commande Auchan (email Gmail exporté en PDF).
    Retourne (products_list, order_num).
    """
    import pdfplumber as _pdfplumber
    import io as _io
    import re as _re

    order_num = None
    products_out = []

    with _pdfplumber.open(_io.BytesIO(content_bytes)) as pdf:
        all_words = []
        y_offset = 0
        for page in pdf.pages:
            ws = page.extract_words(x_tolerance=3, y_tolerance=3)
            for w in ws:
                all_words.append({**w, 'top': w['top'] + y_offset})
            y_offset += page.height

    # Numéro de commande Auchan (format 224XXXXXXXXXX)
    for w in all_words:
        if _re.match(r'^224\d{10}$', w['text']):
            order_num = w['text']
            break
    if not order_num:
        return [], None  # pas une commande Auchan

    # Séparation par zone x :
    # Quantité : x = 240-315
    # Prix     : x > 315
    # Nom      : x < 250
    qty_words   = [w for w in all_words if 240 <= w['x0'] <= 315 and _re.match(r'^\d+$', w['text'])]
    price_words = [w for w in all_words if w['x0'] > 315 and _re.match(r'^\d+[,\.]\d+$', w['text'])]
    name_words  = [w for w in all_words if w['x0'] < 250]

    def _clean_nom(nom):
        nom = _re.sub(r'\s+\d+[,\.]?\d*\s*cl\s+\d.*', '', nom, flags=_re.I)
        nom = _re.sub(r'\s+\d+[,\.]?\d+\s*[Ll]\s+\d.*', '', nom, flags=_re.I)
        nom = _re.sub(r'\s+\d+\.\d+[Ll]\b.*', '', nom)
        nom = _re.sub(r'\s+\d+[,\.]?\d*\s*[Ll]$', '', nom, flags=_re.I)
        nom = _re.sub(r'\s+\d+[,\.]\d+$', '', nom)
        nom = _re.sub(r'\s+(DONT|OFFERTE?S?).*$', '', nom, flags=_re.I)
        nom = _re.sub(r'\s*\d+%.*$', '', nom)
        nom = _re.sub(r'^AUCHAN\s+', '', nom, flags=_re.I)
        nom = _re.sub(r'€/?\s*[Ll]?\s*$', '', nom, flags=_re.I)
        nom = _re.sub(r'\s+\d+[Gg]\s+.*$', '', nom, flags=_re.I)
        nom = _re.sub(r'\s+RI\s+.*$', '', nom, flags=_re.I)
        return _re.sub(r'\s+', ' ', nom).strip()

    for qw in qty_words:
        y_q = qw['top']
        qty = int(qw['text'])
        if qty == 0 or y_q < 530:
            continue

        # Prix proches (±8px)
        near_prices = [w for w in price_words if abs(w['top'] - y_q) <= 8]
        near_prices.sort(key=lambda w: w['x0'])
        if len(near_prices) < 2:
            continue
        prix_ht = float(near_prices[0]['text'].replace(',', '.'))

        # Mots nom dans fenêtre y-15 à y+45
        nom_candidates = {}
        for nw in name_words:
            dy = nw['top'] - y_q
            if not (-15 <= dy <= 45):
                continue
            t = nw['text']
            if _re.match(r'^\d{13}$', t): continue
            if _re.match(r'^https?://', t): continue
            if _re.match(r'^\d{2}/\d{2}', t): continue
            if t in ('€', '€/', 'l', '%', ':', 'À', '!', '-', '/'): continue
            if _re.match(r'^\d{4}_', t): continue
            if _re.match(r'^(Bonjour|Votre|qu\'elle|Nous|SARROLA|Commande|Date|Récap|Prix|Libellé|Quantité|Montant|dont)', t, _re.I): continue
            # Garde les chiffres purs seulement s'ils sont dans la zone gauche du nom
            # (ex : "12" dans "SAINT GEORGES … 12 X33CL", "6" dans "S.PELLEGRINO 6 X 1L")
            if not any(c.isalpha() for c in t):
                if not (nw['x0'] < 235 and _re.match(r'^\d{1,3}$', t)):
                    continue
            ly = round(nw['top'] / 2) * 2
            if ly not in nom_candidates:
                nom_candidates[ly] = []
            nom_candidates[ly].append(nw)

        clean_lines = {}
        for ly, wds in nom_candidates.items():
            texts_here = [w['text'] for w in sorted(wds, key=lambda w: w['x0'])]
            joined = ' '.join(texts_here)
            if _re.match(r'^\d+[,\.]?\d*\s*[Ll]\s+\d+[,\.]', joined, _re.I): continue
            if _re.match(r'^\d+[,\.]?\d*\s*cl\s+\d+[,\.]', joined, _re.I): continue
            if _re.match(r'^\d+\.\d+[Ll]', joined) and any(c.isdigit() for c in joined[3:]): continue
            word_list = joined.split()
            # Filtre les lignes de catégorie (ex: "BOISSONS AUX FRUITS GAZEUSES")
            # SAUF si la ligne est sur la même rangée horizontale que la quantité
            # (ex: "SCHWEPPES AGRUMES BOITE SLIM" est sur la même ligne que qty=3)
            is_on_product_row = abs(ly - y_q) <= 3
            is_cat = (not is_on_product_row and
                      len(word_list) <= 4 and
                      not any(c.isdigit() for c in joined) and
                      '€' not in joined and '%' not in joined and
                      sum(1 for wt in word_list if wt.isupper() and len(wt) > 1) >= len(word_list) * 0.8)
            # Filtre aussi les en-têtes de catégorie Auchan avec volume
            # ex : "EAUX PLATES 1 LITRE", "EAUX GAZEUSES - 1 LITRE"
            is_vol_cat = (not is_on_product_row and len(word_list) <= 5 and
                          _re.match(r'^(EAUX|BIERES?|COLAS?|BOISSONS?|SPIRITUEUX|VINS?)\b',
                                    joined, _re.I))
            if is_cat or is_vol_cat: continue
            if any(c.isalpha() for c in joined):
                clean_lines[ly] = joined

        if not clean_lines:
            continue

        sorted_cl = sorted(clean_lines.items(), key=lambda kv: abs(kv[0] - y_q))[:2]
        sorted_cl.sort(key=lambda kv: kv[0])

        # Appliquer la détection de pack sur le nom BRUT (avant nettoyage)
        # afin que "6X 1.5L" (format Saint Georges) soit bien détecté même si
        # _clean_nom supprime ensuite le suffixe volume "1.5L".
        raw_nom = ' '.join(v for k, v in sorted_cl)
        nom = _clean_nom(raw_nom)
        if not nom or len(nom) < 3:
            continue
        # Ignore les entrées dont le nom n'est qu'un descripteur de volume/format
        # ex : "4X33CL" issu d'une ligne "3 OFFERTES" parasite
        _meaningful = [w for w in nom.split()
                       if len(w) >= 3 and not _re.match(r'^\d+[Xx]?\d*[CLcl]*$', w)
                       and w.upper() not in ('THE', 'LES', 'DES', 'SUR', 'AUX')]
        if not _meaningful:
            continue

        # Calcul quantité individuelle — regex sur le nom BRUT
        # Format NxVol : "6X33CL", "20X25CL", "12X20CL", "6X 1.5L", "24X33"
        m1 = _re.search(r'(\d+)\s*[Xx]\s*\d+[,\.]?\d*\s*(?:cl|CL|L\b|l\b)?', raw_nom)
        # Format VolxN : "20CL X12", "20CLX12"
        m2 = _re.search(r'\d+[,\.]?\d*\s*[Cc][Ll]\s*[Xx]\s*(\d+)', raw_nom)

        # Détermine pack_n pour calculer le prix par unité individuelle
        pack_n = 1
        if m2:
            pack_n = int(m2.group(1))
            quantite = qty * pack_n
        elif m1:
            n = int(m1.group(1))
            if n > 1:
                pack_n = n
                quantite = qty * n
            else:
                quantite = qty
        else:
            quantite = qty


        # Le BL Auchan affiche le prix du PACK (ex : 7,93€ pour 12 canettes).
        # On stocke toujours le prix par UNITÉ INDIVIDUELLE dans la base.
        prix_per_unit = round(prix_ht / pack_n, 6) if pack_n > 1 else prix_ht

        products_out.append({
            'nom': nom,
            'quantite': quantite,
            'prix_unitaire_ht': prix_per_unit,
            'numero_facture': order_num,
        })

    return products_out, order_num


@app.post("/api/import/livraison")
@app.post("/api/import/delivery/analyze")
async def analyze_delivery(file: UploadFile = File(...)):
    import os
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(400, detail="Fichier trop volumineux (max 10 Mo).")
    is_pdf = file.filename.lower().endswith(".pdf")
    # Pour les photos : réoriente selon EXIF (iPhone) avant envoi à l'IA
    if not is_pdf:
        content, _ = _apply_exif_orientation(content, "image/jpeg")

    # ── Pour les PDF : parsers déterministes (pdfplumber) ──
    if is_pdf:
        # 1. Essayer SOCOBO (factures avec caractères doublés)
        try:
            products, invoice_num = _parse_socobo_pdf(content)
            if products:
                return {"products": products, "fournisseur": "Socobo"}
        except Exception:
            pass

        # 2. Essayer PGV Distribution / Empreinte du Vin (PGFA / VVFA)
        try:
            products, invoice_num = _parse_pgv_vv_pdf(content)
            if products:
                return {"products": products, "fournisseur": "PGV Distribution"}
        except Exception:
            pass

        # 3. Essayer Esprit du Vin (factures SARL Esprit du Vin)
        try:
            products, invoice_num = _parse_esprit_du_vin_pdf(content)
            if products:
                return {"products": products, "fournisseur": "Esprit du Vin"}
        except Exception:
            pass

        # 4. Essayer Auchan (commandes Gmail/Auchan)
        try:
            products, order_num = _parse_auchan_pdf(content)
            if products:
                return {"products": products, "fournisseur": "Auchan"}
        except Exception:
            pass

    # ── Pour les photos (ou si pdfplumber n'a rien extrait) : Claude ──
    api_key = os.getenv("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        raise HTTPException(400, detail="Clé API Anthropic non configurée. Vérifiez le fichier .env")

    import anthropic as _anthropic
    client = _anthropic.Anthropic(api_key=api_key)

    prompt_text = """Tu analyses une facture ou un bon de livraison fournisseur.

TÂCHE 1 — Identifie le nom du fournisseur (ex : "Socobo", "PGV Distribution", "Auchan", "Esprit du Vin"…).
Si introuvable, laisse "fournisseur" vide.

TÂCHE 2 — Extrait les produits livrés :
- "nom" : nom complet du produit
- "quantite" : total d'unités individuelles (si NxVol ou N colis de X unités → quantite = N×X)
- "prix_unitaire_ht" : prix unitaire HT en euros (null si absent)
- "numero_facture" : numéro de facture/bon (même valeur pour toutes les lignes, "" si absent)

COLONNES typiques (format Socobo/DSAC) : LIBELLE | COLIS | CONTENANT | COLS | PRIX UNIT HT | REMISE | NET HT | VOLUME EFFECTIF | ALCOOL PUR
⚠️ VOLUME EFFECTIF et ALCOOL PUR = colonnes fiscales en LITRES — NE JAMAIS les utiliser comme quantités.

LIGNES À IGNORER : FRAIS DE REGIE, DECONSIGNE, CONSIGNE, FUT 30 EUROS.
PROMO FOURN / GRATUIT : ajoute les COLIS au produit principal.

Réponds UNIQUEMENT en JSON valide :
{"fournisseur": "Socobo", "produits": [{"nom": "Pastis 51 1L", "quantite": 12, "prix_unitaire_ht": 16.84, "numero_facture": "100051"}]}"""

    try:
        if is_pdf:
            b64 = base64.standard_b64encode(content).decode("utf-8")
            message = client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=4096,
                messages=[{"role": "user", "content": [
                    {"type": "document",
                     "source": {"type": "base64", "media_type": "application/pdf", "data": b64}},
                    {"type": "text", "text": prompt_text}
                ]}],
            )
        else:
            ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else "jpg"
            media_map = {"jpg": "image/jpeg", "jpeg": "image/jpeg",
                         "png": "image/png", "webp": "image/webp",
                         "gif": "image/gif", "heic": "image/jpeg"}
            media_type = media_map.get(ext, "image/jpeg")
            b64 = base64.standard_b64encode(content).decode("utf-8")
            message = client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=4096,
                messages=[{"role": "user", "content": [
                    {"type": "image",
                     "source": {"type": "base64", "media_type": media_type, "data": b64}},
                    {"type": "text", "text": prompt_text}
                ]}],
            )
    except _anthropic.BadRequestError as e:
        raise HTTPException(400, detail=f"Fichier non lisible : {str(e)}")
    except _anthropic.AuthenticationError:
        raise HTTPException(401, detail="Clé API invalide. Vérifiez le fichier .env")
    except Exception as e:
        raise HTTPException(500, detail=f"Erreur lors de l'analyse : {str(e)}")

    raw = message.content[0].text.strip()

    # Essayer d'abord le format objet {"fournisseur": "...", "produits": [...]}
    fournisseur_ia = ""
    products = None
    obj_start = raw.find("{")
    if obj_start != -1:
        obj_end = raw.rfind("}") + 1
        try:
            parsed = json.loads(raw[obj_start:obj_end])
            if isinstance(parsed, dict) and ("produits" in parsed or "products" in parsed):
                products = parsed.get("produits") or parsed.get("products") or []
                fournisseur_ia = parsed.get("fournisseur", "")
        except json.JSONDecodeError:
            pass

    # Fallback : ancien format tableau [...]
    if products is None:
        start = raw.find("[")
        end = raw.rfind("]") + 1
        if start == -1:
            raise HTTPException(400, detail="Impossible d'extraire les données du document. Essayez avec une meilleure photo.")
        try:
            products = json.loads(raw[start:end])
        except json.JSONDecodeError:
            raise HTTPException(400, detail="Le document n'a pas pu être analysé correctement.")

    return {"products": products, "fournisseur": fournisseur_ia}


class DeliveryConfirmIn(BaseModel):
    products: list
    numero_facture: str
    fournisseur: str = ""


def _normalize_tokens(s: str) -> set:
    """Normalise un nom produit en tokens pour matching fuzzy."""
    import unicodedata, re
    # Supprimer accents
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = s.lower()
    # Normaliser "saint" → "st" pour matcher "Saint Georges" ↔ "St Georges"
    s = re.sub(r'\bsaint\b', 'st', s)
    # Synonymes marques : "lipton" → "ice tea" (BL Auchan vs catalogue)
    s = re.sub(r'\blipton\b', 'ice tea', s)
    # Synonymes variantes : "sans sucre" / "low sucres" / "zero sucre" → "zero"
    s = re.sub(r'\bsans\s+sucre\b', 'zero', s)
    s = re.sub(r'\blow\s+sucres?\b', 'zero', s)
    s = re.sub(r'\bzero\s+sucre\b', 'zero', s)
    s = re.sub(r'\bsugar\s+free\b', 'zero', s)
    # Normaliser décimales : "1,5" "1.5" → "15"
    s = re.sub(r'(\d)[,.](\d)', r'\1\2', s)
    # Supprimer multiplicateurs de pack : "6x", "12x", "x6", "33clx6" → "33cl"
    s = re.sub(r'(\d+(?:cl|l))x\d+', r'\1', s)
    s = re.sub(r'\b\d+x\b', ' ', s)   # "6x", "12x"
    s = re.sub(r'\bx\d+\b', ' ', s)   # "x6", "x12"
    # Supprimer "/24" après cl : "33cl/24" → "33cl"
    s = re.sub(r'(\d+cl)/\d+', r'\1', s)
    # Supprimer caractères non-alphanumériques
    s = re.sub(r'[^\w\s]', ' ', s)
    # Mots parasites à ignorer (descripteurs génériques + articles)
    STOP = {'vp','bte','pet','bt','pres','purjus','pur','abc','slim','nectar',
            'pack','carton','lot','de','le','la','les','du','des','et','en',
            'un','une','fut','fut','lx6','lx4','lx12','lx24','bionda',
            # descripteurs eau generiques
            'eau','source','gazeuse','plate','minerale','naturelle','drinking',
            # suffixes pack/format
            'bouteille','bouteilles','canette','canettes','boite','boites','brik'}
    return {t for t in s.split() if len(t) >= 2 and t not in STOP}


def _find_best_product(nom: str, all_products) -> object:
    """Matching intelligent multi-stratégies pour retrouver un produit depuis un nom OCR."""
    from difflib import SequenceMatcher
    import unicodedata, re

    def no_accent(s):
        s = unicodedata.normalize('NFD', s)
        return ''.join(c for c in s if unicodedata.category(c) != 'Mn').lower()

    nom_clean = no_accent(nom)

    # Stratégie 1 : exact
    for p in all_products:
        if no_accent(p.name) == nom_clean:
            return p

    # Stratégie 2 : contient (bidirectionnel)
    for p in all_products:
        pn = no_accent(p.name)
        if pn in nom_clean or nom_clean in pn:
            return p

    # Stratégie 3 : tokens
    import re as _re
    ocr_tokens = _normalize_tokens(nom)

    # Extraire le volume du nom OCR (ex: "75cl", "33cl", "1l")
    ocr_vols = set(_re.findall(r'\d+cl|\d+l\b', no_accent(nom)))

    # Descripteurs génériques qui ne suffisent PAS à identifier un produit seuls
    # (saveur, couleur, style) — évite "Lipton Pêche" → "Pago Pêche" via token "peche"
    GENERIC_DESCRIPTORS = {
        'peche', 'citron', 'menthe', 'fraise', 'framboise', 'cerise',
        'pomme', 'raisin', 'mangue', 'grenade', 'tropical', 'agrumes',
        'rouge', 'blanc', 'rose', 'vert', 'noir', 'bleu',
        'light', 'zero', 'original', 'classic', 'nature', 'special',
        'premium', 'extra', 'brut', 'sec', 'doux', 'vieux',
    }

    best_p, best_score = None, 0.0
    for p in all_products:
        db_tokens = _normalize_tokens(p.name)
        if not db_tokens:
            continue
        common = ocr_tokens & db_tokens
        score = len(common) / len(db_tokens)

        # Pénalité si volumes incompatibles (ex: "75cl" dans OCR mais "33cl" dans DB)
        if ocr_vols:
            db_vols = set(_re.findall(r'\d+cl|\d+l\b', no_accent(p.name)))
            if db_vols and not ocr_vols & db_vols:
                score *= 0.4  # pénalité forte si aucun volume en commun

        # Pénalité si le seul token commun est un descripteur générique (saveur, couleur…)
        # Ex : "LIPTON PECHE SLEEK" ne doit PAS matcher "Pago Pêche" via "peche" seul
        if len(common) == 1 and common <= GENERIC_DESCRIPTORS:
            score *= 0.25

        # Pénalité croisée sur les marqueurs de variante (zero, light, diet…)
        # Si l'OCR a "zero" mais le produit DB n'a pas "zero" (ou vice-versa) → mauvais match
        VARIANT_MARKERS = {'zero', 'light', 'diet', 'sans'}
        ocr_variants = ocr_tokens & VARIANT_MARKERS
        db_variants  = db_tokens  & VARIANT_MARKERS
        if ocr_variants and not (ocr_variants & db_variants):
            score *= 0.1   # OCR = variante, DB = produit normal → très mauvais
        elif db_variants and not (ocr_variants & db_variants):
            score *= 0.1   # DB = variante, OCR = produit normal → très mauvais

        if score > best_score:
            best_score = score
            best_p = p
    if best_score >= 0.5:
        return best_p

    # Stratégie 4 : difflib
    best_p, best_ratio = None, 0.0
    for p in all_products:
        ratio = SequenceMatcher(None, nom_clean, no_accent(p.name)).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_p = p
    if best_ratio >= 0.55:
        return best_p

    return None


@app.post("/api/import/livraison/confirm")
@app.post("/api/import/delivery/confirm")
def confirm_delivery(body: DeliveryConfirmIn, db: Session = Depends(get_db)):
    import traceback
    try:
        return _confirm_delivery_inner(body, db)
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(500, detail=f"Bug: {type(e).__name__}: {str(e)} | {traceback.format_exc()[-300:]}")

def _confirm_delivery_inner(body: DeliveryConfirmIn, db: Session):
    existing = db.query(ImportLog).filter(
        ImportLog.import_type == "delivery",
        ImportLog.reference == body.numero_facture,
    ).first()
    if existing:
        raise HTTPException(
            400,
            detail=f"Facture n°{body.numero_facture} déjà importée le {to_local(existing.created_at).strftime('%d/%m/%Y %H:%M')}."
        )

    all_products = db.query(Product).all()

    # Résoudre le fournisseur (string → objet Supplier)
    supplier_obj = None
    if body.fournisseur:
        supplier_obj = db.query(Supplier).filter(
            func.lower(Supplier.name) == body.fournisseur.strip().lower()
        ).first()
        # Fallback : recherche partielle (ex: "Auchan Bastia" → "Auchan")
        if not supplier_obj:
            fnom = body.fournisseur.strip().lower()
            supplier_obj = db.query(Supplier).filter(
                func.lower(Supplier.name).contains(fnom)
            ).first()
        if not supplier_obj:
            fnom = body.fournisseur.strip().lower()
            for s in db.query(Supplier).all():
                if fnom in s.name.lower() or s.name.lower() in fnom:
                    supplier_obj = s
                    break

    # 1. Résoudre les correspondances et agréger les quantités par produit trouvé
    matched_map = {}   # product_id → {product, total_qty, prix}
    not_found = []

    for item in body.products:
        nom = item.get("nom", "").strip()
        qty = float(item.get("quantite", 0) or 0)
        prix = item.get("prix_unitaire_ht")
        matched_id = item.get("produit_id") or item.get("product_id")

        p = None
        if matched_id:
            p = db.query(Product).get(int(matched_id))
        if not p:
            p = _find_best_product(nom, all_products)

        if p and qty != 0:
            if p.id not in matched_map:
                matched_map[p.id] = {"product": p, "total_qty": 0.0, "prix": None}
            matched_map[p.id]["total_qty"] += qty
            # Ne met à jour le prix que pour les livraisons (qty > 0)
            if qty > 0 and prix is not None and prix != "" and float(prix) > 0:
                matched_map[p.id]["prix"] = float(prix)
        elif qty != 0:
            not_found.append(nom)

    # 2. Appliquer les quantités en tenant compte du conditionnement
    updated = []
    is_avoir = False  # détecte si c'est un avoir (retour fournisseur)

    for pid, entry in matched_map.items():
        p = entry["product"]
        raw_qty = entry["total_qty"]
        prix = entry["prix"]

        # Le stock est toujours en unités individuelles (bouteilles/canettes).
        # qty_per_pack sert uniquement à l'affichage (équivalent cartons).
        actual_qty = raw_qty

        if actual_qty != 0:
            if actual_qty < 0:
                is_avoir = True
            old_price = p.purchase_price
            p.stock = round(p.stock + actual_qty, 4)
            # Ne met à jour le prix que pour les entrées de stock (livraisons)
            if prix is not None and prix > 0 and actual_qty > 0:
                record_price_change(
                    db, p, prix,
                    source="delivery_import",
                    supplier_id=supplier_obj.id if supplier_obj else None,
                    reference=body.numero_facture or "",
                )
                p.purchase_price = prix
                p.is_estimated = False
                # Upsert dans product_suppliers pour ce fournisseur
                if supplier_obj:
                    ps = db.query(ProductSupplier).filter(
                        ProductSupplier.product_id == p.id,
                        ProductSupplier.supplier_id == supplier_obj.id,
                    ).first()
                    if ps:
                        ps.purchase_price = prix
                        ps.updated_at = datetime.utcnow()
                    else:
                        # Premier lien avec ce fournisseur : is_primary si aucun autre
                        is_first = db.query(ProductSupplier).filter(
                            ProductSupplier.product_id == p.id
                        ).count() == 0
                        db.add(ProductSupplier(
                            product_id=p.id,
                            supplier_id=supplier_obj.id,
                            purchase_price=prix,
                            is_primary=is_first,
                        ))
                    # Si c'est le fournisseur principal, sync supplier_id legacy
                    primary_ps = db.query(ProductSupplier).filter(
                        ProductSupplier.product_id == p.id,
                        ProductSupplier.is_primary == True,
                    ).first()
                    if primary_ps and primary_ps.supplier_id == supplier_obj.id:
                        p.supplier_id = supplier_obj.id
            updated.append({
                "product_id": p.id,
                "product": p.name,
                "added": actual_qty,
                "old_price": old_price,
                "new_price": p.purchase_price,
            })

    # Ne créer le log que si quelque chose a été traité
    if updated:
        event_type = "avoir_fournisseur" if is_avoir else "livraison"
        label = "Avoir/retour fournisseur" if is_avoir else "Bon de livraison"
        import_log = ImportLog(
            import_type="delivery",
            reference=body.numero_facture,
            supplier=body.fournisseur,
            details_json=json.dumps({"items": updated, "not_found": not_found}, ensure_ascii=False),
        )
        db.add(import_log)
        log_event(
            db,
            event_type,
            f"{label} n°{body.numero_facture} — {len(updated)} produit(s)",
            {
                "numero_facture": body.numero_facture,
                "fournisseur": body.fournisseur,
                "updated": [{"product": u["product"], "added": u["added"]} for u in updated],
                "not_found": not_found,
            }
        )
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(500, detail=f"Erreur base de données : {str(e)}")
    return {"ok": True, "updated": updated, "not_found": not_found, "is_avoir": is_avoir}


def _parse_import_details(raw_json: str):
    """Parse details_json → (is_annule, items, not_found).
    Gère tous les formats : liste plate (ancien), dict avec items/not_found (nouveau),
    dict avec annule (annulé)."""
    try:
        raw = json.loads(raw_json or '[]')
    except Exception:
        return False, [], []
    if isinstance(raw, list):
        return False, raw, []
    if isinstance(raw, dict):
        is_annule = bool(raw.get('annule'))
        items = raw.get('items', raw.get('details', []))
        not_found = raw.get('not_found', [])
        return is_annule, items, not_found
    return False, [], []


@app.get("/api/imports/{import_id}/detail")
def import_detail(import_id: int, db: Session = Depends(get_db)):
    imp = db.query(ImportLog).filter(ImportLog.id == import_id).first()
    if not imp:
        raise HTTPException(404, detail="Import introuvable")
    is_annule, items, not_found = _parse_import_details(getattr(imp, "details_json", "[]"))
    return {
        "id": imp.id,
        "reference": imp.reference,
        "supplier": imp.supplier,
        "created_at": to_local(imp.created_at).strftime("%d/%m/%Y %H:%M"),
        "annule": is_annule,
        "details": items,
        "not_found": not_found,
    }


@app.get("/api/imports/recent")
def recent_imports(days: int = 7, db: Session = Depends(get_db)):
    from datetime import timedelta
    cutoff = datetime.utcnow() - timedelta(days=days)
    imports = db.query(ImportLog).filter(
        ImportLog.import_type == "delivery",
        ImportLog.created_at >= cutoff,
    ).order_by(ImportLog.created_at.desc()).all()

    result = []
    for imp in imports:
        is_annule, items, _ = _parse_import_details(getattr(imp, 'details_json', '[]'))
        result.append({
            "id": imp.id,
            "reference": imp.reference,
            "supplier": imp.supplier or "",
            "created_at": to_local(imp.created_at).strftime("%d/%m/%Y %H:%M"),
            "annule": is_annule,
            "nb_produits": len(items),
            "details": items,
        })
    return result


@app.delete("/api/imports/{import_id}/lines/{product_id}")
def delete_import_line(import_id: int, product_id: int, db: Session = Depends(get_db)):
    imp = db.query(ImportLog).get(import_id)
    if not imp or imp.import_type != "delivery":
        raise HTTPException(404, "Import introuvable")

    is_annule, items, not_found = _parse_import_details(getattr(imp, 'details_json', '[]'))
    if is_annule:
        raise HTTPException(400, "Cet import a été annulé, modification impossible")

    line = next((x for x in items if x.get('product_id') == product_id), None)
    if not line:
        raise HTTPException(404, "Ligne introuvable dans cet import")

    p = db.query(Product).get(product_id)
    if p:
        p.stock = round(p.stock - line['added'], 3)
        old_price = line.get('old_price')
        if old_price is not None:
            p.purchase_price = old_price

    new_items = [x for x in items if x.get('product_id') != product_id]
    imp.details_json = json.dumps({"items": new_items, "not_found": not_found}, ensure_ascii=False)

    log_event(db, "modif_livraison",
              f"Suppression ligne BL n°{imp.reference} — {line.get('product', '')}",
              {"import_id": import_id, "product_id": product_id, "removed_qty": line['added']})
    db.commit()
    return {"ok": True, "removed": line.get('product', ''), "qty": line['added']}


class ImportLineIn(BaseModel):
    new_qty: float
    new_price: Optional[float] = None


@app.patch("/api/imports/{import_id}/lines/{product_id}")
def update_import_line(import_id: int, product_id: int, body: ImportLineIn, db: Session = Depends(get_db)):
    imp = db.query(ImportLog).get(import_id)
    if not imp or imp.import_type != "delivery":
        raise HTTPException(404, "Import introuvable")

    is_annule, items, not_found = _parse_import_details(getattr(imp, 'details_json', '[]'))
    if is_annule:
        raise HTTPException(400, "Cet import a été annulé, modification impossible")

    line = next((x for x in items if x.get('product_id') == product_id), None)
    if not line:
        raise HTTPException(404, "Ligne introuvable dans cet import")

    old_added = line['added']
    delta = body.new_qty - old_added

    p = db.query(Product).get(product_id)
    if p:
        p.stock = round(p.stock + delta, 3)
        if body.new_price is not None:
            p.purchase_price = body.new_price
            p.is_estimated = False
            if imp.supplier:
                sup = db.query(Supplier).filter(
                    func.lower(Supplier.name) == imp.supplier.lower()
                ).first()
                if sup:
                    ps = db.query(ProductSupplier).filter_by(
                        product_id=product_id, supplier_id=sup.id
                    ).first()
                    if ps:
                        ps.purchase_price = body.new_price

    line['added'] = body.new_qty
    if body.new_price is not None:
        line['new_price'] = body.new_price
    imp.details_json = json.dumps({"items": items, "not_found": not_found}, ensure_ascii=False)

    log_event(db, "modif_livraison",
              f"Modification ligne BL n°{imp.reference} — {line.get('product', '')}",
              {"import_id": import_id, "product_id": product_id,
               "old_qty": old_added, "new_qty": body.new_qty, "new_price": body.new_price})
    db.commit()
    return {"ok": True, "product": line.get('product', ''), "old_qty": old_added, "new_qty": body.new_qty}


@app.post("/api/imports/{import_id}/annuler")
def annuler_import(import_id: int, db: Session = Depends(get_db)):
    from datetime import timedelta
    imp = db.query(ImportLog).get(import_id)
    if not imp or imp.import_type != "delivery":
        raise HTTPException(404, "Import introuvable")

    try:
        details = json.loads(getattr(imp, 'details_json', '[]') or '[]')
    except Exception:
        details = []

    is_annule, items, not_found = _parse_import_details(getattr(imp, 'details_json', '[]'))
    if is_annule:
        raise HTTPException(400, "Cet import a déjà été annulé")

    if datetime.utcnow() - imp.created_at > timedelta(days=7):
        raise HTTPException(400, "Annulation impossible après 7 jours")

    reversed_items = []
    for item in items:
        p = db.query(Product).get(item.get('product_id'))
        if not p:
            continue
        p.stock = round(p.stock - item['added'], 3)
        old_price = item.get('old_price')
        if old_price is not None:
            p.purchase_price = old_price
            p.is_estimated = old_price is None
        reversed_items.append({"product": p.name, "removed": item['added']})

    # Marquer comme annulé en conservant not_found
    imp.details_json = json.dumps({"annule": True, "items": items, "not_found": not_found}, ensure_ascii=False)

    log_event(db, "annulation_livraison",
              f"Annulation BL n°{imp.reference} ({imp.supplier})",
              {"numero_facture": imp.reference, "fournisseur": imp.supplier, "reversed": reversed_items})
    db.commit()
    return {"ok": True, "reversed": reversed_items}


# ══════════════════════════════════════════════════════════════════════════
# ADMINISTRATION — RESET / PURGE
# ══════════════════════════════════════════════════════════════════════════

class AdminActionIn(BaseModel):
    pin: str


def _verify_admin_pin(pin: str):
    """Lève une HTTPException si le PIN est incorrect."""
    manager_pin = os.environ.get("MANAGER_PIN", "1234")
    if pin != manager_pin:
        raise HTTPException(401, "Code PIN incorrect")


@app.post("/api/admin/reset-stocks")
def admin_reset_stocks(body: AdminActionIn, db: Session = Depends(get_db)):
    """Remet tous les stocks à 0 — nécessite le PIN gérant."""
    _verify_admin_pin(body.pin)
    count = db.query(Product).count()
    db.query(Product).update({"stock": 0})
    log_event(db, "admin_reset_stocks",
              f"Remise à zéro de tous les stocks ({count} produits)", {})
    db.commit()
    return {"ok": True, "products_reset": count}


class AdminClearImportsIn(BaseModel):
    pin: str
    include_validated: bool = False   # si True, supprime aussi les validés ET rollback le stock


@app.delete("/api/admin/clear-imports")
def admin_clear_imports(body: AdminClearImportsIn, db: Session = Depends(get_db)):
    """
    Supprime tous les contrôles de livraison + l'historique des imports.
    - Par défaut : supprime les pending_count / counted / partial (aucun impact stock).
    - include_validated=True : supprime aussi les validés, en décrémentant le stock.
    Requiert le PIN gérant.
    """
    _verify_admin_pin(body.pin)

    # Rollback stock pour les validés si demandé
    rolled_back = 0
    if body.include_validated:
        for c in db.query(DeliveryCheck).filter(DeliveryCheck.status == "validated").all():
            for it in c.items:
                if it.stock_applied and it.qty_validated and it.product_id:
                    p = db.query(Product).get(it.product_id)
                    if p:
                        p.stock = max(0, (p.stock or 0) - it.qty_validated)
                        rolled_back += 1

    # Supprimer les DeliveryChecks (les items sont cascade)
    q = db.query(DeliveryCheck)
    if not body.include_validated:
        q = q.filter(DeliveryCheck.status != "validated")
    dc_count = q.count()
    q.delete(synchronize_session=False)

    # Supprimer l'historique des imports IA (livraisons)
    imp_count = db.query(ImportLog).filter(ImportLog.import_type == "delivery").count()
    db.query(ImportLog).filter(ImportLog.import_type == "delivery").delete(synchronize_session=False)

    log_event(
        db, "admin_clear_imports",
        f"Suppression de {dc_count} contrôle(s) de livraison + {imp_count} import(s) IA"
        + (f" (rollback stock sur {rolled_back} items)" if rolled_back else ""),
        {"delivery_checks": dc_count, "import_logs": imp_count, "stock_rollbacks": rolled_back},
    )
    db.commit()
    return {"ok": True, "delivery_checks_deleted": dc_count, "import_logs_deleted": imp_count, "stock_rollbacks": rolled_back}


# ── Snapshots stock (pour démarque auto hebdomadaire) ─────────────────────
@app.post("/api/stock-snapshots")
def create_stock_snapshot(label: str = "user", db: Session = Depends(get_db)):
    """
    Enregistre une photo du stock actuel de tous les produits non archivés.
    Appelé manuellement ou par le scheduler hebdo.
    """
    products = db.query(Product).filter(
        (Product.archived == False) | (Product.archived == None)
    ).all()
    now = datetime.utcnow()
    for p in products:
        db.add(StockSnapshot(
            product_id=p.id,
            taken_at=now,
            stock=p.stock or 0,
            label=label,
        ))
    db.commit()
    return {"ok": True, "snapshots": len(products), "taken_at": now.isoformat() + "Z"}


@app.get("/api/demarque/auto")
def get_demarque_auto(db: Session = Depends(get_db)):
    """
    Calcule la démarque théorique vs réelle depuis le dernier snapshot
    de type 'weekly_auto' (à défaut, le dernier snapshot tout label).

    perte_théorique = stock_snapshot + livraisons − ventes − pertes_declarées − stock_actuel

    Retourne la liste des produits avec un écart > 0 (perte potentielle non déclarée).
    """
    # Trouver le snapshot de référence : weekly_auto le plus récent, ou fallback
    latest_dt_row = (
        db.query(StockSnapshot.taken_at)
        .filter(StockSnapshot.label == "weekly_auto")
        .order_by(StockSnapshot.taken_at.desc())
        .first()
    )
    if not latest_dt_row:
        latest_dt_row = (
            db.query(StockSnapshot.taken_at)
            .order_by(StockSnapshot.taken_at.desc())
            .first()
        )
    if not latest_dt_row:
        return {
            "has_data": False,
            "message": "Aucun snapshot encore enregistré. Créez-en un pour démarrer le suivi.",
        }

    snapshot_dt = latest_dt_row[0]
    # Toutes les lignes de ce snapshot (il y en a plusieurs à la même date)
    snapshots = (
        db.query(StockSnapshot)
        .filter(StockSnapshot.taken_at == snapshot_dt)
        .all()
    )
    stock_at_snapshot = {s.product_id: s.stock for s in snapshots}

    # Ventes entre snapshot_dt et maintenant (via extract_sales sur import_cashpad)
    sales_hist = (
        db.query(StockHistory)
        .filter(
            StockHistory.event_type.in_(["import_cashpad", "cashpad_sync"]),
            StockHistory.created_at >= snapshot_dt,
        )
        .all()
    )
    # On utilise le champ "deductions" (qty_effectivement_déduite du stock)
    ventes_qty = {}   # product_id → total deduction
    for h in sales_hist:
        try:
            data = json.loads(h.data_json or "{}")
        except Exception:
            continue
        if isinstance(data, dict):
            for ded in data.get("deductions", []):
                pid = ded.get("product_id")
                qty = abs(float(ded.get("quantity", ded.get("deducted", 0)) or 0))
                if pid:
                    ventes_qty[pid] = ventes_qty.get(pid, 0) + qty

    # Livraisons (imports BL) depuis le snapshot
    livraisons_hist = (
        db.query(StockHistory)
        .filter(
            StockHistory.event_type == "livraison",
            StockHistory.created_at >= snapshot_dt,
        )
        .all()
    )
    livraisons_qty = {}
    for h in livraisons_hist:
        try:
            data = json.loads(h.data_json or "{}")
            items = data.get("items") or data.get("lines") or data.get("deductions") or []
            for it in items:
                pid = it.get("product_id")
                qty = float(it.get("qty") or it.get("quantity") or it.get("added") or 0)
                if pid and qty > 0:
                    livraisons_qty[pid] = livraisons_qty.get(pid, 0) + qty
        except Exception:
            pass

    # Mouvements manuels (+/-)
    manuel_hist = (
        db.query(StockHistory)
        .filter(
            StockHistory.event_type == "mouvement_manuel",
            StockHistory.created_at >= snapshot_dt,
        )
        .all()
    )
    manuel_net = {}   # product_id → delta net (positif = ajout, négatif = retrait)
    for h in manuel_hist:
        try:
            data = json.loads(h.data_json or "{}")
            pid = data.get("product_id")
            qty = float(data.get("quantity") or data.get("qty") or 0)
            if pid:
                manuel_net[pid] = manuel_net.get(pid, 0) + qty
        except Exception:
            pass

    # Pertes déclarées (ManualLoss) depuis le snapshot
    losses = (
        db.query(ManualLoss)
        .filter(ManualLoss.date >= snapshot_dt)
        .all()
    )
    pertes_qty = {}
    for l in losses:
        pertes_qty[l.product_id] = pertes_qty.get(l.product_id, 0) + abs(float(l.quantity or 0))

    # Calcul final par produit
    products = db.query(Product).filter(
        (Product.archived == False) | (Product.archived == None)
    ).all()
    items = []
    total_lost_eur = 0.0
    total_lost_qty = 0.0
    for p in products:
        if p.id not in stock_at_snapshot:
            continue
        s_start = stock_at_snapshot[p.id]
        s_end   = p.stock or 0
        v  = ventes_qty.get(p.id, 0)
        l  = livraisons_qty.get(p.id, 0)
        m  = manuel_net.get(p.id, 0)
        pe = pertes_qty.get(p.id, 0)

        theoretical_end = s_start + l + m - v - pe
        ecart = theoretical_end - s_end   # > 0 → manque, produit "disparu"

        if abs(ecart) < 0.05:
            continue   # rien de notable

        price = p.purchase_price or 0
        loss_eur = round(ecart * price, 2) if ecart > 0 else 0
        if ecart > 0:
            total_lost_eur += loss_eur
            total_lost_qty += ecart
        items.append({
            "product_id":    p.id,
            "product_name":  p.name,
            "category":      p.category,
            "stock_start":   round(s_start, 3),
            "stock_end":     round(s_end, 3),
            "ventes":        round(v, 3),
            "livraisons":    round(l, 3),
            "manuel_net":    round(m, 3),
            "pertes_declarees": round(pe, 3),
            "theoretical_end":  round(theoretical_end, 3),
            "ecart":         round(ecart, 3),
            "loss_eur":      loss_eur,
            "unit":          p.unit or "u",
        })

    # Tri : plus grosses pertes en €
    items.sort(key=lambda x: -x["loss_eur"])

    return {
        "has_data":      True,
        "snapshot_date": snapshot_dt.isoformat() + "Z",
        "snapshot_age_days": (datetime.utcnow() - snapshot_dt).days,
        "items":         items,
        "total_lost_eur": round(total_lost_eur, 2),
        "total_lost_qty": round(total_lost_qty, 2),
        "n_products_with_loss": sum(1 for i in items if i["ecart"] > 0),
    }


class SeasonResetIn(BaseModel):
    pin: str
    confirmation: str   # doit valoir exactement "RESET"


@app.post("/api/admin/reset-season")
def admin_reset_season(body: SeasonResetIn, db: Session = Depends(get_db)):
    """
    RÉINITIALISATION SAISON — purge les données de rodage.
    Supprime : imports Cashpad (StockHistory + ImportLog), mouvements manuels,
    alertes inventaire, ventes historiques, session inventaires, service alerts.
    Remet les stocks à 0.
    Conserve : produits, cocktails, fournisseurs, mappings Cashpad, événements,
    objectif saison, commandes fournisseurs.
    """
    _verify_admin_pin(body.pin)
    if (body.confirmation or "").strip() != "RESET":
        raise HTTPException(400, "Confirmation invalide — tapez RESET en majuscules.")

    stats = {}

    # 1. StockHistory — purger tous les événements "transactionnels"
    event_types_to_purge = [
        "import_cashpad", "cashpad_sync", "mouvement_manuel",
        "alerte_inventaire", "ventes_historiques", "controle_flash",
        "livraison", "avoir_fournisseur",
    ]
    stats["stock_history_deleted"] = (
        db.query(StockHistory)
        .filter(StockHistory.event_type.in_(event_types_to_purge))
        .delete(synchronize_session=False)
    )

    # 2. ImportLog (Cashpad + livraisons)
    stats["import_logs_deleted"] = (
        db.query(ImportLog).delete(synchronize_session=False)
    )

    # 3. Sessions inventaires (anciens inventaires flash)
    stats["inventory_sessions_deleted"] = (
        db.query(InventorySession).delete(synchronize_session=False)
    )

    # 4. Service alerts
    stats["service_alerts_deleted"] = (
        db.query(ServiceAlert).delete(synchronize_session=False)
    )

    # 5. Pertes manuelles (casse, vol déclarés pendant rodage)
    stats["manual_losses_deleted"] = (
        db.query(ManualLoss).delete(synchronize_session=False)
    )

    # 6. Remettre stocks à 0
    stats["products_reset"] = db.query(Product).count()
    db.query(Product).update({"stock": 0})

    # 7. Reset settings Cashpad
    for key in ["cashpad_last_sync", "cashpad_last_sequential_id",
                "weather_cache", "weather_cache_ts"]:
        db.query(AppSetting).filter(AppSetting.key == key).delete(synchronize_session=False)

    # Log la purge elle-même (reste dans l'historique comme marqueur)
    log_event(db, "admin_reset_season",
              f"Réinitialisation saison effectuée — purge complète", stats)
    db.commit()

    return {"ok": True, **stats}


# ══════════════════════════════════════════════════════════════════════════
# STATISTIQUES DE CONSOMMATION
# ══════════════════════════════════════════════════════════════════════════

@app.get("/api/stats/consommation")
def stats_consommation(periode: int = 30, db: Session = Depends(get_db)):
    """
    Agrège les consommations sur N jours (sorties réserve + Cashpad).
    Retourne : top produits, par catégorie, évolution journalière.
    """
    cutoff = datetime.utcnow() - timedelta(days=periode)
    rows = db.query(StockHistory).filter(
        StockHistory.created_at >= cutoff,
        StockHistory.event_type.in_(["mouvement_manuel", "import_cashpad"])
    ).order_by(StockHistory.created_at).all()

    # Récupère les infos produits pour la catégorie
    products_map = {p.id: p for p in db.query(Product).all()}

    by_product = {}   # product_name → {qty, category, product_id}
    by_day = {}       # "YYYY-MM-DD" → qty totale

    for row in rows:
        try:
            d = json.loads(row.data_json)
        except Exception:
            continue

        day = to_local(row.created_at).strftime("%Y-%m-%d")

        if row.event_type == "mouvement_manuel":
            qty = d.get("quantity", 0)
            if qty >= 0:
                continue  # ignore les entrées, seulement les sorties
            qty = abs(qty)
            name = d.get("product", "?")
            pid = d.get("product_id")
            cat = products_map[pid].category if pid and pid in products_map else "Autres"
            key = name
            if key not in by_product:
                by_product[key] = {"qty": 0, "category": cat, "product_id": pid}
            by_product[key]["qty"] += qty
            by_day[day] = by_day.get(day, 0) + qty

        elif row.event_type == "import_cashpad":
            deductions = d.get("deductions", [])
            for ded in deductions:
                qty = abs(float(ded.get("quantity", ded.get("qty", 0)) or 0))
                if qty <= 0:
                    continue
                name = ded.get("product", ded.get("nom", "?"))
                pid = ded.get("product_id")
                cat = products_map[pid].category if pid and pid in products_map else "Autres"
                key = name
                if key not in by_product:
                    by_product[key] = {"qty": 0, "category": cat, "product_id": pid}
                by_product[key]["qty"] += qty
                by_day[day] = by_day.get(day, 0) + qty

    # Top produits triés
    top_products = sorted(
        [{"name": k, "qty": round(v["qty"], 2), "category": v["category"]} for k, v in by_product.items()],
        key=lambda x: x["qty"], reverse=True
    )

    # Par catégorie
    by_cat = {}
    for item in top_products:
        c = item["category"]
        by_cat[c] = round(by_cat.get(c, 0) + item["qty"], 2)
    by_category = sorted([{"category": k, "qty": v} for k, v in by_cat.items()], key=lambda x: x["qty"], reverse=True)

    # Évolution journalière (tous les jours de la période)
    from datetime import date, timedelta as td
    daily = []
    for i in range(periode):
        d_str = (date.today() - td(days=periode - 1 - i)).strftime("%Y-%m-%d")
        daily.append({"date": d_str, "qty": round(by_day.get(d_str, 0), 2)})

    return {
        "periode": periode,
        "top_products": top_products[:20],
        "by_category": by_category,
        "daily": daily,
        "total": round(sum(x["qty"] for x in top_products), 2),
    }


@app.post("/api/stats/import-historique")
async def import_historique(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Importe un fichier Cashpad Synthèse (Excel) pour alimenter les stats historiques.
    Utilise la feuille 'Ventes-par-produit-par-date-4' (données mensuelles).
    """
    import openpyxl, io as _io
    content = await file.read()
    wb = openpyxl.load_workbook(_io.BytesIO(content), data_only=True)

    # Cherche la feuille par-date
    sheet_name = None
    for s in wb.sheetnames:
        if "par-date" in s.lower() or "date" in s.lower():
            sheet_name = s
            break
    if not sheet_name:
        raise HTTPException(400, "Feuille 'par date' introuvable dans ce fichier")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # Ligne 11 (index 11) = en-têtes avec les dates des mois
    header_row = rows[11]
    # Colonnes paires = Quantité, impaires = Montant (à partir de col 2)
    months = []  # liste de (col_index_qty, datetime_mois)
    for ci, cell in enumerate(header_row):
        if ci < 2:
            continue
        from datetime import datetime as _dt
        if hasattr(cell, 'year') or (isinstance(cell, _dt)):
            months.append((ci, cell))

    if not months:
        raise HTTPException(400, "Impossible de trouver les colonnes de dates")

    # Supprime les anciens imports historiques de ce fichier (évite doublons)
    source_tag = file.filename or "cashpad_historique"

    imported = 0
    errors = []

    for row in rows[13:]:
        if not row[0] and not row[1]:
            continue
        cat = str(row[0] or "").strip()
        nom = str(row[1] or "").strip()
        if not nom or nom.startswith("Total"):
            continue

        for ci, mois_dt in months:
            qty_val = row[ci] if ci < len(row) else None
            if not qty_val or float(qty_val or 0) <= 0:
                continue
            qty = float(qty_val)

            # Crée une entrée StockHistory par mois/produit
            from datetime import datetime as _dt2
            if hasattr(mois_dt, 'year'):
                event_date = _dt2(mois_dt.year, mois_dt.month, 15)  # milieu du mois
            else:
                continue

            # Récupère aussi le CA (colonne suivante)
            ca_val = row[ci + 1] if ci + 1 < len(row) else None
            ca = round(float(ca_val), 2) if ca_val else 0.0

            h = StockHistory(
                event_type="ventes_historiques",
                description=f"Historique {mois_dt.strftime('%b %Y') if hasattr(mois_dt,'strftime') else mois_dt} — {nom}",
                data_json=json.dumps({
                    "product": nom,
                    "category": cat,
                    "quantity": qty,
                    "ca": ca,
                    "mois": event_date.strftime("%Y-%m"),
                    "source": source_tag,
                }, ensure_ascii=False),
                created_at=event_date,
            )
            db.add(h)
            imported += 1

    db.commit()
    return {"ok": True, "imported": imported, "mois": len(months), "source": source_tag}


@app.get("/api/stats/historique")
def stats_historique(db: Session = Depends(get_db)):
    """Retourne les ventes historiques avec analyses décisionnelles."""
    rows = db.query(StockHistory).filter(
        StockHistory.event_type == "ventes_historiques"
    ).all()

    by_month = {}     # "2025-05" → {qty, ca}
    by_product = {}   # nom → {qty, ca, category, par_mois}
    by_cat = {}       # cat → {qty, ca}

    for row in rows:
        try:
            d = json.loads(row.data_json)
        except Exception:
            continue
        mois = d.get("mois", "")
        qty  = float(d.get("quantity", 0))
        ca   = float(d.get("ca", 0))
        nom  = d.get("product", "?")
        cat  = d.get("category", "Autres")

        if mois not in by_month:
            by_month[mois] = {"qty": 0, "ca": 0}
        by_month[mois]["qty"] = round(by_month[mois]["qty"] + qty, 1)
        by_month[mois]["ca"]  = round(by_month[mois]["ca"]  + ca,  2)

        if nom not in by_product:
            by_product[nom] = {"qty": 0, "ca": 0, "category": cat, "par_mois": {}}
        by_product[nom]["qty"] = round(by_product[nom]["qty"] + qty, 1)
        by_product[nom]["ca"]  = round(by_product[nom]["ca"]  + ca,  2)
        by_product[nom]["par_mois"][mois] = round(by_product[nom]["par_mois"].get(mois, 0) + qty, 1)

        if cat not in by_cat:
            by_cat[cat] = {"qty": 0, "ca": 0}
        by_cat[cat]["qty"] = round(by_cat[cat]["qty"] + qty, 1)
        by_cat[cat]["ca"]  = round(by_cat[cat]["ca"]  + ca,  2)

    total_qty = round(sum(v["qty"] for v in by_month.values()), 1)
    total_ca  = round(sum(v["ca"]  for v in by_month.values()), 2)

    # Classement produits
    all_products = sorted(
        [{"name": k, "qty": v["qty"], "ca": v["ca"], "category": v["category"],
          "par_mois": v["par_mois"],
          "pct": round(v["qty"] / total_qty * 100, 1) if total_qty else 0}
         for k, v in by_product.items()],
        key=lambda x: x["qty"], reverse=True
    )

    # Top 20 par catégorie
    by_cat_products = {}
    for p in all_products:
        c = p["category"]
        if c not in by_cat_products:
            by_cat_products[c] = []
        by_cat_products[c].append(p)

    # Mois peak
    peak = max(by_month.items(), key=lambda x: x[1]["qty"]) if by_month else ("—", {"qty": 0})

    monthly = sorted([{"mois": k, "qty": v["qty"], "ca": v["ca"]} for k, v in by_month.items()], key=lambda x: x["mois"])
    by_category = sorted([{"category": k, "qty": v["qty"], "ca": v["ca"]} for k, v in by_cat.items()], key=lambda x: x["qty"], reverse=True)

    return {
        "total_qty": total_qty,
        "total_ca": total_ca,
        "peak_mois": peak[0],
        "peak_qty": peak[1]["qty"],
        "monthly": monthly,
        "by_category": by_category,
        "top_products": all_products[:20],
        "bottom_products": all_products[-30:][::-1],  # 30 moins vendus, du moins au plus
        "by_cat_products": {k: v[:20] for k, v in by_cat_products.items()},
        "nb_produits": len(all_products),
    }


# ══════════════════════════════════════════════════════════════════════════
# INVENTAIRE DU SOIR
# ══════════════════════════════════════════════════════════════════════════

PRIORITY_CATEGORIES = [
    "Spiritueux", "Rhums", "Gins", "Whiskies", "Vodkas",
    "Champagnes", "Apéritifs", "Digestifs", "Anisés", "Tequilas", "Cachaça"
]

# Ordre d'affichage des catégories dans l'inventaire du soir
INVENTORY_CATEGORY_ORDER = [
    "Bières", "Spiritueux", "Rhums", "Gins", "Whiskies", "Vodkas",
    "Tequilas", "Cachaça", "Apéritifs", "Digestifs", "Anisés",
    "Champagnes", "Vins", "Jus de fruit", "Sodas", "Eaux", "Sirops"
]


@app.get("/api/inventaire/session")
@app.get("/api/inventory/session")
def get_inventory_session(db: Session = Depends(get_db)):
    """Retourne TOUS les produits à comptabiliser groupés par catégorie.
    Priorité : alcools + grandes bouteilles de mixing (volume >= 75cl, qty_per_pack=1).
    Exclut les produits en carton (comptés à la réception) et fûts (comptés par le débit).
    """
    all_products = db.query(Product).filter(
        Product.unit != 'Fût',       # Les fûts sont suivis via Cashpad
    ).order_by(Product.category, Product.name).all()

    result = []
    for p in all_products:
        # Cartons → compté à la réception, pas à l'inventaire soir
        if p.unit and 'Carton' in p.unit:
            continue
        result.append({
            "id": p.id,
            "nom": p.name,
            "name": p.name,
            "categorie": p.category,
            "category": p.category,
            "unite": p.unit,
            "unit": p.unit,
            "volume_cl": p.volume_cl,
            "qty_per_pack": p.qty_per_pack,
            "quantite_theorique": round(p.stock, 3),
            "theoretical": round(p.stock, 3),
            "is_grande_bouteille": bool(p.volume_cl and p.volume_cl >= 75 and (p.qty_per_pack or 1) == 1 and p.unit not in ('Fût',)),
        })

    # Trier : ordre des catégories prioritaires, puis le reste
    def sort_key(item):
        cat = item["category"] or "ZZZ"
        try:
            return (INVENTORY_CATEGORY_ORDER.index(cat), item["name"])
        except ValueError:
            return (99, item["name"])

    result.sort(key=sort_key)
    return result


class InventoryCountIn(BaseModel):
    counts: List[dict]
    staff_name: str = ""


@app.post("/api/inventaire/compter")
@app.post("/api/inventory/submit")
def submit_inventory(body: InventoryCountIn, db: Session = Depends(get_db)):
    results = []
    alerts = []
    for item in body.counts:
        pid = item.get("product_id") or item.get("produit_id")
        p = db.query(Product).get(pid)
        if not p:
            continue
        theoretical = p.stock
        actual = float(item.get("actual", item.get("quantite_reelle", 0)))
        diff = actual - theoretical
        sess = InventorySession(
            product_id=p.id,
            theoretical_qty=theoretical,
            actual_qty=actual,
            difference=diff,
            staff_name=body.staff_name,
        )
        db.add(sess)
        p.stock = actual
        result = {
            "product": p.name,
            "theoretical": round(theoretical, 3),
            "actual": actual,
            "diff": round(diff, 3),
            "ecart": round(diff, 3),
        }
        results.append(result)
        if diff < -0.5:
            alerts.append(f"Écart anormal : {p.name} (écart {diff:+.3f})")
            log_event(db, "alerte_inventaire", f"Écart anormal détecté : {p.name}", result)

    log_event(db, "inventaire_soir", f"Inventaire du soir — {len(results)} produits comptés", {
        "staff": body.staff_name,
        "results": results,
        "alerts": alerts,
    })
    db.commit()
    return {"ok": True, "results": results, "alerts": alerts}


# ═══════════════════════════════════════════════════════════════════════════
#  INVENTAIRE FLASH — comptage par photo IA
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/api/inventory/flash-test")
def flash_test_api():
    """Endpoint de diagnostic : vérifie que l'API Anthropic est joignable."""
    api_key = os.getenv("ANTHROPIC_API_KEY", "").strip().strip()
    if not api_key:
        return {"ok": False, "error": "ANTHROPIC_API_KEY non configurée"}
    masked = api_key[:8] + "…" + api_key[-4:] if len(api_key) > 12 else "***"
    try:
        payload = json.dumps({
            "model": "claude-haiku-4-5-20251001",
            "max_tokens": 20,
            "messages": [{"role": "user", "content": "Dis juste OK"}],
        }).encode("utf-8")
        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=payload,
            headers={
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        return {"ok": True, "key": masked, "response": data["content"][0]["text"]}
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        return {"ok": False, "key": masked, "error": f"HTTP {e.code}: {body[:300]}"}
    except Exception as e:
        return {"ok": False, "key": masked, "error": f"{type(e).__name__}: {str(e)}"}


class FindProductIn(BaseModel):
    description: str


@app.post("/api/inventory/find-product")
def find_product_ai(body: FindProductIn, db: Session = Depends(get_db)):
    """
    IA : reçoit une description libre (ex: "bouteille bleue de gin en carafe")
    et retourne jusqu'à 3 produits du catalogue qui correspondent le mieux.
    """
    api_key = os.getenv("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        raise HTTPException(400, "Clé API Anthropic non configurée.")
    desc = (body.description or "").strip()
    if len(desc) < 3:
        raise HTTPException(400, "Description trop courte (min 3 caractères).")

    # Catalogue actif (non archivé) — envoyé au modèle
    products = db.query(Product).filter(
        (Product.archived == False) | (Product.archived == None)
    ).all()
    if not products:
        return {"candidates": []}

    # Compact catalogue : id, nom, catégorie, unité
    catalog = "\n".join(
        f"{p.id}|{p.name}|{p.category or ''}|{p.unit or ''}"
        for p in products
    )

    import anthropic as _anthropic
    client = _anthropic.Anthropic(api_key=api_key)
    prompt = (
        "Tu aides un serveur de bar à retrouver un produit dans le catalogue. "
        "Voici le catalogue (format: id|nom|catégorie|unité), un produit par ligne:\n\n"
        f"{catalog}\n\n"
        f"Description donnée par le serveur: \"{desc}\"\n\n"
        "Retourne les 1 à 3 produits qui correspondent le mieux, au format JSON strict:\n"
        '{"candidates":[{"id":123,"reason":"courte raison"}, ...]}\n'
        "Classe du meilleur au moins bon. Si rien ne correspond vraiment, retourne une liste vide."
    )
    try:
        response = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=400,
            messages=[{"role": "user", "content": prompt}],
        )
        text = ""
        for block in response.content:
            if hasattr(block, "text"):
                text += block.text
        # Extraire le JSON même si entouré de texte
        import re, json as _json
        m = re.search(r"\{.*\}", text, re.DOTALL)
        if not m:
            return {"candidates": []}
        data = _json.loads(m.group(0))
        raw_candidates = data.get("candidates", [])[:3]
        # Enrichir avec les infos produit
        out = []
        for c in raw_candidates:
            pid = c.get("id")
            p = next((x for x in products if x.id == pid), None)
            if not p:
                continue
            out.append({
                "id": p.id,
                "name": p.name,
                "category": p.category or "",
                "unit": p.unit or "",
                "reason": (c.get("reason") or "").strip()[:120],
            })
        return {"candidates": out}
    except Exception as e:
        raise HTTPException(500, f"Erreur IA : {str(e)[:160]}")


@app.post("/api/inventory/flash-analyze")
async def flash_analyze_photo(
    file: UploadFile = File(...),
    zone: str = Form(""),
    db: Session = Depends(get_db),
):
    """Analyse une photo de frigo/étagère et détecte les bouteilles visibles.
    Utilise Claude vision pour identifier et compter les bouteilles."""
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(400, detail="Image trop volumineuse (max 20 Mo)")

    # Compresser + réorienter EXIF (photos téléphone = 5-12 Mo, souvent en portrait iPhone)
    try:
        from PIL import Image, ImageOps
        img = Image.open(io.BytesIO(content))
        img = ImageOps.exif_transpose(img)
        # Convertir RGBA/P → RGB si nécessaire
        if img.mode in ("RGBA", "P", "LA"):
            img = img.convert("RGB")
        # Redimensionner si > 1600px de large (suffisant pour identifier des bouteilles)
        max_dim = 1600
        if max(img.size) > max_dim:
            img.thumbnail((max_dim, max_dim), Image.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=85, optimize=True)
        content = buf.getvalue()
    except Exception:
        pass  # si Pillow échoue, on envoie l'image originale

    api_key = os.getenv("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        raise HTTPException(400, detail="Clé API Anthropic non configurée. Vérifiez le fichier .env")

    # Récupérer la liste des produits connus pour aider l'IA à identifier
    all_prods = db.query(Product).filter(Product.unit != 'Fût').order_by(Product.category, Product.name).all()
    known_products = []
    for p in all_prods:
        if p.unit and 'Carton' in p.unit:
            continue
        known_products.append({"id": p.id, "name": p.name, "category": p.category})

    product_list_str = "\n".join(
        f"- ID {p['id']}: {p['name']} (catégorie: {p['category']})"
        for p in known_products
    )

    prompt_text = f"""Tu analyses une photo d'un frigo, d'une étagère ou d'une zone de stockage de boissons dans un bar/restaurant.

TÂCHE : Identifie et compte TOUTES les bouteilles, canettes et contenants de boissons visibles sur la photo.

INSTRUCTIONS :
1. Pour chaque type de produit visible, indique le nombre exact de bouteilles/canettes.
2. Essaie de reconnaître les marques et types de produits (Coca-Cola, Pietra, Orezza, etc.)
3. Si tu vois des bouteilles partiellement cachées, inclus-les dans ton comptage avec une note.
4. Regroupe les produits identiques ensemble.
5. Sois précis : il vaut mieux signaler une incertitude que donner un chiffre faux.

PRODUITS CONNUS DANS NOTRE BASE :
{product_list_str}

IMPORTANT : Quand tu reconnais un produit, associe-le à l'ID correspondant de la liste ci-dessus.
Si tu ne reconnais pas un produit ou s'il n'est pas dans la liste, mets product_id à null.

Zone indiquée par l'utilisateur : {zone or "non précisée"}

Réponds UNIQUEMENT en JSON valide avec ce format :
{{
  "zone_description": "Description courte de ce que tu vois (ex: frigo principal, étagère haute...)",
  "total_bottles": 22,
  "confidence": "high/medium/low",
  "items": [
    {{
      "product_name": "Coca-Cola 33cl",
      "product_id": 15,
      "quantity": 6,
      "confidence": "high",
      "notes": ""
    }},
    {{
      "product_name": "Bouteille inconnue (verre vert)",
      "product_id": null,
      "quantity": 2,
      "confidence": "low",
      "notes": "partiellement cachées derrière les Coca"
    }}
  ],
  "observations": "Remarques générales (éclairage, visibilité, bouteilles possiblement cachées...)"
}}"""

    # Appel direct HTTP à l'API Anthropic (contourne les bugs du SDK)
    media_type = "image/jpeg"
    b64 = base64.standard_b64encode(content).decode("utf-8")
    img_size_kb = len(content) // 1024

    api_payload = json.dumps({
        "model": "claude-haiku-4-5-20251001",
        "max_tokens": 4096,
        "messages": [{"role": "user", "content": [
            {"type": "image",
             "source": {"type": "base64", "media_type": media_type, "data": b64}},
            {"type": "text", "text": prompt_text},
        ]}],
    }).encode("utf-8")

    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=api_payload,
        headers={
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
    )

    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            api_result = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")[:200]
        raise HTTPException(e.code, detail=f"API Anthropic erreur HTTP {e.code}: {body}")
    except urllib.error.URLError as e:
        raise HTTPException(502, detail=f"Connexion API impossible (image: {img_size_kb}Ko): {str(e.reason)}")
    except Exception as e:
        raise HTTPException(500, detail=f"Erreur ({type(e).__name__}): {str(e)} — image: {img_size_kb}Ko")

    raw = api_result["content"][0]["text"].strip()

    # Parser le JSON de la réponse
    obj_start = raw.find("{")
    obj_end = raw.rfind("}") + 1
    if obj_start == -1:
        raise HTTPException(400, detail="L'IA n'a pas pu analyser cette image. Essayez avec un meilleur éclairage.")

    try:
        result = json.loads(raw[obj_start:obj_end])
    except json.JSONDecodeError:
        raise HTTPException(400, detail="L'analyse n'a pas produit un résultat exploitable. Réessayez.")

    # Enrichir chaque item avec le stock théorique actuel
    for item in result.get("items", []):
        pid = item.get("product_id")
        if pid:
            prod = db.query(Product).get(pid)
            if prod:
                item["current_stock"] = round(prod.stock, 3)
                item["category"] = prod.category
            else:
                item["product_id"] = None
                item["current_stock"] = None
                item["category"] = None
        else:
            item["current_stock"] = None
            item["category"] = None

    return result


class FlashControlIn(BaseModel):
    counts: List[dict]    # [{product_id, product_name, actual, theoretical}]
    staff_name: str = ""
    zone: str = ""


@app.post("/api/inventory/flash-save")
def flash_save_control(body: FlashControlIn, db: Session = Depends(get_db)):
    """Enregistre un rapport de contrôle Inventaire Flash SANS modifier le stock.
    C'est un constat : on compare le comptage IA/corrigé avec le stock théorique."""
    items = []
    alerts = []
    for item in body.counts:
        pid = item.get("product_id")
        if not pid:
            continue
        p = db.query(Product).get(pid)
        if not p:
            continue
        theoretical = round(p.stock, 3)
        actual = float(item.get("actual", 0))
        diff = round(actual - theoretical, 3)
        items.append({
            "product_id": p.id,
            "product_name": p.name,
            "category": p.category,
            "theoretical": theoretical,
            "actual": actual,
            "diff": diff,
            "corrected": False,
        })
        if diff < -0.5:
            alerts.append(f"Écart : {p.name} ({diff:+.3f})")

    # Enregistrer le rapport complet dans stock_history
    report = {
        "staff": body.staff_name,
        "zone": body.zone,
        "items": items,
        "alerts": alerts,
        "date": datetime.now(_LOCAL_TZ).strftime("%d/%m/%Y %H:%M"),
    }
    h = log_event(db, "controle_flash",
                  f"Contrôle Flash — {len(items)} produits, zone: {body.zone or 'non précisée'}",
                  report)
    db.commit()
    return {"ok": True, "control_id": h.id, "items": items, "alerts": alerts}


class FlashCorrectIn(BaseModel):
    qty: Optional[float] = None


@app.post("/api/inventory/flash-correct/{control_id}/{product_id}")
def flash_correct_product(control_id: int, product_id: int, body: FlashCorrectIn = None, db: Session = Depends(get_db)):
    if body is None:
        body = FlashCorrectIn()
    """Corrige le stock d'UN seul produit suite à un contrôle flash.
    Accepte un body JSON optionnel {"qty": X} pour corriger la quantité avant envoi."""
    h = db.query(StockHistory).get(control_id)
    if not h or h.event_type != "controle_flash":
        raise HTTPException(404, detail="Rapport de contrôle introuvable")

    report = json.loads(h.data_json)
    items = report.get("items", [])

    target = None
    for item in items:
        if item.get("product_id") == product_id:
            target = item
            break
    if not target:
        raise HTTPException(404, detail="Produit non trouvé dans ce contrôle")

    if target.get("corrected"):
        raise HTTPException(400, detail="Ce produit a déjà été corrigé")

    p = db.query(Product).get(product_id)
    if not p:
        raise HTTPException(404, detail="Produit introuvable")

    old_stock = p.stock
    new_stock = target["actual"]

    if body.qty is not None:
        new_stock = body.qty
        target["actual"] = new_stock
        target["diff"] = round(new_stock - target["theoretical"], 3)
    diff = round(new_stock - old_stock, 3)
    p.stock = new_stock

    # Marquer comme corrigé dans le rapport
    target["corrected"] = True
    target["corrected_at"] = datetime.now(_LOCAL_TZ).strftime("%d/%m/%Y %H:%M")
    h.data_json = json.dumps(report, ensure_ascii=False)

    # Enregistrer la correction dans l'historique aussi
    log_event(db, "correction_flash",
              f"Correction stock (flash) : {p.name} — {old_stock} → {new_stock} (écart {diff:+.3f})",
              {"product_id": p.id, "product_name": p.name, "old_stock": old_stock,
               "new_stock": new_stock, "diff": diff, "control_id": control_id})

    db.commit()
    return {"ok": True, "product_name": p.name, "old_stock": round(old_stock, 3),
            "new_stock": new_stock, "diff": diff}


@app.get("/api/inventory/flash-history")
def flash_control_history(db: Session = Depends(get_db)):
    """Historique des contrôles flash."""
    controls = db.query(StockHistory).filter(
        StockHistory.event_type == "controle_flash"
    ).order_by(StockHistory.created_at.desc()).limit(50).all()

    result = []
    for c in controls:
        data = json.loads(c.data_json) if c.data_json else {}
        items = data.get("items", [])
        nb_ecarts = sum(1 for it in items if abs(it.get("diff", 0)) > 0.1)
        nb_corrected = sum(1 for it in items if it.get("corrected"))
        result.append({
            "id": c.id,
            "date": to_local(c.created_at).strftime("%d/%m/%Y %H:%M") if c.created_at else "",
            "staff": data.get("staff", ""),
            "zone": data.get("zone", ""),
            "nb_products": len(items),
            "nb_ecarts": nb_ecarts,
            "nb_corrected": nb_corrected,
            "items": items,
        })
    return result


# ═══════════════════════════════════════════════════════════════════════════
#  MODULE ÉVÉNEMENTS / BOOST
# ═══════════════════════════════════════════════════════════════════════════

class EventRequirementIn(BaseModel):
    product_id: int
    quantity: float
    notes: Optional[str] = ""


class EventIn(BaseModel):
    name: str
    event_type: str = "Autre"
    date: str                                       # "YYYY-MM-DD" — date de début
    end_date: Optional[str] = None                  # "YYYY-MM-DD" — date de fin (optionnel)
    start_time: Optional[str] = ""                  # "HH:MM" optionnel
    end_time: Optional[str] = ""                    # "HH:MM" optionnel
    notes: str = ""
    requirements: Optional[List[EventRequirementIn]] = None


def _event_to_dict(e: Event) -> dict:
    reqs = []
    for r in (e.requirements or []):
        reqs.append({
            "id": r.id,
            "product_id": r.product_id,
            "product_name": r.product.name if r.product else "",
            "unit": r.product.unit if r.product else "",
            "stock": r.product.stock if r.product else 0,
            "quantity": r.quantity,
            "notes": r.notes or "",
        })
    return {
        "id": e.id,
        "name": e.name,
        "event_type": e.event_type,
        "date": e.date.strftime("%Y-%m-%d"),
        "end_date": e.end_date.strftime("%Y-%m-%d") if e.end_date else None,
        "start_time": e.start_time or "",
        "end_time": e.end_time or "",
        "notes": e.notes,
        "created_at": e.created_at.strftime("%d/%m/%Y") if e.created_at else "",
        "requirements": reqs,
    }


@app.get("/api/events")
def list_events(db: Session = Depends(get_db)):
    evs = db.query(Event).order_by(Event.date.desc()).all()
    return [_event_to_dict(e) for e in evs]


def _apply_event_body(ev: Event, body: EventIn, db: Session):
    ev.name = body.name
    ev.event_type = body.event_type or "Autre"
    ev.date = datetime.strptime(body.date, "%Y-%m-%d")
    if body.end_date:
        end = datetime.strptime(body.end_date, "%Y-%m-%d")
        if end < ev.date:
            raise HTTPException(400, "La date de fin doit être postérieure à la date de début.")
        ev.end_date = end
    else:
        ev.end_date = None
    ev.start_time = (body.start_time or "").strip()
    ev.end_time = (body.end_time or "").strip()
    ev.notes = body.notes
    # Requirements : on remplace intégralement la liste
    if body.requirements is not None:
        for old in list(ev.requirements or []):
            db.delete(old)
        for r in body.requirements:
            if not r.product_id or r.quantity is None or r.quantity <= 0:
                continue
            ev.requirements.append(EventRequirement(
                product_id=r.product_id,
                quantity=float(r.quantity),
                notes=(r.notes or "").strip(),
            ))


@app.post("/api/events")
def create_event(body: EventIn, db: Session = Depends(get_db)):
    ev = Event()
    db.add(ev)
    _apply_event_body(ev, body, db)
    db.commit()
    db.refresh(ev)
    return {"id": ev.id}


@app.put("/api/events/{eid}")
def update_event(eid: int, body: EventIn, db: Session = Depends(get_db)):
    ev = db.query(Event).get(eid)
    if not ev:
        raise HTTPException(404, "Événement introuvable")
    _apply_event_body(ev, body, db)
    db.commit()
    return {"ok": True}


@app.delete("/api/events/{eid}")
def delete_event(eid: int, db: Session = Depends(get_db)):
    ev = db.query(Event).get(eid)
    if not ev:
        raise HTTPException(404)
    db.delete(ev)
    db.commit()
    return {"ok": True}


# ── Analyse IA d'une demande client (PDF / email) ─────────────────────────
@app.post("/api/events/parse-request")
async def parse_event_request(
    file: Optional[UploadFile] = File(None),
    text: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    """
    Reçoit un PDF (bon de commande, devis client) ou un texte (email collé).
    Utilise Claude pour extraire la liste des boissons demandées et les
    associer aux produits du catalogue.
    Retourne: [{product_id, product_name, quantity, matched, raw_name}]
    """
    # ── 1. Extraire le texte ──────────────────────────────────────────────
    extracted_text = (text or "").strip()
    pdf_bytes = None

    if file is not None:
        content = await file.read()
        fname = (file.filename or "").lower()
        if fname.endswith(".pdf") or (file.content_type and "pdf" in file.content_type):
            pdf_bytes = content
            try:
                import pdfplumber
                with pdfplumber.open(io.BytesIO(content)) as pdf:
                    pages = []
                    for page in pdf.pages:
                        t = page.extract_text() or ""
                        if t.strip():
                            pages.append(t)
                    pdf_text = "\n".join(pages).strip()
                if pdf_text:
                    extracted_text = (extracted_text + "\n" + pdf_text).strip()
            except Exception:
                pass
        else:
            try:
                extracted_text = (extracted_text + "\n" + content.decode("utf-8", errors="ignore")).strip()
            except Exception:
                pass

    if not extracted_text and not pdf_bytes:
        raise HTTPException(400, detail="Aucun contenu à analyser (PDF vide ou texte manquant).")

    # ── 2. Préparer le catalogue (pour que l'IA fasse le matching) ─────────
    products = db.query(Product).filter(
        (Product.archived == False) | (Product.archived == None)
    ).all()
    catalogue_lines = [
        f"- id={p.id} | {p.name} ({p.category}, {p.unit})"
        for p in products
    ]
    catalogue_str = "\n".join(catalogue_lines)

    # ── 3. Appel Claude ───────────────────────────────────────────────────
    api_key = os.getenv("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        raise HTTPException(400, detail="Clé API Anthropic non configurée (ANTHROPIC_API_KEY).")

    import anthropic as _anthropic
    client = _anthropic.Anthropic(api_key=api_key)

    system_prompt = (
        "Tu es un assistant qui analyse des demandes de boissons pour un événement "
        "(anniversaire, mariage, soirée privée, tournoi…). "
        "Tu reçois un catalogue de produits disponibles et la demande du client "
        "(email ou bon de commande PDF). "
        "Extrait chaque boisson demandée avec sa quantité, et associe-la à un produit "
        "du catalogue si possible (matching sémantique, tolérer fautes de frappe, synonymes "
        "ex: « Deutz » → « Champagne Deutz », « bière Pietra » → « Fût Pietra Blonde 30L » ou "
        "variante selon contexte). "
        "Si aucun produit du catalogue ne correspond, retourne product_id=null et mets le "
        "nom brut dans raw_name. "
        "Réponds UNIQUEMENT en JSON valide, sans texte avant ou après."
    )

    user_text = (
        "CATALOGUE PRODUITS (id | nom | catégorie | unité) :\n"
        f"{catalogue_str}\n\n"
        "FORMAT DE RÉPONSE (JSON strict) :\n"
        '{"items": [{"product_id": 42 | null, "raw_name": "Champagne Deutz Brut", '
        '"quantity": 20, "unit": "bouteille"}]}\n\n'
        "DEMANDE DU CLIENT :\n"
        "────────────────────\n"
        f"{extracted_text if extracted_text else '(voir le PDF joint)'}"
    )

    try:
        content_blocks: list = []
        if pdf_bytes and not extracted_text:
            b64 = base64.standard_b64encode(pdf_bytes).decode("utf-8")
            content_blocks.append({
                "type": "document",
                "source": {"type": "base64", "media_type": "application/pdf", "data": b64},
            })
        content_blocks.append({"type": "text", "text": user_text})

        message = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=2048,
            system=system_prompt,
            messages=[{"role": "user", "content": content_blocks}],
        )
    except _anthropic.AuthenticationError:
        raise HTTPException(401, detail="Clé API Anthropic invalide.")
    except _anthropic.BadRequestError as e:
        raise HTTPException(400, detail=f"Demande non lisible : {str(e)}")
    except Exception as e:
        raise HTTPException(500, detail=f"Erreur d'analyse IA : {str(e)}")

    raw = message.content[0].text.strip() if message.content else ""
    # Récupérer le JSON
    start = raw.find("{")
    end = raw.rfind("}") + 1
    if start < 0 or end <= start:
        raise HTTPException(500, detail=f"Réponse IA invalide : {raw[:200]}")
    try:
        parsed = json.loads(raw[start:end])
    except json.JSONDecodeError as e:
        raise HTTPException(500, detail=f"JSON IA invalide : {e}")

    # ── 4. Normaliser la réponse et enrichir avec info produit ────────────
    products_by_id = {p.id: p for p in products}
    items = []
    for it in parsed.get("items", []):
        pid = it.get("product_id")
        raw_name = (it.get("raw_name") or "").strip()
        qty = it.get("quantity") or 0
        try:
            qty = float(qty)
        except Exception:
            qty = 0
        if qty <= 0:
            continue
        if pid and pid in products_by_id:
            p = products_by_id[pid]
            items.append({
                "product_id": p.id,
                "product_name": p.name,
                "raw_name": raw_name,
                "quantity": qty,
                "unit": p.unit or "",
                "stock": p.stock,
                "matched": True,
            })
        else:
            items.append({
                "product_id": None,
                "product_name": raw_name,
                "raw_name": raw_name,
                "quantity": qty,
                "unit": it.get("unit") or "",
                "stock": 0,
                "matched": False,
            })

    return {"ok": True, "items": items, "raw_text_preview": extracted_text[:500]}


@app.get("/api/events/analysis")
def get_events_analysis(db: Session = Depends(get_db)):
    """
    Pour chaque type d'événement, compare la consommation lors des événements
    vs la consommation moyenne des jours 'normaux'. Retourne le boost % par produit.
    """
    from collections import defaultdict

    events = db.query(Event).order_by(Event.date.desc()).all()
    if not events:
        return []

    # ── 1. Construction du map consommation journalière ──────────────────
    all_imports = db.query(StockHistory).filter(
        StockHistory.event_type == "import_cashpad"
    ).all()

    # daily_consumption[day][product_id] = qty_totale_ce_jour
    daily_consumption: dict = {}
    for imp in all_imports:
        day = imp.created_at.strftime("%Y-%m-%d")
        d = json.loads(imp.data_json)
        if day not in daily_consumption:
            daily_consumption[day] = {}
        for ded in d.get("deductions", []):
            pid = ded.get("product_id")
            qty = abs(float(ded.get("quantity", ded.get("qty", 0)) or 0))
            if pid and qty > 0:
                daily_consumption[day][pid] = daily_consumption[day].get(pid, 0) + qty

    # ── 2. Baseline : jours sans événement ───────────────────────────────
    def _event_days(ev):
        start = ev.date.date() if hasattr(ev.date, "date") else ev.date
        end = (ev.end_date.date() if ev.end_date and hasattr(ev.end_date, "date") else (ev.end_date or start))
        if end < start:
            end = start
        days = []
        cur = start
        while cur <= end:
            days.append(cur.strftime("%Y-%m-%d"))
            cur += timedelta(days=1)
        return days

    all_event_days = {d for ev in events for d in _event_days(ev)}
    non_event_days = {d: c for d, c in daily_consumption.items() if d not in all_event_days}

    baseline: dict = {}
    if non_event_days:
        n_base = len(non_event_days)
        for day_data in non_event_days.values():
            for pid, qty in day_data.items():
                baseline[pid] = baseline.get(pid, 0) + qty
        baseline = {pid: total / n_base for pid, total in baseline.items()}

    # ── 3. Regroupement par type ──────────────────────────────────────────
    type_events: dict = defaultdict(list)
    for ev in events:
        type_events[ev.event_type or "Autre"].append(ev)

    # Noms de produits
    products_map = {p.id: p.name for p in db.query(Product).all()}

    results = []
    for et, evs in type_events.items():
        # Consommation totale sur les jours d'événement de ce type
        event_consumption: dict = defaultdict(float)
        n_with_data = 0
        event_list = []
        for ev in evs:
            days = _event_days(ev)
            date_label = ev.date.strftime("%d/%m/%Y")
            if ev.end_date and ev.end_date != ev.date:
                date_label = f"{ev.date.strftime('%d/%m/%Y')} → {ev.end_date.strftime('%d/%m/%Y')}"
            event_list.append({"id": ev.id, "name": ev.name, "date": date_label})
            had_data = False
            for day in days:
                if day in daily_consumption:
                    had_data = True
                    for pid, qty in daily_consumption[day].items():
                        event_consumption[pid] += qty
            if had_data:
                n_with_data += 1

        if not event_consumption:
            results.append({
                "event_type": et,
                "count": len(evs),
                "events": event_list,
                "boosts": [],
                "no_data": True,
                "n_with_data": 0,
            })
            continue

        n_ev = max(n_with_data, 1)
        event_avg = {pid: total / n_ev for pid, total in event_consumption.items()}

        boosts = []
        for pid, avg in event_avg.items():
            base = baseline.get(pid, 0)
            if base > 0:
                pct = round((avg / base - 1) * 100, 1)
            else:
                pct = None   # pas de baseline → nouveau produit
            boosts.append({
                "product_id": pid,
                "product_name": products_map.get(pid, f"#{pid}"),
                "event_avg": round(avg, 2),
                "baseline_avg": round(base, 2),
                "boost_pct": pct,
            })

        # Tri : plus fort boost d'abord ; produits sans baseline à la fin
        boosts.sort(key=lambda x: (x["boost_pct"] is None, -(x["boost_pct"] or 0)))

        results.append({
            "event_type": et,
            "count": len(evs),
            "n_with_data": n_with_data,
            "events": event_list,
            "boosts": boosts[:15],   # top 15 produits
        })

    # Tri résultats : plus d'événements d'abord
    results.sort(key=lambda x: -x["count"])
    return results


# ═══════════════════════════════════════════════════════════════════════════
#  MODULE DÉMARQUE INCONNUE
# ═══════════════════════════════════════════════════════════════════════════

LOSS_REASONS = ["Casse", "Offert maison", "Dégustation", "Vol suspecté", "Périmé", "Autre"]


class ManualLossIn(BaseModel):
    product_id: int
    quantity: float
    reason: str = "Autre"
    notes: str = ""
    date: Optional[str] = None   # "YYYY-MM-DD", défaut = aujourd'hui
    staff_name: str = ""
    update_stock: bool = True    # déduire du stock immédiatement


@app.get("/api/losses")
def list_losses(limit: int = 200, db: Session = Depends(get_db)):
    rows = (
        db.query(ManualLoss)
        .order_by(ManualLoss.date.desc())
        .limit(limit)
        .all()
    )
    return [
        {
            "id": r.id,
            "product_id": r.product_id,
            "product_name": r.product.name if r.product else f"#{r.product_id}",
            "category": r.product.category if r.product else "",
            "quantity": r.quantity,
            "unit": r.product.unit if r.product else "",
            "reason": r.reason,
            "notes": r.notes,
            "date": r.date.strftime("%d/%m/%Y") if r.date else "",
            "staff_name": r.staff_name,
            "stock_updated": r.stock_updated,
            "value_eur": round(r.quantity * (r.product.purchase_price or 0), 2) if r.product else 0,
        }
        for r in rows
    ]


@app.post("/api/losses")
def create_loss(body: ManualLossIn, db: Session = Depends(get_db)):
    p = db.query(Product).get(body.product_id)
    if not p:
        raise HTTPException(404, "Produit introuvable")

    loss_date = (
        datetime.strptime(body.date, "%Y-%m-%d")
        if body.date else datetime.utcnow()
    )

    loss = ManualLoss(
        product_id=body.product_id,
        quantity=body.quantity,
        reason=body.reason,
        notes=body.notes,
        date=loss_date,
        staff_name=body.staff_name,
        stock_updated=body.update_stock,
    )
    db.add(loss)

    if body.update_stock:
        p.stock = max(0, p.stock - body.quantity)
        log_event(db, "perte_declaree",
                  f"Perte déclarée : {p.name} — {body.quantity} {p.unit} ({body.reason})",
                  {"product_id": p.id, "quantity": body.quantity,
                   "reason": body.reason, "staff": body.staff_name})

    db.commit()
    db.refresh(loss)
    return {"id": loss.id, "new_stock": round(p.stock, 3)}


@app.delete("/api/losses/{lid}")
def delete_loss(lid: int, db: Session = Depends(get_db)):
    loss = db.query(ManualLoss).get(lid)
    if not loss:
        raise HTTPException(404)
    # Restituer le stock si la perte avait été déduite
    if loss.stock_updated and loss.product:
        loss.product.stock = loss.product.stock + loss.quantity
    db.delete(loss)
    db.commit()
    return {"ok": True}


@app.get("/api/shrinkage/summary")
def get_shrinkage_summary(db: Session = Depends(get_db)):
    """
    Par produit :
      - inventory_loss  : somme des écarts NÉGATIFS d'inventaire (ce qu'on n'arrive pas à expliquer)
      - declared_losses : pertes déclarées manuellement (casse, offert, etc.)
      - unexplained     : inventory_loss − declared_losses (vraie démarque inconnue)
      - value_eur       : unexplained × prix d'achat
    """
    from sqlalchemy import func

    # ── Écarts d'inventaire ──────────────────────────────────────────────
    inv_rows = (
        db.query(
            InventorySession.product_id,
            func.sum(InventorySession.difference).label("total_diff"),
            func.count(InventorySession.id).label("nb"),
        )
        .group_by(InventorySession.product_id)
        .all()
    )
    inv_map = {r.product_id: (float(r.total_diff or 0), int(r.nb)) for r in inv_rows}

    # ── Pertes déclarées ────────────────────────────────────────────────
    loss_rows = (
        db.query(
            ManualLoss.product_id,
            func.sum(ManualLoss.quantity).label("total_qty"),
        )
        .group_by(ManualLoss.product_id)
        .all()
    )
    loss_map = {r.product_id: float(r.total_qty or 0) for r in loss_rows}

    products_map = {p.id: p for p in db.query(Product).all()}
    all_pids = set(inv_map) | set(loss_map)

    result = []
    for pid in all_pids:
        p = products_map.get(pid)
        if not p:
            continue

        total_diff, nb_inv = inv_map.get(pid, (0.0, 0))
        declared = loss_map.get(pid, 0.0)

        # Perte d'inventaire = somme des écarts négatifs uniquement
        inv_loss = abs(min(0.0, total_diff))

        # Les pertes déclarées AVEC update_stock ne créent pas d'écart inventaire
        # (stock déjà déduit). Les pertes sans update_stock apparaissent dans inv_loss.
        # On montre les deux colonnes séparément pour que le gérant comprenne.
        unexplained = max(0.0, inv_loss - declared)

        price = p.purchase_price or 0
        value_eur = round(unexplained * price, 2)

        result.append({
            "product_id": pid,
            "product_name": p.name,
            "category": p.category or "",
            "unit": p.unit or "Bouteille",
            "stock_actuel": round(p.stock, 2),
            "purchase_price": price,
            "inventory_loss": round(inv_loss, 2),
            "inventory_gain": round(max(0.0, total_diff), 2),
            "declared_losses": round(declared, 2),
            "unexplained": round(unexplained, 2),
            "value_eur": value_eur,
            "nb_inventaires": nb_inv,
        })

    result.sort(key=lambda x: -x["value_eur"])
    return result


@app.get("/api/shrinkage/history")
def get_shrinkage_history(db: Session = Depends(get_db)):
    """Historique mensuel des pertes déclarées + écarts inventaire."""
    from sqlalchemy import func

    # Pertes déclarées par mois
    losses = db.query(ManualLoss).order_by(ManualLoss.date).all()
    inv_sessions = db.query(InventorySession).order_by(InventorySession.created_at).all()
    products_map = {p.id: p for p in db.query(Product).all()}

    monthly: dict = {}

    for loss in losses:
        month = loss.date.strftime("%Y-%m") if loss.date else "??-??"
        if month not in monthly:
            monthly[month] = {"declared": 0.0, "declared_eur": 0.0,
                               "inventory": 0.0, "inventory_eur": 0.0}
        p = products_map.get(loss.product_id)
        price = (p.purchase_price or 0) if p else 0
        monthly[month]["declared"] += loss.quantity
        monthly[month]["declared_eur"] += loss.quantity * price

    for sess in inv_sessions:
        if sess.difference >= 0:
            continue   # gain → on ne compte pas
        month = sess.created_at.strftime("%Y-%m") if sess.created_at else "??-??"
        if month not in monthly:
            monthly[month] = {"declared": 0.0, "declared_eur": 0.0,
                               "inventory": 0.0, "inventory_eur": 0.0}
        p = products_map.get(sess.product_id)
        price = (p.purchase_price or 0) if p else 0
        monthly[month]["inventory"] += abs(sess.difference)
        monthly[month]["inventory_eur"] += abs(sess.difference) * price

    rows = [
        {
            "month": m,
            "label": _fmt_month_short_fr(datetime.strptime(m, "%Y-%m")) if m != "??-??" else m,
            **v
        }
        for m, v in sorted(monthly.items())
    ]
    return rows


# ═══════════════════════════════════════════════════════════════════════════
#  CONNEXION CASHPAD API — Sync automatique
# ═══════════════════════════════════════════════════════════════════════════

_CASHPAD_BASE = os.getenv("CASHPAD_BASE_URL", "https://www3.cashpad.net")


def _get_setting(db: Session, key: str, default: str = "") -> str:
    row = db.query(AppSetting).filter(AppSetting.key == key).first()
    return row.value if row else default


def _set_setting(db: Session, key: str, value: str) -> None:
    row = db.query(AppSetting).filter(AppSetting.key == key).first()
    if row:
        row.value = value
        row.updated_at = datetime.utcnow()
    else:
        db.add(AppSetting(key=key, value=value))


def _cashpad_get(path: str, params: dict) -> dict:
    """Requête GET vers l'API Cashpad avec gestion d'erreurs."""
    qs = urllib.parse.urlencode(params)
    url = f"{_CASHPAD_BASE}{path}?{qs}"
    req = urllib.request.Request(url, headers={"Accept": "application/json", "User-Agent": "MarinadiLava/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            body = json.loads(resp.read().decode("utf-8"))
    except _urllib_err.HTTPError as e:
        raise HTTPException(502, f"Cashpad HTTP {e.code} — {e.reason}")
    except Exception as e:
        raise HTTPException(502, f"Impossible de joindre Cashpad : {e}")

    if body.get("error"):
        raise HTTPException(502, f"Cashpad API error : {body['error']}")
    return body.get("data", body)


def _process_cashpad_archive(archive_data: dict, db: Session) -> dict:
    """
    Traite une archive Cashpad (équivalent d'un import CSV).
    Retourne les stats de la sync.
    """
    seq_id = archive_data.get("sequential_id", "?")
    start_date = archive_data.get("start_date", "")[:10]
    end_date   = archive_data.get("end_date",   "")[:10]

    # Les lignes de ventes peuvent arriver sous plusieurs noms selon la version API
    lines = (
        archive_data.get("order_lines") or
        archive_data.get("lines")       or
        archive_data.get("sales")       or
        archive_data.get("products")    or
        archive_data.get("items")       or
        []
    )

    if not lines:
        return {"seq_id": seq_id, "synced": 0, "skipped": 0, "skipped_names": []}

    # Charge les mappings Cashpad (nom_cashpad → mapping)
    mappings = {m.nom_cashpad.lower(): m for m in db.query(CashpadMapping).all()}

    deductions = []
    sales      = []
    skipped    = []

    for line in lines:
        # Normalise les noms de champs (l'API peut varier)
        name = (
            line.get("name") or line.get("nom") or
            line.get("product_name") or line.get("label") or
            line.get("item_name") or ""
        ).strip()

        qty = float(
            line.get("qty") or line.get("quantity") or
            line.get("quantite") or line.get("sold_qty") or 0
        )

        # Prix de vente TTC éventuellement présent dans la ligne
        line_price = 0.0
        for _f in ("unit_price_ttc", "price_ttc", "unit_price", "price", "ttc", "prix_ttc", "prix_unitaire"):
            if line.get(_f) is not None:
                try: line_price = float(line[_f]); break
                except Exception: pass

        if not name or qty <= 0:
            continue

        mapping = mappings.get(name.lower())
        if not mapping or mapping.ignored:
            skipped.append(name)
            continue

        if mapping.mapping_type == "cocktail" and mapping.cocktail_id:
            cocktail = db.query(Cocktail).get(mapping.cocktail_id)
            if cocktail:
                cost_matiere = 0.0
                has_alcohol = False
                for ing in cocktail.ingredients:
                    if ing.product:
                        dose_cl = ing.dose_cl * qty
                        vol = (ing.product.volume_cl or 70)
                        bottles = dose_cl / vol
                        ing.product.stock = max(0.0, ing.product.stock - bottles)
                        deductions.append({
                            "product_id":   ing.product.id,
                            "product_name": ing.product.name,
                            "quantity":     round(bottles, 4),
                        })
                        if ing.product.purchase_price is not None:
                            cost_matiere += (ing.product.purchase_price / vol) * ing.dose_cl
                        if (ing.product.vat_rate or 0) >= 0.20:
                            has_alcohol = True
                sales.append({
                    "ref_type":       "cocktail",
                    "ref_id":         cocktail.id,
                    "product_name":   cocktail.name,
                    "name":           cocktail.name,
                    "qty_sold":       qty,
                    "sale_price_ttc": line_price if line_price > 0 else (cocktail.sale_price_ttc or 0),
                    "purchase_price": round(cost_matiere, 4),
                    "vat_rate":       0.20 if has_alcohol else 0.10,
                    "unit":           "u",
                })

        elif mapping.product_id:
            p = db.query(Product).get(mapping.product_id)
            if p:
                dose_cl = (mapping.dose_cl or 0) * qty
                if dose_cl > 0 and p.volume_cl:
                    bottles = dose_cl / p.volume_cl
                else:
                    bottles = qty   # déduction unitaire directe
                p.stock = max(0.0, p.stock - bottles)
                deductions.append({
                    "product_id":   p.id,
                    "product_name": p.name,
                    "quantity":     round(bottles, 4),
                })
                # Figer le coût matière associé à la vente
                if mapping.dose_cl and p.volume_cl:
                    unit_cost = (p.purchase_price / p.volume_cl) * mapping.dose_cl if p.purchase_price is not None else None
                else:
                    unit_cost = p.purchase_price
                sales.append({
                    "ref_type":       "direct",
                    "ref_id":         p.id,
                    "product_id":     p.id,
                    "product_name":   mapping.nom_cashpad,
                    "name":           mapping.nom_cashpad,
                    "qty_sold":       qty,
                    "sale_price_ttc": line_price if line_price > 0 else (p.sale_price_ttc or 0),
                    "purchase_price": round(unit_cost, 4) if unit_cost is not None else None,
                    "vat_rate":       p.vat_rate if p.vat_rate is not None else 0.20,
                    "unit":           p.unit or "u",
                })

    if deductions:
        log_event(
            db, "import_cashpad",
            f"Sync Cashpad API — archive #{seq_id} ({start_date} → {end_date})",
            {"deductions": deductions, "sales": sales, "source": "api_sync", "sequential_id": seq_id},
        )
        db.commit()

    return {
        "seq_id":        seq_id,
        "period":        f"{start_date} → {end_date}",
        "synced":        len(deductions),
        "skipped":       len(skipped),
        "skipped_names": list(set(skipped)),
    }


def _run_cashpad_sync(db: Session) -> dict:
    """
    Logique centrale de sync : récupère les nouvelles archives et les traite.
    Utilisée par le endpoint ET le scheduler.
    """
    email      = os.getenv("CASHPAD_EMAIL", "").strip()
    token      = os.getenv("CASHPAD_TOKEN", "").strip()
    install_id = os.getenv("CASHPAD_INSTALLATION_ID", "").strip()

    if not all([email, token, install_id]):
        return {
            "ok": False,
            "error": (
                "Cashpad non configuré. Ajoutez CASHPAD_EMAIL, CASHPAD_TOKEN "
                "et CASHPAD_INSTALLATION_ID dans vos variables Railway."
            ),
        }

    auth = {"apiuser_email": email, "apiuser_token": token}
    last_id = int(_get_setting(db, "cashpad_last_sequential_id", "0") or 0)

    # 1. Récupère la liste des archives depuis la dernière sync
    params = {**auth}
    if last_id:
        params["start_sequential_id"] = last_id + 1

    try:
        archives_payload = _cashpad_get(
            f"/api/salesdata/v2/{install_id}/archives", params
        )
    except HTTPException as e:
        return {"ok": False, "error": e.detail}

    # Normalise la liste (peut être une list ou {"archives": [...]})
    archives = (
        archives_payload
        if isinstance(archives_payload, list)
        else archives_payload.get("archives", [])
    )

    if not archives:
        _set_setting(db, "cashpad_last_sync", datetime.utcnow().isoformat())
        db.commit()
        return {"ok": True, "message": "Stock à jour — aucune nouvelle archive", "archives": 0}

    results   = []
    max_seq   = last_id

    for archive in archives:
        seq_id = archive.get("sequential_id") or archive.get("id")
        if seq_id is None:
            continue
        seq_id = int(seq_id)

        try:
            content = _cashpad_get(
                f"/api/salesdata/v2/{install_id}/archive_content",
                {**auth, "sequential_id": seq_id},
            )
            result = _process_cashpad_archive(content, db)
            results.append(result)
            max_seq = max(max_seq, seq_id)
        except HTTPException as e:
            results.append({"seq_id": seq_id, "error": e.detail})
        except Exception as e:
            results.append({"seq_id": seq_id, "error": str(e)})

    # Sauvegarde le curseur et l'heure de sync
    if max_seq > last_id:
        _set_setting(db, "cashpad_last_sequential_id", str(max_seq))
    _set_setting(db, "cashpad_last_sync", datetime.utcnow().isoformat())
    db.commit()

    total_synced  = sum(r.get("synced",  0) for r in results)
    total_skipped = sum(r.get("skipped", 0) for r in results)

    return {
        "ok":               True,
        "archives":         len(results),
        "total_synced":     total_synced,
        "total_skipped":    total_skipped,
        "last_seq_id":      max_seq,
        "results":          results,
    }


# ── Endpoints ────────────────────────────────────────────────────────────

@app.get("/api/cashpad/sync-status")
def cashpad_sync_status(db: Session = Depends(get_db)):
    email      = os.getenv("CASHPAD_EMAIL", "").strip()
    token      = os.getenv("CASHPAD_TOKEN", "").strip()
    install_id = os.getenv("CASHPAD_INSTALLATION_ID", "").strip()
    configured = bool(email and token and install_id)

    last_sync  = _get_setting(db, "cashpad_last_sync", "")
    last_seq   = _get_setting(db, "cashpad_last_sequential_id", "0")

    # Formate la date de dernière sync
    last_sync_label = ""
    if last_sync:
        try:
            dt = datetime.fromisoformat(last_sync)
            local = to_local(dt)
            last_sync_label = f"{local.day} {_MONTHS_FR_SHORT[local.month - 1]} {local.year} à {local.strftime('%H:%M')}"
        except Exception:
            last_sync_label = last_sync[:16]

    return {
        "configured":       configured,
        "email":            email[:3] + "***" if email else "",
        "installation_id":  install_id,
        "last_sync":        last_sync_label,
        "last_sequential_id": last_seq,
    }


@app.post("/api/cashpad/sync")
def cashpad_sync_now(db: Session = Depends(get_db)):
    """Déclenche une sync manuelle immédiate."""
    result = _run_cashpad_sync(db)
    if not result.get("ok"):
        raise HTTPException(400, result.get("error", "Erreur inconnue"))
    return result


@app.post("/api/cashpad/reset-cursor")
def cashpad_reset_cursor(db: Session = Depends(get_db)):
    """Remet le curseur à zéro pour re-syncer depuis le début."""
    _set_setting(db, "cashpad_last_sequential_id", "0")
    db.commit()
    return {"ok": True}


# ── Scheduler APScheduler — sync automatique toutes les 30 min ───────────

try:
    from apscheduler.schedulers.background import BackgroundScheduler
    from database import SessionLocal as _SessionLocal

    def _auto_sync_job():
        db = _SessionLocal()
        try:
            r = _run_cashpad_sync(db)
            if r.get("archives", 0) > 0:
                print(f"[Cashpad auto-sync] ✅ {r['archives']} archives, {r['total_synced']} déductions")
        except Exception as e:
            print(f"[Cashpad auto-sync] ❌ {e}")
        finally:
            db.close()

    def _weekly_snapshot_job():
        """Prend une photo du stock tous les lundis 02:00 pour la démarque auto."""
        db = _SessionLocal()
        try:
            products = db.query(Product).filter(
                (Product.archived == False) | (Product.archived == None)
            ).all()
            now = datetime.utcnow()
            for p in products:
                db.add(StockSnapshot(
                    product_id=p.id, taken_at=now,
                    stock=p.stock or 0, label="weekly_auto",
                ))
            db.commit()
            print(f"[Snapshot] 📸 {len(products)} produits figés le {now.strftime('%Y-%m-%d %H:%M')}")
        except Exception as e:
            print(f"[Snapshot] ❌ {e}")
        finally:
            db.close()

    _scheduler = BackgroundScheduler(timezone="Europe/Paris")
    _scheduler.add_job(_auto_sync_job, "interval", minutes=30, id="cashpad_sync", replace_existing=True)
    _scheduler.add_job(_weekly_snapshot_job, "cron", day_of_week="mon", hour=2, minute=0,
                       id="weekly_snapshot", replace_existing=True)
    _scheduler.add_job(_run_weekly_backup, "cron", day_of_week="sun", hour=3, minute=0,
                       id="weekly_backup", replace_existing=True)
    _scheduler.start()
    print("[Cashpad] 🔄 Scheduler démarré — sync 30min + snapshot lundi 02h + backup dimanche 03h")

except Exception as _sched_err:
    print(f"[Cashpad] ⚠️ Scheduler non démarré : {_sched_err}")


# ═══════════════════════════════════════════════════════════════════════════
#  MODULE MÉTÉO — Alerte Pic de Chaleur
# ═══════════════════════════════════════════════════════════════════════════

# Sensibilité à la chaleur par mot-clé (catégorie ou nom produit)
# Valeur = % de hausse consommation à 30°C
_HEAT_SENSITIVITY: dict = {
    "eau":        0.35, "water":     0.35, "minérale":  0.35,
    "bière":      0.28, "biere":     0.28, "beer":      0.28,
    "pils":       0.25, "lager":     0.25, "blonde":    0.22,
    "rosé":       0.22, "rose":      0.22,
    "soda":       0.20, "cola":      0.20, "limonade":  0.20,
    "tonic":      0.18, "schweppes": 0.18, "agrume":    0.18,
    "jus":        0.15, "juice":     0.15, "nectar":    0.15,
    "cidre":      0.15,
    "champagne":  0.12, "prosecco":  0.12, "crémant":   0.12,
    "ice tea":    0.20, "ice-tea":   0.20,
    "mojito":     0.18, "spritz":    0.15,
}

# Multiplicateur par niveau de chaleur (appliqué sur la sensibilité de base)
_HEAT_MULTIPLIERS = {
    "canicule":  1.50,   # ≥35°C  → ×1.5 sur la sensibilité
    "chaud":     1.00,   # 30–34°C → base
    "tiede":     0.45,   # 25–29°C → ×0.45
}


def _fetch_openweather(lat: str, lon: str, api_key: str) -> dict:
    """Appelle OpenWeather Forecast 5 jours et retourne les données brutes."""
    qs = urllib.parse.urlencode({
        "lat": lat, "lon": lon, "appid": api_key,
        "units": "metric", "lang": "fr", "cnt": 16,   # 16×3h = 2 jours
    })
    url = f"https://api.openweathermap.org/data/2.5/forecast?{qs}"
    req = urllib.request.Request(url, headers={"Accept": "application/json"})
    with urllib.request.urlopen(req, timeout=10) as resp:
        return json.loads(resp.read().decode("utf-8"))


@app.get("/api/weather")
def get_weather(refresh: bool = False, db: Session = Depends(get_db)):
    api_key = (os.getenv("OPENWEATHER_API_KEY") or os.getenv("OPENWEATHER-API-KEY") or "").strip()
    lat     = os.getenv("WEATHER_LAT", "41.9267").strip()   # Ajaccio, Corse (défaut)
    lon     = os.getenv("WEATHER_LON", "8.7369").strip()

    if not api_key:
        return {"configured": False}

    # ── Cache 1h ────────────────────────────────────────────────────────
    if not refresh:
        cached    = _get_setting(db, "weather_cache", "")
        cache_ts  = _get_setting(db, "weather_cache_ts", "")
        if cached and cache_ts:
            try:
                age = (datetime.utcnow() - datetime.fromisoformat(cache_ts)).total_seconds()
                if age < 3600:
                    return json.loads(cached)
            except Exception:
                pass

    # ── Appel API ────────────────────────────────────────────────────────
    try:
        raw = _fetch_openweather(lat, lon, api_key)
    except _urllib_err.HTTPError as e:
        return {"configured": True, "error": f"OpenWeather HTTP {e.code} : {e.reason}"}
    except Exception as e:
        return {"configured": True, "error": str(e)}

    city      = raw.get("city", {}).get("name", "")
    country   = raw.get("city", {}).get("country", "")
    forecasts = raw.get("list", [])
    if not forecasts:
        return {"configured": True, "error": "Aucune donnée météo disponible"}

    # ── Températures ─────────────────────────────────────────────────────
    current     = forecasts[0]
    current_t   = round(current["main"]["temp"], 1)
    current_desc= (current["weather"][0]["description"] if current.get("weather") else "").capitalize()
    current_icon= current["weather"][0]["icon"] if current.get("weather") else "01d"

    # Demain : créneaux 8→15 (24h→48h après maintenant, 3h chacun)
    tomorrow    = forecasts[8:16] if len(forecasts) >= 16 else forecasts[4:8]
    if not tomorrow:
        tomorrow = forecasts[1:]

    t_max  = round(max(s["main"].get("temp_max", s["main"]["temp"]) for s in tomorrow), 1)
    t_min  = round(min(s["main"].get("temp_min", s["main"]["temp"]) for s in tomorrow), 1)
    mid    = tomorrow[len(tomorrow)//2]
    t_desc = (mid["weather"][0]["description"] if mid.get("weather") else "").capitalize()
    t_icon = mid["weather"][0]["icon"] if mid.get("weather") else "01d"

    # ── Niveau d'alerte ──────────────────────────────────────────────────
    if t_max >= 35:
        level, emoji, label, mult = "canicule", "🔥", f"Canicule prévue demain ({t_max:.0f}°C) — préparez-vous !", _HEAT_MULTIPLIERS["canicule"]
    elif t_max >= 30:
        level, emoji, label, mult = "chaud",    "☀️", f"Forte chaleur prévue demain ({t_max:.0f}°C)", _HEAT_MULTIPLIERS["chaud"]
    elif t_max >= 25:
        level, emoji, label, mult = "tiede",    "🌤️", f"Belle journée prévue demain ({t_max:.0f}°C)", _HEAT_MULTIPLIERS["tiede"]
    else:
        level, emoji, label, mult = "normal",   "⛅", f"Temps normal prévu demain ({t_max:.0f}°C)", 0.0

    # ── Suggestions produits ──────────────────────────────────────────────
    suggestions = []
    if mult > 0:
        products = db.query(Product).filter(Product.stock > 0).all()
        for p in products:
            haystack = f"{(p.category or '').lower()} {(p.name or '').lower()}"
            base_boost = max(
                (pct for kw, pct in _HEAT_SENSITIVITY.items() if kw in haystack),
                default=0.0,
            )
            if base_boost <= 0:
                continue
            effective = round(base_boost * mult * 100)   # en %
            if effective < 8:
                continue
            extra = max(1, round(p.stock * base_boost * mult))
            suggestions.append({
                "product_id":   p.id,
                "product_name": p.name,
                "category":     p.category or "",
                "current_stock": round(p.stock, 1),
                "unit":          p.unit or "u",
                "boost_pct":     effective,
                "extra_units":   extra,
            })
        suggestions.sort(key=lambda x: -x["boost_pct"])
        suggestions = suggestions[:10]

    result = {
        "configured":    True,
        "city":          f"{city}, {country}" if city else "Corse",
        "current_temp":  current_t,
        "current_desc":  current_desc,
        "current_icon":  current_icon,
        "tomorrow_max":  t_max,
        "tomorrow_min":  t_min,
        "tomorrow_desc": t_desc,
        "tomorrow_icon": t_icon,
        "alert_level":   level,
        "alert_emoji":   emoji,
        "alert_label":   label,
        "suggestions":   suggestions,
    }

    # Mise en cache
    _set_setting(db, "weather_cache",    json.dumps(result))
    _set_setting(db, "weather_cache_ts", datetime.utcnow().isoformat())
    db.commit()
    return result


# ══════════════════════════════════════════════════════════════════════════
# TABLEAU DE BORD
# ══════════════════════════════════════════════════════════════════════════

_MONTHS_FR = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
              "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
_MONTHS_FR_SHORT = ["janv.", "févr.", "mars", "avr.", "mai", "juin",
                    "juil.", "août", "sept.", "oct.", "nov.", "déc."]

def _fmt_month_fr(d) -> str:
    """Retourne 'Avril 2026' au lieu de 'April 2026' (locale système non garantie sur Railway)."""
    return f"{_MONTHS_FR[d.month - 1]} {d.year}"

def _fmt_month_short_fr(d) -> str:
    """Retourne 'avr. 2026' au lieu de 'Apr 2026'."""
    return f"{_MONTHS_FR_SHORT[d.month - 1]} {d.year}"


@app.get("/api/manque-a-gagner")
def get_manque_a_gagner(db: Session = Depends(get_db)):
    """Calcule le manque à gagner dû aux ruptures de stock ce mois."""
    now_local = datetime.now(_LOCAL_TZ)
    month_start = now_local.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # Récupérer les alertes de rupture du mois
    ruptures = db.query(ServiceAlert).filter(
        ServiceAlert.is_rupture == True,
        ServiceAlert.created_at >= month_start.astimezone(timezone.utc).replace(tzinfo=None),
    ).all()

    if not ruptures:
        return {"total_lost": 0, "items": [], "month": _fmt_month_fr(now_local)}

    # Calculer les ventes moyennes par produit (sur 30 derniers jours d'historique)
    thirty_days_ago = datetime.utcnow() - timedelta(days=30)
    sales_history = db.query(StockHistory).filter(
        StockHistory.event_type.in_(["import_cashpad", "cashpad_sync"]),
        StockHistory.created_at >= thirty_days_ago,
    ).all()

    daily_sales = {}
    for h in sales_history:
        for item in extract_sales(h):
            name = item.get("product_name") or item.get("name", "")
            qty = float(item.get("qty_sold", 0) or 0)
            price = float(item.get("sale_price_ttc", 0) or 0)
            if name and qty > 0 and price > 0:
                if name not in daily_sales:
                    daily_sales[name] = {"total_qty": 0, "price": price}
                daily_sales[name]["total_qty"] += qty

    # Calculer le manque pour chaque produit en rupture
    items = []
    total_lost = 0.0
    for r in ruptures:
        p = r.product
        if not p:
            continue
        # Durée de rupture en heures
        end = r.resolved_at or datetime.utcnow()
        hours = (end - r.created_at).total_seconds() / 3600

        # Ventes moyennes par heure (sur 30j, environ 10h d'ouverture/jour)
        avg = daily_sales.get(p.name, {})
        total_qty = avg.get("total_qty", 0)
        price = avg.get("price", p.sale_price_ttc or 0)
        daily_avg = total_qty / 30 if total_qty > 0 else 0
        hourly_avg = daily_avg / 10  # ~10h d'ouverture par jour

        lost_qty = hourly_avg * hours
        lost_eur = lost_qty * price

        if lost_eur > 0.5:
            total_lost += lost_eur
            items.append({
                "product_name": p.name,
                "hours_rupture": round(hours, 1),
                "lost_qty": round(lost_qty, 1),
                "lost_eur": round(lost_eur, 2),
                "reported_by": r.staff_name,
                "date": to_local(r.created_at).strftime("%d/%m %H:%M") if r.created_at else "",
            })

    items.sort(key=lambda x: -x["lost_eur"])
    return {
        "total_lost": round(total_lost, 2),
        "items": items,
        "month": _fmt_month_fr(now_local),
    }


@app.get("/api/dashboard")
def get_dashboard(db: Session = Depends(get_db)):
    now_local = datetime.now(_LOCAL_TZ)
    today     = now_local.date()

    # ── Météo (depuis cache) ──────────────────────────────────────────────
    api_key = (os.getenv("OPENWEATHER_API_KEY") or os.getenv("OPENWEATHER-API-KEY") or "").strip()
    weather_summary: dict
    if not api_key:
        weather_summary = {"configured": False}
    else:
        cached   = _get_setting(db, "weather_cache", "")
        cache_ts = _get_setting(db, "weather_cache_ts", "")
        if cached and cache_ts:
            try:
                age = (datetime.utcnow() - datetime.fromisoformat(cache_ts)).total_seconds()
                if age < 7200:          # 2h de tolérance pour le dashboard
                    full = json.loads(cached)
                    weather_summary = {
                        "configured":   True,
                        "current_temp": full.get("current_temp"),
                        "alert_level":  full.get("alert_level"),
                        "alert_emoji":  full.get("alert_emoji"),
                        "alert_label":  full.get("alert_label"),
                        "city":         full.get("city", "Corse"),
                    }
                else:
                    weather_summary = {"configured": True, "stale": True}
            except Exception:
                weather_summary = {"configured": True, "stale": True}
        else:
            # Pas encore en cache : tenter un appel léger
            lat = os.getenv("WEATHER_LAT", "41.9267").strip()
            lon = os.getenv("WEATHER_LON", "8.7369").strip()
            try:
                raw = _fetch_openweather(lat, lon, api_key)
                forecasts = raw.get("list", [])
                city_name = raw.get("city", {}).get("name", "")
                country   = raw.get("city", {}).get("country", "")
                if forecasts:
                    current_t = round(forecasts[0]["main"]["temp"], 1)
                    tomorrow  = forecasts[8:16] if len(forecasts) >= 16 else forecasts[1:]
                    t_max = round(max(s["main"].get("temp_max", s["main"]["temp"]) for s in tomorrow), 1) if tomorrow else current_t
                    if t_max >= 35:
                        level, emoji, label = "canicule", "🔥", f"Canicule prévue ({t_max:.0f}°C)"
                    elif t_max >= 30:
                        level, emoji, label = "chaud",    "☀️", f"Forte chaleur ({t_max:.0f}°C)"
                    elif t_max >= 25:
                        level, emoji, label = "tiede",    "🌤️", f"Belle journée ({t_max:.0f}°C)"
                    else:
                        level, emoji, label = "normal",   "⛅", f"Temps normal ({t_max:.0f}°C)"
                    weather_summary = {
                        "configured":   True,
                        "current_temp": current_t,
                        "alert_level":  level,
                        "alert_emoji":  emoji,
                        "alert_label":  label,
                        "city":         f"{city_name}, {country}" if city_name else "Corse",
                    }
                else:
                    weather_summary = {"configured": True, "stale": True}
            except Exception:
                weather_summary = {"configured": True, "stale": True}

    # ── Prochains événements (inclut événements multi-jours en cours) ─────
    today_dt = datetime.combine(today, datetime.min.time())
    upcoming_events_q = (
        db.query(Event)
        .filter(
            (Event.date >= today_dt) |
            ((Event.end_date != None) & (Event.end_date >= today_dt))
        )
        .order_by(Event.date.asc())
        .limit(15)
        .all()
    )
    events_upcoming = [
        {
            "id":         ev.id,
            "name":       ev.name,
            "event_type": ev.event_type,
            "date":       ev.date.isoformat(),
            "end_date":   ev.end_date.isoformat() if ev.end_date else None,
            "start_time": ev.start_time or "",
            "end_time":   ev.end_time or "",
        }
        for ev in upcoming_events_q
    ]

    # ── Alertes stock urgentes ────────────────────────────────────────────
    urgent_products = (
        db.query(Product)
        .filter(Product.stock <= Product.alert_threshold)
        .filter((Product.archived == False) | (Product.archived == None))
        .order_by((Product.stock / (Product.alert_threshold + 0.001)).asc())
        .limit(5)
        .all()
    )
    urgent_alerts = [
        {
            "id":              p.id,
            "name":            p.name,
            "stock":           round(p.stock, 2),
            "alert_threshold": p.alert_threshold,
            "unit":            p.unit or "u",
        }
        for p in urgent_products
    ]

    # ── CA hier & CA semaine + Top produits ──────────────────────────────
    yesterday_start = datetime.combine(today - timedelta(days=1), datetime.min.time())
    today_start     = datetime.combine(today,                     datetime.min.time())
    week_start      = datetime.combine(today - timedelta(days=7), datetime.min.time())

    cashpad_history = (
        db.query(StockHistory)
        .filter(
            StockHistory.event_type == "import_cashpad",
            StockHistory.created_at >= week_start,
        )
        .all()
    )

    ca_yesterday = 0.0
    ca_week      = 0.0
    product_totals: dict = {}   # name → {"qty": float, "unit": str}

    for h in cashpad_history:
        for item in extract_sales(h):
            qty   = float(item.get("qty_sold", 0) or 0)
            price = float(item.get("sale_price_ttc", 0) or 0)
            name  = item.get("product_name") or item.get("name", "")
            unit  = item.get("unit", "u")
            revenue = qty * price
            ca_week += revenue
            if yesterday_start <= h.created_at < today_start:
                ca_yesterday += revenue
            if name:
                if name not in product_totals:
                    product_totals[name] = {"qty": 0.0, "unit": unit}
                product_totals[name]["qty"] += qty

    top_products = sorted(
        [{"name": k, "qty_sold": round(v["qty"], 2), "unit": v["unit"]} for k, v in product_totals.items()],
        key=lambda x: -x["qty_sold"],
    )[:3]

    # ── CA N-1 (même jour de semaine, même semaine de l'année, année -1) ──
    yesterday = today - timedelta(days=1)
    iso_year, iso_week, iso_day = yesterday.isocalendar()
    # Trouver le même jour (iso_week, iso_day) de l'année précédente
    from datetime import date as _date
    jan4_prev = _date(iso_year - 1, 1, 4)  # le 4 janvier est toujours en semaine ISO 1
    start_w1 = jan4_prev - timedelta(days=jan4_prev.isocalendar()[2] - 1)  # lundi S1 de l'an passé
    target_n1 = start_w1 + timedelta(weeks=iso_week - 1, days=iso_day - 1)

    n1_start = datetime.combine(target_n1, datetime.min.time())
    n1_end   = datetime.combine(target_n1 + timedelta(days=1), datetime.min.time())

    ca_n1 = 0.0
    n1_history = (
        db.query(StockHistory)
        .filter(
            StockHistory.event_type == "import_cashpad",
            StockHistory.created_at >= n1_start,
            StockHistory.created_at < n1_end,
        ).all()
    )
    for h in n1_history:
        for item in extract_sales(h):
            qty   = float(item.get("qty_sold", 0) or 0)
            price = float(item.get("sale_price_ttc", 0) or 0)
            ca_n1 += qty * price

    day_names = ["lundi","mardi","mercredi","jeudi","vendredi","samedi","dimanche"]

    # ── Widget 1 : santé Cashpad (dernière sync + activité 24h) ───────────
    last_sync_raw = _get_setting(db, "cashpad_last_sync", "")
    # Activité des 24 dernières heures
    since_24h = datetime.utcnow() - timedelta(hours=24)
    tx_24h = 0
    ca_24h = 0.0
    hist_24h = (
        db.query(StockHistory)
        .filter(
            StockHistory.event_type.in_(["import_cashpad", "cashpad_sync"]),
            StockHistory.created_at >= since_24h,
        )
        .all()
    )
    for h in hist_24h:
        for item in extract_sales(h):
            qty = float(item.get("qty_sold", 0) or 0)
            price = float(item.get("sale_price_ttc", 0) or 0)
            if qty > 0:
                tx_24h += 1
                ca_24h += qty * price

    last_sync_info = {"has_sync": False, "tx_24h": tx_24h, "ca_24h": round(ca_24h, 2)}
    if last_sync_raw:
        try:
            last_sync_dt = datetime.fromisoformat(last_sync_raw)
            delta = datetime.utcnow() - last_sync_dt
            mins = int(delta.total_seconds() // 60)
            hours = mins // 60
            days = hours // 24
            if mins < 60:
                label = f"il y a {max(mins, 0)} min"
            elif hours < 24:
                label = f"il y a {hours} h"
            elif days < 7:
                label = f"il y a {days} jour{'s' if days > 1 else ''}"
            else:
                label = f"il y a {days} jours"
            # Santé : combine fraîcheur + activité
            if hours >= 24:
                status = "error"
            elif hours >= 4 or tx_24h == 0:
                status = "warn"
            else:
                status = "ok"
            last_sync_info = {
                "has_sync":    True,
                "iso":         last_sync_raw,
                "label":       label,
                "status":      status,
                "date_fr":     to_local(last_sync_dt).strftime("%d/%m/%Y %H:%M"),
                "tx_24h":      tx_24h,
                "ca_24h":      round(ca_24h, 2),
            }
        except Exception:
            pass

    # ── Widget 2 : commandes fournisseurs en attente (envoyées/partielles) ─
    pending_orders_q = (
        db.query(SupplierOrder)
        .filter(SupplierOrder.status.in_(["sent", "partial"]))
        .order_by(SupplierOrder.sent_at.asc())
        .limit(10)
        .all()
    )
    pending_orders = []
    for o in pending_orders_q:
        ref_date = o.sent_at or o.created_at
        days_since = (datetime.utcnow() - ref_date).days if ref_date else 0
        pending_orders.append({
            "id":           o.id,
            "reference":    o.reference,
            "supplier":     o.supplier.name if o.supplier else "",
            "status":       o.status,
            "sent_at":      o.sent_at.isoformat() if o.sent_at else None,
            "days_since":   days_since,
            "n_items":      len(o.items or []),
        })

    # ── Widget 3 : CA par jour de la semaine (4 dernières semaines) ───────
    range_start = datetime.combine(today - timedelta(days=28), datetime.min.time())
    weekday_hist = (
        db.query(StockHistory)
        .filter(
            StockHistory.event_type == "import_cashpad",
            StockHistory.created_at >= range_start,
        )
        .all()
    )
    # weekday 0 = lundi
    ca_by_weekday = [0.0] * 7
    cnt_by_weekday = [0] * 7
    seen_days = {}   # weekday → set of days
    for h in weekday_hist:
        day_key = h.created_at.strftime("%Y-%m-%d")
        wd = h.created_at.weekday()
        revenue_day = 0.0
        for item in extract_sales(h):
            qty   = float(item.get("qty_sold", 0) or 0)
            price = float(item.get("sale_price_ttc", 0) or 0)
            revenue_day += qty * price
        ca_by_weekday[wd] += revenue_day
        if wd not in seen_days:
            seen_days[wd] = set()
        seen_days[wd].add(day_key)
    for wd, days_set in seen_days.items():
        cnt_by_weekday[wd] = len(days_set)
    ca_weekday = [
        {
            "day":      day_names[i],
            "avg":      round(ca_by_weekday[i] / cnt_by_weekday[i], 2) if cnt_by_weekday[i] else 0,
            "total":    round(ca_by_weekday[i], 2),
            "n_days":   cnt_by_weekday[i],
        }
        for i in range(7)
    ]

    # ── Widget : objectif mois en cours + comparaison N-1 ─────────────────
    import calendar as _cal
    monthly_goal = {"configured": False}
    try:
        monthly_json = _get_setting(db, "monthly_goals_json", "")
        monthly_goals = json.loads(monthly_json) if monthly_json else {}
        if not isinstance(monthly_goals, dict):
            monthly_goals = {}
    except Exception:
        monthly_goals = {}

    try:
        cur_year, cur_month = today.year, today.month
        cur_key = f"{cur_year:04d}-{cur_month:02d}"
        month_start = datetime(cur_year, cur_month, 1)
        days_in_month = _cal.monthrange(cur_year, cur_month)[1]
        next_month = (month_start + timedelta(days=days_in_month)).replace(day=1)

        # CA mois en cours
        ca_month = 0.0
        hist_m = db.query(StockHistory).filter(
            StockHistory.event_type.in_(["import_cashpad", "cashpad_sync"]),
            StockHistory.created_at >= month_start,
            StockHistory.created_at <  next_month,
        ).all()
        for h in hist_m:
            for item in extract_sales(h):
                ca_month += float(item.get("qty_sold", 0) or 0) * float(item.get("sale_price_ttc", 0) or 0)

        # CA même mois N-1 (année précédente)
        prev_month_start = datetime(cur_year - 1, cur_month, 1)
        prev_days_in_month = _cal.monthrange(cur_year - 1, cur_month)[1]
        prev_next = (prev_month_start + timedelta(days=prev_days_in_month)).replace(day=1)
        ca_n1_month = 0.0
        hist_n1 = db.query(StockHistory).filter(
            StockHistory.event_type.in_(["import_cashpad", "cashpad_sync"]),
            StockHistory.created_at >= prev_month_start,
            StockHistory.created_at <  prev_next,
        ).all()
        for h in hist_n1:
            for item in extract_sales(h):
                ca_n1_month += float(item.get("qty_sold", 0) or 0) * float(item.get("sale_price_ttc", 0) or 0)

        goal_amount_m = float(monthly_goals.get(cur_key, 0) or 0)
        days_elapsed = (today - month_start.date()).days + 1
        days_remaining = days_in_month - days_elapsed + 1
        pct = round((ca_month / goal_amount_m) * 100, 1) if goal_amount_m > 0 else None

        # Rythme nécessaire pour atteindre l'objectif (si défini)
        rythme = None
        if goal_amount_m > 0 and days_remaining > 0:
            rythme = round((goal_amount_m - ca_month) / days_remaining, 2)

        # Comparaison mois prorata (pour éviter mois incomplet vs mois complet N-1)
        n1_prorata = ca_n1_month * (days_elapsed / prev_days_in_month) if prev_days_in_month > 0 else 0
        delta_vs_n1 = ca_month - n1_prorata
        delta_pct_n1 = round((delta_vs_n1 / n1_prorata) * 100, 1) if n1_prorata > 0 else None

        months_fr = ["janvier","février","mars","avril","mai","juin","juillet","août","septembre","octobre","novembre","décembre"]
        monthly_goal = {
            "configured":       goal_amount_m > 0,
            "month_key":        cur_key,
            "month_label":      months_fr[cur_month - 1],
            "year":             cur_year,
            "goal_amount":      round(goal_amount_m, 2),
            "ca_month":         round(ca_month, 2),
            "ca_n1_month":      round(ca_n1_month, 2),
            "ca_n1_prorata":    round(n1_prorata, 2),
            "delta_vs_n1":      round(delta_vs_n1, 2),
            "delta_pct_n1":     delta_pct_n1,
            "pct":              pct,
            "days_elapsed":     days_elapsed,
            "days_remaining":   max(0, days_remaining),
            "total_days":       days_in_month,
            "rythme_daily":     rythme,
            "all_goals":        monthly_goals,   # pour l'édition
        }
    except Exception as e:
        print(f"[monthly goal] {e}")

    # ── Widget : contrôles de livraison en attente ───────────────────────
    dc_pending_count = (
        db.query(DeliveryCheck)
        .filter(DeliveryCheck.status == "counted")
        .count()
    )
    dc_pending_to_count = (
        db.query(DeliveryCheck)
        .filter(DeliveryCheck.status == "pending_count")
        .count()
    )
    dc_pending_list = (
        db.query(DeliveryCheck)
        .filter(DeliveryCheck.status == "counted")
        .order_by(DeliveryCheck.counted_at.desc())
        .limit(5)
        .all()
    )
    delivery_checks_info = {
        "pending_validation": dc_pending_count,
        "pending_count":      dc_pending_to_count,
        "items": [
            {
                "id":             c.id,
                "supplier_name":  c.supplier.name if c.supplier else "",
                "checked_by":     c.checked_by or "",
                "counted_at":     c.counted_at.isoformat() + "Z" if c.counted_at else None,
                "n_items":        len(c.items or []),
            }
            for c in dc_pending_list
        ],
    }

    return {
        "weather":         weather_summary,
        "events_upcoming": events_upcoming,
        "urgent_alerts":   urgent_alerts,
        "ca_yesterday":    round(ca_yesterday, 2),
        "ca_week":         round(ca_week, 2),
        "ca_n1":           round(ca_n1, 2),
        "ca_n1_date":      target_n1.strftime("%d/%m/%Y"),
        "ca_n1_day":       day_names[target_n1.weekday()],
        "top_products":    top_products,
        "last_sync":       last_sync_info,
        "pending_orders":  pending_orders,
        "ca_weekday":      ca_weekday,
        "monthly_goal":    monthly_goal,
        "delivery_checks": delivery_checks_info,
        "critical_alerts": _dashboard_critical_alerts(db),
        "stock_value_trend": _dashboard_stock_value_trend(db),
        "next_event_coverage": _dashboard_next_event_coverage(db),
        "dormant_products": _dashboard_dormant_products(db),
    }


def _dashboard_dormant_products(db: Session, days: int = 30) -> dict:
    """
    Produits en stock n'ayant été vendus dans aucun import Cashpad
    des `days` derniers jours. Retourne top 3 par valeur immobilisée.
    """
    since = datetime.utcnow() - timedelta(days=days)
    # Ensemble des noms produits vendus sur la période
    sold_names: set = set()
    for h in db.query(StockHistory).filter(
        StockHistory.event_type.in_(["import_cashpad", "cashpad_sync"]),
        StockHistory.created_at >= since,
    ).all():
        for item in extract_sales(h):
            name = (item.get("product_name") or item.get("name") or "").strip().lower()
            if name:
                sold_names.add(name)

    # Produits actifs avec stock > 0 + purchase_price défini, non vendus
    candidates = db.query(Product).filter(
        (Product.archived == False) | (Product.archived == None),
        Product.stock > 0,
        Product.purchase_price > 0,
    ).all()
    dormant = []
    for p in candidates:
        if p.name.strip().lower() not in sold_names:
            dormant.append({
                "id":        p.id,
                "name":      p.name,
                "stock":     round(float(p.stock or 0), 2),
                "unit":      p.unit or "u",
                "immobilized": round(float(p.stock or 0) * float(p.purchase_price or 0), 2),
            })
    dormant.sort(key=lambda x: -x["immobilized"])
    total_immob = round(sum(d["immobilized"] for d in dormant), 2)
    return {
        "days":         days,
        "count":        len(dormant),
        "total_immob":  total_immob,
        "top":          dormant[:3],
    }


def _dashboard_critical_alerts(db: Session) -> dict:
    """Comptes pour le widget 'alertes critiques'."""
    active_products = db.query(Product).filter(
        (Product.archived == False) | (Product.archived == None)
    )
    ruptures = active_products.filter(Product.stock <= 0).count()
    stock_bas = active_products.filter(
        Product.stock > 0, Product.stock <= Product.alert_threshold
    ).count()
    service_open = db.query(ServiceAlert).filter(ServiceAlert.status == "open").count()
    return {
        "ruptures":     ruptures,
        "stock_bas":    stock_bas,
        "service_open": service_open,
        "total":        ruptures + stock_bas + service_open,
    }


def _dashboard_stock_value_trend(db: Session) -> dict:
    """Valeur stock actuelle + tendance vs 7 jours (snapshots hebdo)."""
    current = 0.0
    for p in db.query(Product).filter(
        (Product.archived == False) | (Product.archived == None)
    ).all():
        if p.stock and p.purchase_price:
            current += float(p.stock) * float(p.purchase_price)

    # Valeur d'il y a ~7 jours depuis les StockSnapshot (label=weekly_auto)
    seven_days_ago = datetime.utcnow() - timedelta(days=7)
    # On prend le snapshot le plus récent avant J-7 pour chaque produit
    last_week_val = 0.0
    has_trend = False
    for p in db.query(Product).filter(
        (Product.archived == False) | (Product.archived == None)
    ).all():
        snap = (
            db.query(StockSnapshot)
            .filter(
                StockSnapshot.product_id == p.id,
                StockSnapshot.taken_at <= seven_days_ago,
            )
            .order_by(StockSnapshot.taken_at.desc())
            .first()
        )
        if snap and p.purchase_price:
            last_week_val += float(snap.stock) * float(p.purchase_price)
            has_trend = True

    if has_trend:
        delta = current - last_week_val
        delta_pct = round((delta / last_week_val) * 100, 1) if last_week_val > 0 else None
    else:
        delta = None
        delta_pct = None

    return {
        "current":   round(current, 2),
        "last_week": round(last_week_val, 2) if has_trend else None,
        "delta":     round(delta, 2) if delta is not None else None,
        "delta_pct": delta_pct,
    }


def _dashboard_next_event_coverage(db: Session) -> dict:
    """Prochain événement + nombre de produits requis non couverts."""
    today_dt = datetime.combine(datetime.now(_LOCAL_TZ).date(), datetime.min.time())
    ev = (
        db.query(Event)
        .filter(
            (Event.date >= today_dt) |
            ((Event.end_date != None) & (Event.end_date >= today_dt))
        )
        .order_by(Event.date.asc())
        .first()
    )
    if not ev:
        return None
    reqs = db.query(EventRequirement).filter(EventRequirement.event_id == ev.id).all()
    missing = []
    for r in reqs:
        p = db.query(Product).get(r.product_id)
        if not p:
            continue
        if (p.stock or 0) < (r.quantity or 0):
            missing.append({
                "product_name": p.name,
                "needed":       r.quantity,
                "in_stock":     p.stock or 0,
                "missing":      max(0, (r.quantity or 0) - (p.stock or 0)),
            })
    return {
        "event": {
            "id":         ev.id,
            "name":       ev.name,
            "event_type": ev.event_type,
            "date":       ev.date.isoformat(),
            "end_date":   ev.end_date.isoformat() if ev.end_date else None,
            "start_time": ev.start_time or "",
            "end_time":   ev.end_time or "",
        },
        "total_requirements": len(reqs),
        "missing_items":      missing,
        "covered":            len(reqs) > 0 and len(missing) == 0,
    }


# ── Endpoint pour configurer les objectifs mensuels ───────────────────────
class MonthlyGoalsIn(BaseModel):
    goals: dict    # {"2026-05": 18000, "2026-06": 22000, ...}


@app.get("/api/monthly-goals")
def get_monthly_goals(db: Session = Depends(get_db)):
    raw = _get_setting(db, "monthly_goals_json", "")
    try:
        goals = json.loads(raw) if raw else {}
        if not isinstance(goals, dict):
            goals = {}
    except Exception:
        goals = {}
    return {"goals": goals}


@app.post("/api/monthly-goals")
def set_monthly_goals(body: MonthlyGoalsIn, db: Session = Depends(get_db)):
    # Nettoyer : clés au format YYYY-MM, valeurs numériques positives
    clean = {}
    for k, v in (body.goals or {}).items():
        if not isinstance(k, str) or len(k) != 7 or k[4] != "-":
            continue
        try:
            amount = float(v)
            if amount > 0:
                clean[k] = amount
        except Exception:
            continue
    _set_setting(db, "monthly_goals_json", json.dumps(clean))
    db.commit()
    return {"ok": True, "goals": clean}
